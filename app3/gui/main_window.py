from __future__ import annotations

import calendar
import csv
import threading
import time
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from queue import Queue

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk  # Treeview + diálogos

import logging

from app3.config import metadata_dir
from app3.core.catalog import CatalogManager
from app3.core.classification_utils import (
    classify_transaction,
    filter_records_by_tab,
    get_classification_label,
    get_tab_statistics,
)
from app3.core.classifier import ClassificationDB, build_dest_folder, classify_record
from app3.core.factura_index import FacturaIndexer
from app3.core.models import FacturaRecord
from app3.core.session import ClientSession
from app3.gui.loading_modal import LoadingOverlay
from app3.gui.pdf_viewer import PDFViewer
from app3.gui.session_view import SessionView

logger = logging.getLogger(__name__)

# ── PALETA (misma que session_view) ──────────────────────────────────────────
BG      = "#0d0f14"
SURFACE = "#13161e"
CARD    = "#181c26"
BORDER  = "#252a38"
TEAL    = "#2dd4bf"
TEAL_DIM= "#1a9e8f"
TEXT    = "#e8eaf0"
MUTED   = "#6b7280"
DANGER  = "#f87171"
SUCCESS = "#34d399"
WARNING = "#fbbf24"

# ── FUENTES (lazy -- se crean solo despues de que existe la ventana raiz) ─────
_fonts: dict = {}

def _f(key: str, size: int, weight: str = "normal") -> ctk.CTkFont:
    if key not in _fonts:
        _fonts[key] = ctk.CTkFont(family="Segoe UI", size=size, weight=weight)
    return _fonts[key]

def F_TITLE()  -> ctk.CTkFont: return _f("mw_title",  11, "bold")
def F_LABEL()  -> ctk.CTkFont: return _f("mw_label",  12)
def F_SMALL()  -> ctk.CTkFont: return _f("mw_small",  11)
def F_BTN()    -> ctk.CTkFont: return _f("mw_btn",    13, "bold")

ESTADO_ICON = {
    "clasificado":   "✓",
    "pendiente":     "·",
    "pendiente_pdf": "!",
    "sin_xml":       "--",
}

# Texto corto para la columna Estado
ESTADO_LABEL = {
    "clasificado":   "clasificado",
    "pendiente":     "pendiente",
    "pendiente_pdf": "sin PDF",
    "sin_xml":       "sin XML",
}

def _fmt_amount(value: str) -> str:
    """Formatea montos como App 2: 137 131,77 (miles con espacio, decimales con coma)."""
    try:
        v = str(value).strip().replace(",", ".")
        f = float(v)
        # Separador de miles = espacio, decimal = coma
        parts = f"{abs(f):,.2f}".split(".")
        integer_part = parts[0].replace(",", " ")
        result = f"{integer_part},{parts[1]}"
        return f"-{result}" if f < 0 else result
    except (ValueError, TypeError):
        return str(value) if value else "--"

def _short_name(name: str, max_len: int = 34) -> str:
    """Abrevia razones sociales como App 2."""
    base = str(name or "").strip().upper()
    for long_form, short_form in [
        ("SOCIEDAD ANONIMA", "S.A."),
        ("SOCIEDAD ANÓNIMA", "S.A."),
        ("SOCIEDAD DE RESPONSABILIDAD LIMITADA", "S.R.L."),
        ("SOCIEDAD RESPONSABILIDAD LIMITADA", "S.R.L."),
        ("COMPANIA LIMITADA", "LTDA."),
        ("COMPAÑIA LIMITADA", "LTDA."),
        ("LIMITADA", "LTDA."),
    ]:
        base = base.replace(long_form, short_form)
    base = " ".join(base.split())
    return base[:max_len]

# Inyectar estilos oscuros en el Treeview de ttk
_TREE_STYLE_DONE = False

def _apply_tree_style():
    global _TREE_STYLE_DONE
    if _TREE_STYLE_DONE:
        return
    style = ttk.Style()
    style.theme_use("default")
    style.configure("Dark.Treeview",
        background=CARD, foreground=TEXT,
        fieldbackground=CARD, borderwidth=0,
        rowheight=19, font=("Segoe UI", 9),
    )
    style.configure("Dark.Treeview.Heading",
        background=SURFACE, foreground=MUTED,
        borderwidth=0, font=("Segoe UI", 8, "bold"),
    )
    style.map("Dark.Treeview",
        background=[("selected", "#1a3a36")],
        foreground=[("selected", TEAL)],
    )
    style.configure("Dark.Vertical.TScrollbar",
        background=SURFACE, troughcolor=BG,
        borderwidth=0, arrowsize=12,
    )
    _TREE_STYLE_DONE = True


class DatePickerPopup(ctk.CTkToplevel):
    def __init__(self, parent, on_pick, initial_value: str = ""):
        super().__init__(parent)
        self.title("Seleccionar fecha")
        self.configure(fg_color=CARD)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self._on_pick = on_pick

        dt = self._parse_date(initial_value) or datetime.today()
        self._year = dt.year
        self._month = dt.month

        self._body = ctk.CTkFrame(self, fg_color=CARD)
        self._body.pack(padx=8, pady=8, fill="both", expand=True)
        self._draw()

    @staticmethod
    def _parse_date(text: str):
        raw = (text or "").strip()
        if not raw:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
        return None

    def _draw(self):
        for w in self._body.winfo_children():
            w.destroy()

        top = ctk.CTkFrame(self._body, fg_color="transparent")
        top.pack(fill="x", padx=4, pady=(2, 8))
        ctk.CTkButton(top, text="◀", width=32, height=28, fg_color=SURFACE,
                      hover_color=BORDER, command=self._prev_month).pack(side="left")
        ctk.CTkLabel(top, text=f"{calendar.month_name[self._month]} {self._year}",
                      text_color=TEXT, font=F_LABEL()).pack(side="left", expand=True)
        ctk.CTkButton(top, text="▶", width=32, height=28, fg_color=SURFACE,
                      hover_color=BORDER, command=self._next_month).pack(side="right")

        grid = ctk.CTkFrame(self._body, fg_color="transparent")
        grid.pack(padx=4, pady=(0, 8))
        for i, name in enumerate(["Lu", "Ma", "Mi", "Ju", "Vi", "Sá", "Do"]):
            ctk.CTkLabel(grid, text=name, text_color=MUTED, font=F_SMALL()).grid(row=0, column=i, padx=3, pady=2)

        for r, week in enumerate(calendar.monthcalendar(self._year, self._month), start=1):
            for c, day in enumerate(week):
                if day == 0:
                    ctk.CTkLabel(grid, text=" ", width=30).grid(row=r, column=c, padx=2, pady=2)
                    continue
                ctk.CTkButton(
                    grid,
                    text=str(day),
                    width=30,
                    height=28,
                    fg_color=SURFACE,
                    hover_color=TEAL_DIM,
                    text_color=TEXT,
                    command=lambda dd=day: self._pick_day(dd),
                ).grid(row=r, column=c, padx=2, pady=2)

        bottom = ctk.CTkFrame(self._body, fg_color="transparent")
        bottom.pack(fill="x", padx=4)
        ctk.CTkButton(bottom, text="Hoy", width=70, fg_color=SURFACE, hover_color=BORDER,
                      command=self._pick_today).pack(side="left")
        ctk.CTkButton(bottom, text="Limpiar", width=80, fg_color=SURFACE, hover_color=BORDER,
                      command=lambda: self._emit("")).pack(side="right")

    def _prev_month(self):
        self._month -= 1
        if self._month == 0:
            self._month = 12
            self._year -= 1
        self._draw()

    def _next_month(self):
        self._month += 1
        if self._month == 13:
            self._month = 1
            self._year += 1
        self._draw()

    def _pick_day(self, day: int):
        self._emit(f"{day:02d}/{self._month:02d}/{self._year:04d}")

    def _pick_today(self):
        now = datetime.today()
        self._emit(f"{now.day:02d}/{now.month:02d}/{now.year:04d}")

    def _emit(self, value: str):
        self._on_pick(value)
        self.destroy()


def _parse_date_any(text: str):
    raw = (text or "").strip()
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


class DatePickerDropdown(ctk.CTkToplevel):
    """Calendario anclado tipo dropdown (sin barra de título)."""

    def __init__(self, parent, on_pick, initial_value: str = "", x: int = 120, y: int = 120):
        super().__init__(parent)
        self.overrideredirect(True)
        # CTkToplevel no permite transparencia en fg_color.
        self.configure(fg_color=CARD)
        self.attributes("-topmost", True)
        self._on_pick = on_pick
        dt = _parse_date_any(initial_value) or datetime.today()
        self._year = dt.year
        self._month = dt.month
        self.geometry(f"290x270+{x}+{y}")

        self._card = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10, border_width=1, border_color=BORDER)
        self._card.pack(fill="both", expand=True)

        self.bind("<Escape>", lambda _e: self._safe_close())
        self.bind("<FocusOut>", lambda _e: self._safe_close())

        self._draw()
        self.after(10, self.focus_force)

    def _safe_close(self):
        try:
            self.destroy()
        except Exception:
            pass

    def _draw(self):
        for w in self._card.winfo_children():
            w.destroy()

        top = ctk.CTkFrame(self._card, fg_color="transparent")
        top.pack(fill="x", padx=8, pady=(8, 6))
        ctk.CTkButton(top, text="◀", width=30, height=28, fg_color=SURFACE,
                      hover_color=BORDER, command=self._prev_month).pack(side="left")
        ctk.CTkLabel(top, text=f"{calendar.month_name[self._month]} {self._year}",
                     text_color=TEXT, font=F_LABEL()).pack(side="left", expand=True)
        ctk.CTkButton(top, text="▶", width=30, height=28, fg_color=SURFACE,
                      hover_color=BORDER, command=self._next_month).pack(side="right")

        grid = ctk.CTkFrame(self._card, fg_color="transparent")
        grid.pack(padx=8, pady=(0, 8))
        for i, name in enumerate(["Lu", "Ma", "Mi", "Ju", "Vi", "Sá", "Do"]):
            ctk.CTkLabel(grid, text=name, text_color=MUTED, font=F_SMALL()).grid(row=0, column=i, padx=3, pady=2)

        for r, week in enumerate(calendar.monthcalendar(self._year, self._month), start=1):
            for c, day in enumerate(week):
                if day == 0:
                    ctk.CTkLabel(grid, text=" ", width=30).grid(row=r, column=c, padx=2, pady=2)
                    continue
                ctk.CTkButton(
                    grid,
                    text=str(day),
                    width=30,
                    height=28,
                    fg_color=SURFACE,
                    hover_color=TEAL_DIM,
                    text_color=TEXT,
                    command=lambda dd=day: self._pick_day(dd),
                ).grid(row=r, column=c, padx=2, pady=2)

        bottom = ctk.CTkFrame(self._card, fg_color="transparent")
        bottom.pack(fill="x", padx=8, pady=(0, 8))
        ctk.CTkButton(bottom, text="Hoy", width=70, fg_color=SURFACE, hover_color=BORDER,
                      command=self._pick_today).pack(side="left")
        ctk.CTkButton(bottom, text="Limpiar", width=80, fg_color=SURFACE, hover_color=BORDER,
                      command=lambda: self._emit("")).pack(side="right")

    def _prev_month(self):
        self._month -= 1
        if self._month == 0:
            self._month = 12
            self._year -= 1
        self._draw()

    def _next_month(self):
        self._month += 1
        if self._month == 13:
            self._month = 1
            self._year += 1
        self._draw()

    def _pick_day(self, day: int):
        self._emit(f"{day:02d}/{self._month:02d}/{self._year:04d}")

    def _pick_today(self):
        now = datetime.today()
        self._emit(f"{now.day:02d}/{now.month:02d}/{now.year:04d}")

    def _emit(self, value: str):
        self._on_pick(value)
        self._safe_close()


class NewCuentaDialog(ctk.CTkToplevel):
    """Modal dialog for adding a new account."""

    def __init__(self, parent, categoria: str, subtipo: str, existing_cuentas: list,
                 catalog_mgr, on_success):
        super().__init__(parent)
        self.overrideredirect(True)
        self.configure(fg_color=CARD)
        self.attributes("-topmost", True)
        self._on_success = on_success
        self._categoria = categoria
        self._subtipo = subtipo
        self._existing_cuentas = existing_cuentas
        self._catalog_mgr = catalog_mgr
        self.geometry("350x180+400+300")

        self._card = ctk.CTkFrame(self, fg_color=CARD, corner_radius=10,
                                   border_width=1, border_color=BORDER)
        self._card.pack(fill="both", expand=True, padx=10, pady=10)

        self.bind("<Escape>", lambda _e: self._cancel())
        self.bind("<Return>", lambda _e: self._save())

        # Title
        ctk.CTkLabel(
            self._card, text="Agregar Nueva Cuenta",
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            text_color=TEXT,
        ).pack(fill="x", padx=12, pady=(12, 8))

        # Input field
        self._input_var = ctk.StringVar()
        ctk.CTkLabel(
            self._card, text="Nombre de la cuenta",
            font=F_SMALL(), text_color=MUTED,
        ).pack(fill="x", padx=12, pady=(0, 4))

        self._entry = ctk.CTkEntry(
            self._card, textvariable=self._input_var,
            fg_color=SURFACE, border_color=BORDER, text_color=TEXT,
            font=F_LABEL(), height=32, corner_radius=6,
        )
        self._entry.pack(fill="x", padx=12, pady=(0, 12))
        self._entry.focus()

        # Error message label
        self._error_label = ctk.CTkLabel(
            self._card, text="", text_color=DANGER,
            font=F_SMALL(), wraplength=300,
        )
        self._error_label.pack(fill="x", padx=12, pady=(0, 8))

        # Button row
        btn_frame = ctk.CTkFrame(self._card, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))

        ctk.CTkButton(
            btn_frame, text="Cancelar", width=100,
            fg_color=SURFACE, hover_color=BORDER,
            font=F_SMALL(), text_color=TEXT,
            command=self._cancel,
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_frame, text="Guardar", width=100,
            fg_color=TEAL, hover_color=TEAL_DIM,
            font=F_SMALL(), text_color="#0d1a18",
            command=self._save,
        ).pack(side="right")

        self.after(10, self.focus_force)

    def _show_error(self, msg: str):
        self._error_label.configure(text=msg)

    def _validate_and_save(self) -> str | None:
        """Validate input and return new account name or None if invalid."""
        name = self._input_var.get().strip()

        if not name:
            self._show_error("El nombre no puede estar vacío.")
            return None

        if len(name) > 100:
            self._show_error("El nombre no puede exceder 100 caracteres.")
            return None

        if name.upper() in [c.upper() for c in self._existing_cuentas]:
            self._show_error(f"La cuenta '{name}' ya existe.")
            return None

        return name

    def _save(self):
        name = self._validate_and_save()
        if name:
            try:
                self._catalog_mgr.add_cuenta(self._categoria, self._subtipo, name)
                self._on_success(name)
                self._safe_close()
            except Exception as e:
                self._show_error(f"Error al guardar: {str(e)}")

    def _cancel(self):
        self._safe_close()

    def _safe_close(self):
        try:
            self.destroy()
        except Exception:
            pass


def _format_amount_es(number: Decimal) -> str:
    sign = "-" if number < 0 else ""
    n = abs(number)
    text = f"{n:,.2f}"
    text = text.replace(",", "_").replace(".", ",").replace("_", " ")
    return f"{sign}{text}"


def _month_name_es(dt: datetime) -> str:
    months = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
    }
    return months.get(dt.month, "MES")


def _default_export_filename(client_name: str, from_date: str, to_date: str) -> str:
    base_dt = _parse_date_any(from_date) or _parse_date_any(to_date) or datetime.now()
    year = base_dt.strftime("%Y")
    month_txt = _month_name_es(base_dt)
    client_clean = (str(client_name or "REPORTE")
                    .replace("/", " ")
                    .replace("\\", " ")
                    .strip())
    if len(client_clean) > 42:
        client_clean = client_clean[:42].strip()
    return f"PF-{year} - {client_clean} - REPORTE - {month_txt}.xlsx"


def _safe_excel_sheet_name(raw_name: str, used_names: set[str]) -> str:
    """Sanitiza y hace único un nombre de hoja de Excel (máx. 31 chars)."""
    invalid_chars = {"\\", "/", "*", "?", ":", "[", "]"}
    cleaned = "".join("_" if ch in invalid_chars else ch for ch in str(raw_name or "").strip())
    cleaned = cleaned.strip("'")
    base = (cleaned or "SIN CLASIFICAR")[:31]

    candidate = base
    suffix = 1
    while candidate in used_names:
        suffix_txt = f" ({suffix})"
        allowed = 31 - len(suffix_txt)
        candidate = f"{base[:allowed]}{suffix_txt}"
        suffix += 1

    used_names.add(candidate)
    return candidate


_GASTO_PREFIX = {
    "GASTOS GENERALES": "GG",
    "GASTOS ESPECÍFICOS": "GE",
    "GASTOS ESPECIFICOS": "GE",
}


def _write_gasto_grouped(
    ws, sheet_df, display_cols,
    numeric_columns, text_columns, date_column,
    pretty_headers, owner_name, sheet_name, date_from_label, date_to_label,
    title_fill, subtitle_fill, summary_fill, header_fill, credit_fill,
    title_font, subtitle_font, summary_font, header_font,
):
    """Hoja Gasto -- agrupación por (subtipo, nombre_cuenta).

    Layout por grupo:
        [filas de datos -- sin color]
        [fila subtotal: sumas numéricas + label "GG/GE / NOMBRE" en última col]  ← fill azul
        [fila vacía]

    display_cols: columnas a mostrar.  subtipo/nombre_cuenta se leen del DataFrame
    para agrupar aunque no aparezcan en display_cols.
    """
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    n_cols = len(display_cols)

    # ── Filas 1-3: título / subtítulo / resumen ───────────────────────────────
    for row in (1, 2, 3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)

    ws.cell(row=1, column=1).value = str(owner_name).upper()
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).fill = title_fill

    ws.cell(row=2, column=1).value = (
        f"REPORTE DE {sheet_name.upper()} - Período: {date_from_label} al {date_to_label}"
    )
    ws.cell(row=2, column=1).font = subtitle_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).fill = subtitle_fill

    monto_total = Decimal("0")
    if "total_comprobante" in sheet_df.columns:
        for v in sheet_df["total_comprobante"].dropna():
            try:
                monto_total += Decimal(str(v))
            except Exception:
                pass

    monedas = (
        sorted({str(m).strip() for m in sheet_df["moneda"].dropna() if str(m).strip()})
        if "moneda" in sheet_df.columns else []
    )
    moneda_value = (
        "N/A" if not monedas
        else monedas[0] if len(monedas) == 1
        else "MIXTA: " + ", ".join(monedas)
    )

    ws.cell(row=3, column=1).value = (
        f"Total filas: {len(sheet_df)}   |   Monto Total: {_format_amount_es(monto_total)}   |   "
        f"Moneda: {moneda_value}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws.cell(row=3, column=1).font = summary_font
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=1).fill = summary_fill

    # ── Fila 5: encabezados de columna ────────────────────────────────────────
    for col_idx, col_name in enumerate(display_cols, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = pretty_headers.get(col_name, col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tipo_col_idx   = (display_cols.index("tipo_documento") + 1) if "tipo_documento" in display_cols else None
    # CRÍTICO: numeric_display solo contiene columnas que existen en display_cols
    numeric_display = [c for c in display_cols if c in numeric_columns and c in display_cols]

    subtotal_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    subtotal_font = Font(bold=True)

    # ── Ordenar y agrupar (claves de agrupación vienen del df, no de display) ─
    group_cols = [c for c in ("subtipo", "nombre_cuenta") if c in sheet_df.columns]
    sort_cols  = group_cols + (["fecha_emision"] if "fecha_emision" in sheet_df.columns else [])
    sorted_df  = sheet_df.sort_values(sort_cols) if sort_cols else sheet_df

    def _safe(v):
        try:
            return None if pd.isna(v) else v
        except (TypeError, ValueError):
            return v

    current_row = 6

    if group_cols:
        for group_keys, group_df in sorted_df.groupby(group_cols, sort=False):
            if isinstance(group_keys, tuple) and len(group_keys) == 2:
                subtipo_val = str(group_keys[0]).strip().upper()
                cuenta_val  = str(group_keys[1]).strip()
            else:
                subtipo_val = ""
                cuenta_val  = str(group_keys).strip()

            # CRÍTICO: group_sums solo para columnas numéricas que existen en display_cols
            group_sums: dict[str, Decimal] = {c: Decimal("0") for c in numeric_display}

            # Filas de datos (sin color)
            for _, row_data in group_df.iterrows():
                for col_idx, col_name in enumerate(display_cols, start=1):
                    # CRÍTICO: Verificar que col_name existe en row_data
                    val = _safe(row_data.get(col_name)) if col_name in row_data.index else None
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = val
                    if col_name in text_columns:
                        cell.number_format = "@"
                        cell.value = "" if cell.value is None else str(cell.value)
                    elif col_name == date_column and cell.value is not None:
                        cell.number_format = "dd/mm/yyyy"
                    elif col_name in numeric_columns and cell.value is not None:
                        cell.number_format = "#,##0.00"
                        if isinstance(cell.value, Decimal):
                            cell.value = float(cell.value)

                if tipo_col_idx and ws.cell(row=current_row, column=tipo_col_idx).value == "Nota de Crédito":
                    for c in range(1, n_cols + 1):
                        ws.cell(row=current_row, column=c).fill = credit_fill

                for col_name in numeric_display:
                    tv = _safe(row_data.get(col_name)) if col_name in row_data.index else None
                    try:
                        if tv is not None:
                            group_sums[col_name] += Decimal(str(tv))
                    except Exception:
                        pass

                current_row += 1

            # ── Fila de subtotal: sumas + label en la ÚLTIMA columna ──────────
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=current_row, column=col_idx).fill = subtotal_fill

            for col_name in numeric_display:
                ci = display_cols.index(col_name) + 1
                tc = ws.cell(row=current_row, column=ci)
                tc.value         = float(group_sums[col_name])
                tc.number_format = "#,##0.00"
                tc.font          = subtotal_font

            prefix      = _GASTO_PREFIX.get(subtipo_val, "")
            cuenta_label = cuenta_val.upper() if cuenta_val else subtipo_val
            label        = f"{prefix} / {cuenta_label}" if prefix else cuenta_label
            lbl = ws.cell(row=current_row, column=n_cols)
            lbl.value     = label
            lbl.font      = subtotal_font
            lbl.alignment = Alignment(horizontal="right", vertical="center")

            current_row += 2  # subtotal + fila vacía

    else:
        # Sin agrupación: filas directas
        for _, row_data in sorted_df.iterrows():
            for col_idx, col_name in enumerate(display_cols, start=1):
                val  = _safe(row_data[col_name])
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                if col_name in text_columns:
                    cell.number_format = "@"
                    cell.value = "" if cell.value is None else str(cell.value)
                elif col_name == date_column and cell.value is not None:
                    cell.number_format = "dd/mm/yyyy"
                elif col_name in numeric_columns and cell.value is not None:
                    cell.number_format = "#,##0.00"
            current_row += 1

    # ── Ancho de columnas ─────────────────────────────────────────────────────
    for col_idx in range(1, n_cols + 1):
        max_len = 0
        for row_idx in range(5, current_row):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[
            ws.cell(row=5, column=col_idx).column_letter
        ].width = min(max(max_len + 3, 12), 65)

    ws.freeze_panes = ws["A6"]


class App3Window(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        self.title("App 3 -- Clasificador Contable")
        self.geometry("1440x860")
        self.minsize(1100, 680)
        self.configure(fg_color=BG)
        self.grid_rowconfigure(0, weight=0)  # Header (can expand with tabs)
        self.grid_rowconfigure(1, weight=1)  # Body
        self.grid_columnconfigure(0, weight=1)

        self.session: ClientSession | None = None
        self.db: ClassificationDB | None = None
        self.catalog_mgr: CatalogManager | None = None
        self.records: list[FacturaRecord] = []
        self.all_records: list[FacturaRecord] = []
        self._db_records: dict[str, dict] = {}
        self.selected: FacturaRecord | None = None
        self.selected_records: list[FacturaRecord] = []  # Multi-selección
        self._load_queue: Queue = Queue()
        self._active_calendar: DatePickerDropdown | None = None
        self._load_generation: int = 0
        self._all_cuentas: list[str] = []  # Unfiltered account list
        self._loading_overlay: LoadingOverlay | None = None  # Overlay de carga
        self._tree_clave_map: dict[str, FacturaRecord] = {}  # Mapeo: clave -> record (para mantener orden)
        self._active_tab: str = "todas"  # Pestaña activa de facturas del período
        self._tab_buttons: dict[str, ctk.CTkButton] = {}  # Botones de pestañas

        _apply_tree_style()

        # Crear body_container (para header y body) -- inicialmente oculto
        self._body_container = ctk.CTkFrame(self, fg_color=BG)
        self._body_container.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self._body_container.grid_rowconfigure(0, weight=0)  # Header
        self._body_container.grid_rowconfigure(1, weight=1)  # Body
        self._body_container.grid_columnconfigure(0, weight=1)
        self._build(self._body_container)
        self._body_container.grid_remove()  # Ocultar inicialmente

        # Crear SessionView (visible al inicio) -- llena toda la ventana
        self._session_frame = SessionView(self, on_session_resolved=self._on_session_resolved)
        self._session_frame.grid(row=0, column=0, rowspan=2, sticky="nsew")

    # ── SESIÓN ────────────────────────────────────────────────────────────────
    def _open_session_view(self):
        # Mostrar SessionView y ocultar body_container
        self._body_container.grid_remove()
        self._session_frame.grid()
        self._session_frame.focus_force()

    def _on_session_resolved(self, session: ClientSession):
        # Ocultar SessionView y mostrar body_container
        self._session_frame.grid_remove()
        self._body_container.grid()
        self.focus_force()
        self._load_session(session)

    def _clear_cache_and_reload(self):
        """Limpia el cache de PDFs y recarga todos los datos.

        Útil cuando los XMLs quedan desvinculados de sus PDFs tras clasificación.
        """
        if not self.session:
            return

        if not self._ask("Limpiar cache",
                        "Se eliminará el cache de PDFs y se recalcularán todas las vinculaciones.\n\n"
                        "Esto toma ~40-50 segundos. ¿Deseas continuar?"):
            return

        try:
            # Limpiar cache de PDFs
            mdir = metadata_dir(self.session.folder)
            cache_file = mdir / "pdf_cache.json"
            if cache_file.exists():
                cache_file.unlink()
                logger.info(f"Cache limpiado: {cache_file}")

            # Recalcular registros (fuerza rescan de PDFs)
            self._load_session(self.session)
        except Exception as e:
            self._show_error("Error al limpiar cache", str(e))

    def _load_session(self, session: ClientSession):
        import time
        self._load_generation += 1
        generation = self._load_generation
        self._load_queue = Queue()

        # Mostrar overlay de carga integrado
        if not hasattr(self, '_loading_overlay') or not self._loading_overlay:
            self._loading_overlay = LoadingOverlay(self)
            self._loading_overlay.grid(row=0, column=0, rowspan=2, sticky="nsew")
        else:
            self._loading_overlay.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self._loading_overlay.update_status("Iniciando...")

        # 🔥 CRÍTICO: Forzar actualización visual del overlay
        # Sin esto, el overlay no se ve en pantalla durante el worker thread
        self.update_idletasks()
        self.update()

        def worker():
            try:
                start_total = time.perf_counter()

                # Fase 1: Setup
                mdir = metadata_dir(session.folder)
                self.after(0, lambda: self._loading_overlay.update_status("📂 Preparando cliente..."))
                self.after(0, lambda: self._loading_overlay.update_progress(10, 100))
                catalog = CatalogManager(mdir).load()
                db = ClassificationDB(mdir)
                indexer = FacturaIndexer()

                # Fase 2: Load (XML + PDF)
                self.after(0, lambda: self._loading_overlay.update_status("📄 Leyendo XMLs..."))
                self.after(0, lambda: self._loading_overlay.update_progress(20, 100))

                # La fase larga: XML + PDF
                self.after(0, lambda: self._loading_overlay.update_status("🔍 Escaneando PDFs (esto toma ~40s)..."))
                self.after(0, lambda: self._loading_overlay.update_progress(30, 100))

                start_load = time.perf_counter()
                records = indexer.load_period(
                    session.folder,
                    from_date="",
                    to_date="",
                    include_pdf_scan=True,
                    allow_pdf_content_fallback=True,
                )
                load_time = time.perf_counter() - start_load
                logger.info(f"load_period() tardó {load_time:.2f}s para {len(records)} registros")

                # Escanear PDFs huérfanos
                from app3.core.classification_utils import find_orphaned_pdfs, create_orphaned_record
                pf_root = session.folder.parent.parent
                contabilidades_root = pf_root / "Contabilidades"
                local_db_records = db.get_records_map()
                orphaned_list = find_orphaned_pdfs(contabilidades_root, local_db_records)

                # Agregar registros dummy para PDFs huérfanos
                for orphaned_info in orphaned_list:
                    orphaned_record = create_orphaned_record(orphaned_info)
                    records.append(orphaned_record)

                if orphaned_list:
                    logger.info(f"Agregados {len(orphaned_list)} registros huérfanos")

                # Casi listo
                self.after(0, lambda: self._loading_overlay.update_status("✅ Finalizando..."))
                self.after(0, lambda: self._loading_overlay.update_progress(90, 100))

                total_time = time.perf_counter() - start_total
                logger.info(f"Worker total: {total_time:.2f}s")

                self._load_queue.put(("ok", (generation, session, catalog, db, records, indexer.parse_errors)))
            except Exception as exc:
                logger.exception("Error en worker")
                self._load_queue.put(("error", (generation, str(exc))))

        threading.Thread(target=worker, daemon=True).start()
        self.after(150, self._poll_load)

    def _poll_load(self):
        import time
        if self._load_queue.empty():
            self.after(150, self._poll_load)
            return

        status, payload = self._load_queue.get()

        # Ocultar overlay de carga
        if hasattr(self, '_loading_overlay') and self._loading_overlay:
            self._loading_overlay.grid_remove()

        if status == "error":
            generation, message = payload
            if generation != self._load_generation:
                return
            self._set_status("Error al cargar")
            self._show_error("Error al cargar cliente", message)
            return

        generation, session, catalog_mgr, db, records, parse_errors = payload
        if generation != self._load_generation:
            return
        self.session = session
        self.catalog_mgr = catalog_mgr
        self.db = db
        self.all_records = records
        self._db_records = db.get_records_map()
        self.records = self._apply_filters()
        self.selected = None
        self.selected_records = []

        # Actualizar header
        self._lbl_cliente.configure(text=session.folder.name)
        self._lbl_year.configure(text=f"PF-{session.year}")

        # Actualizar catálogo
        cats = catalog_mgr.categorias()
        self._cat_cb.configure(values=cats)
        if cats:
            self._cat_var.set(cats[0])
            self._on_categoria_change()

        self.pdf_viewer.clear()

        # Actualizar overlay: renderizando tabla
        if hasattr(self, '_loading_overlay') and self._loading_overlay and self._loading_overlay.winfo_exists():
            self._loading_overlay.update_status("Renderizando lista de facturas...")
            self._loading_overlay.update_progress(0, len(self.records))

        # Timing del refresh
        start_refresh = time.perf_counter()
        self._refresh_tree()
        refresh_time = time.perf_counter() - start_refresh
        logger.info(f"_refresh_tree() tardó {refresh_time:.2f}s para {len(self.records)} registros")

        self._update_progress()
        self._set_status("Listo")

        if parse_errors:
            self._show_warning(
                "Advertencias al cargar",
                f"Facturas cargadas: {len(records)}\n"
                f"XML con error (omitidos): {len(parse_errors)}\n\n"
                + "\n".join(parse_errors[:5])
            )

    def _apply_pdf_enrichment(self, generation: int, enriched_records: list[FacturaRecord], parse_errors: list[str]):
        """Ya no se usa (PDFs cargados en _load_session). Mantenido por compatibilidad."""
        if generation != self._load_generation:
            return

        self.all_records = enriched_records
        self.records = self._apply_date_filter(self.all_records)
        self._refresh_tree()
        self._update_progress()
        self._set_status("Listo")

        if parse_errors:
            self._show_warning(
                "Advertencias en PDFs",
                f"Se detectaron incidencias en {len(parse_errors)} PDF(s).\n\n" + "\n".join(parse_errors[:5]),
            )

    # ── CONSTRUCCIÓN UI ───────────────────────────────────────────────────────
    def _build(self, parent=None):
        if parent is None:
            parent = self
        self._build_header(parent)
        self._build_body(parent)

    def _build_header(self, parent=None):
        if parent is None:
            parent = self
        hdr = ctk.CTkFrame(parent, fg_color=SURFACE, corner_radius=0)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_columnconfigure(1, weight=1)
        hdr.grid_columnconfigure(2, weight=1)

        # Logo
        ctk.CTkLabel(hdr, text="📊", fg_color="#1a3a36", corner_radius=8,
                      width=32, height=32,
                      font=ctk.CTkFont(size=16)).grid(row=0, column=0, padx=(16,8), pady=14)
        ctk.CTkLabel(hdr, text="Clasificador  Contable",
                      font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
                      text_color=TEXT).grid(row=0, column=1, sticky="w", padx=(0, 12))

        # Info cliente activo (bloque izquierdo)
        client_frame = ctk.CTkFrame(hdr, fg_color="transparent")
        client_frame.grid(row=0, column=2, sticky="w")

        ctk.CTkLabel(client_frame, text="Cliente:",
                      font=F_SMALL(), text_color=MUTED).pack(side="left", padx=(0,6))
        self._lbl_cliente = ctk.CTkLabel(client_frame, text="Sin sesión",
                                          font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
                                          text_color=TEXT)
        self._lbl_cliente.pack(side="left")
        self._lbl_year = ctk.CTkLabel(client_frame, text="",
                                       font=F_SMALL(), text_color=MUTED,
                                       fg_color=CARD, corner_radius=20)
        self._lbl_year.pack(side="left", padx=(10,0), ipadx=8, ipady=2)

        # Filtros de fecha (bloque central)
        date_frame = ctk.CTkFrame(hdr, fg_color="transparent")
        date_frame.grid(row=0, column=3, sticky="e", padx=8)

        ctk.CTkLabel(date_frame, text="Desde:", font=F_SMALL(),
                      text_color=MUTED).pack(side="left", padx=(0,4))
        self.from_var = ctk.StringVar()
        self._from_entry = ctk.CTkEntry(
            date_frame,
            textvariable=self.from_var,
            width=100,
            placeholder_text="DD/MM/AAAA",
            fg_color=CARD,
            border_color=BORDER,
            text_color=TEXT,
            font=F_SMALL(),
            height=32,
            corner_radius=8,
        )
        self._from_entry.pack(side="left")
        ctk.CTkButton(date_frame, text="📅", width=30, height=32,
                      fg_color=SURFACE, hover_color=BORDER, command=lambda: self._open_date_picker("from")).pack(side="left", padx=(4, 8))

        ctk.CTkLabel(date_frame, text="Hasta:", font=F_SMALL(),
                      text_color=MUTED).pack(side="left", padx=(10,4))
        self.to_var = ctk.StringVar()
        self._to_entry = ctk.CTkEntry(
            date_frame,
            textvariable=self.to_var,
            width=100,
            placeholder_text="DD/MM/AAAA",
            fg_color=CARD,
            border_color=BORDER,
            text_color=TEXT,
            font=F_SMALL(),
            height=32,
            corner_radius=8,
        )
        self._to_entry.pack(side="left")
        ctk.CTkButton(date_frame, text="📅", width=30, height=32,
                      fg_color=SURFACE, hover_color=BORDER, command=lambda: self._open_date_picker("to")).pack(side="left", padx=(4, 8))

        ctk.CTkButton(date_frame, text="Filtrar", width=70, height=32,
                       fg_color=TEAL, hover_color=TEAL_DIM, text_color="#0d1a18",
                       font=F_SMALL(), corner_radius=8,
                       command=self._on_filter).pack(side="left", padx=(8,0))

        # Botón limpiar cache y recargar
        ctk.CTkButton(date_frame, text="🔄  Limpiar cache", width=130, height=32,
                       fg_color=SURFACE, hover_color=BORDER, text_color=TEXT,
                       border_color=WARNING, border_width=1,
                       font=F_SMALL(), corner_radius=8,
                       command=self._clear_cache_and_reload).pack(side="left", padx=(8,0))

        # Botón cambiar cliente
        ctk.CTkButton(hdr, text="⇄  Cambiar cliente", width=150, height=32,
                       fg_color=CARD, hover_color=SURFACE, text_color=MUTED,
                       border_color=BORDER, border_width=1,
                       font=F_SMALL(), corner_radius=8,
                       command=self._open_session_view).grid(
            row=0, column=4, padx=(10, 8), pady=14)

        # Status
        self._status_var = ctk.StringVar(value="")
        ctk.CTkLabel(hdr, textvariable=self._status_var,
                      font=F_SMALL(), text_color=MUTED).grid(
            row=0, column=5, padx=(0, 16))

        # Pestañas de clasificación (fila 1)
        tabs_container = ctk.CTkFrame(hdr, fg_color="transparent")
        tabs_container.grid(row=1, column=0, columnspan=6, sticky="ew", padx=16, pady=(8, 8))
        tabs_container.grid_columnconfigure(0, weight=1)

        tab_configs = [
            ("todas", "Todas"),
            ("ingreso", "Ingresos"),
            ("egreso", "Egresos"),
            ("sin_receptor", "Sin Receptor"),
            ("ors", "ORS"),
            ("pendiente", "Pendientes"),
            ("sin_clave", "📄 Sin clave"),
            ("omitidos", "⊘ Omitidos"),
            ("huerfanos", "🔍 Huérfanos"),
        ]

        for tab_id, tab_label in tab_configs:
            btn = ctk.CTkButton(
                tabs_container,
                text=tab_label,
                width=90,
                height=28,
                fg_color=CARD,
                hover_color=BORDER,
                text_color=TEXT,
                font=F_SMALL(),
                corner_radius=6,
                command=lambda t=tab_id: self._on_tab_clicked(t),
            )
            btn.pack(side="left", padx=4)
            self._tab_buttons[tab_id] = btn

        # Marcar pestaña inicial como activa
        self._update_tab_appearance("todas")

        # Separador
        ctk.CTkFrame(self, fg_color=BORDER, height=1, corner_radius=0).grid(
            row=0, column=0, columnspan=6, sticky="ew", pady=(0, 0)
        )

    def _build_body(self, parent=None):
        if parent is None:
            parent = self
        body = ctk.CTkFrame(parent, fg_color="transparent")
        body.grid(row=1, column=0, sticky="nsew", padx=8, pady=(8, 8))
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(0, weight=33, minsize=360)  # lista
        body.grid_columnconfigure(1, weight=52, minsize=520)  # visor PDF
        body.grid_columnconfigure(2, weight=15, minsize=260)  # clasificación

        self._build_list_panel(body)
        self._build_pdf_panel(body)
        self._build_classify_panel(body)

    # ── PANEL IZQUIERDO -- LISTA ───────────────────────────────────────────────
    def _build_list_panel(self, parent):
        frame = ctk.CTkFrame(parent, fg_color=SURFACE, corner_radius=10, border_width=1, border_color=BORDER)
        frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        frame.grid_rowconfigure(1, weight=1)
        frame.grid_columnconfigure(0, weight=1)  # Treeview

        # Header del panel
        top = ctk.CTkFrame(frame, fg_color=CARD, corner_radius=10, height=44)
        top.grid(row=0, column=0, sticky="ew")
        top.grid_propagate(False)
        top.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(top, text="Facturas del período",
                      font=F_TITLE(), text_color=TEXT).grid(
            row=0, column=0, sticky="w", padx=14, pady=12)

        ctk.CTkButton(top, text="Exportar reporte", width=120, height=30,
                      fg_color=SURFACE, hover_color=BORDER, text_color=TEXT,
                      font=F_SMALL(), corner_radius=8,
                      command=self._export_report).grid(row=0, column=2, sticky="e", padx=(0, 10))

        self._progress_var = ctk.StringVar(value="")
        ctk.CTkLabel(top, textvariable=self._progress_var,
                      font=F_SMALL(), text_color=TEAL).grid(
            row=0, column=1, sticky="e", padx=14)

        # Treeview con estilo oscuro
        tree_frame = ctk.CTkFrame(frame, fg_color=CARD, corner_radius=10)
        tree_frame.grid(row=1, column=0, sticky="nsew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Columnas: Emisor, Fecha, Tipo, IVA 13%, Impuesto, Total (sin Estado)
        cols = ("emisor", "fecha", "tipo", "iva_13", "impuesto", "total")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                  selectmode="extended", style="Dark.Treeview")
        self.tree.heading("emisor",    text="Emisor")
        self.tree.heading("fecha",     text="Fecha")
        self.tree.heading("tipo",      text="Tipo")
        self.tree.heading("iva_13",    text="IVA 13%")
        self.tree.heading("impuesto",  text="Impuesto")
        self.tree.heading("total",     text="Total")

        self.tree.column("emisor",    width=140)
        self.tree.column("fecha",     width=78,  stretch=False)
        self.tree.column("tipo",      width=46,  stretch=False)
        self.tree.column("iva_13",    width=80,  stretch=False, anchor="e")
        self.tree.column("impuesto",  width=90,  stretch=False, anchor="e")
        self.tree.column("total",     width=96,  stretch=False, anchor="e")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self.tree.yview, style="Dark.Vertical.TScrollbar")
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        # Etiquetas para colores de fondo según estado
        self.tree.tag_configure("clasificado",   background="#1a4d3d", foreground=TEXT)      # Verde oscuro
        self.tree.tag_configure("pendiente",     background="",        foreground=TEXT)      # Sin color (normal)
        self.tree.tag_configure("pendiente_pdf", background="#1a3d4d", foreground=TEXT)     # Azul oscuro
        self.tree.tag_configure("sin_xml",       background="#2d2d2d", foreground=MUTED)    # Gris oscuro
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Return>", lambda _e: self._classify_selected())

    # ── PANEL CENTRAL -- VISOR PDF ─────────────────────────────────────────────
    def _build_pdf_panel(self, parent):
        frame = ctk.CTkFrame(parent, fg_color=SURFACE, corner_radius=10, border_width=1, border_color=BORDER)
        frame.grid(row=0, column=1, sticky="nsew", padx=6)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        # NO usar grid_propagate(False) -- colapsa el frame a tamaño 0
        # El layout estable viene de minsize en las columnas del body (en _build_body)

        # PDFViewer ocupa TODO el panel -- incluye su propia toolbar internamente
        self.pdf_viewer = PDFViewer(frame)
        self.pdf_viewer.grid(row=0, column=0, sticky="nsew")

    # ── PANEL DERECHO -- CLASIFICACIÓN ────────────────────────────────────────
    def _build_classify_panel(self, parent):
        frame = ctk.CTkFrame(parent, fg_color=SURFACE, corner_radius=10, border_width=1, border_color=BORDER)
        frame.grid(row=0, column=2, sticky="nsew", padx=(6, 0))
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)

        # Header fijo
        top = ctk.CTkFrame(frame, fg_color=CARD, corner_radius=10, height=44)
        top.grid(row=0, column=0, sticky="ew")
        top.grid_propagate(False)
        ctk.CTkLabel(top, text="Clasificación",
                      font=F_TITLE(), text_color=TEXT).pack(side="left", padx=12, pady=10)

        # Contenido scrollable
        scroll = ctk.CTkScrollableFrame(
            frame, fg_color="transparent",
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=MUTED,
        )
        scroll.grid(row=1, column=0, sticky="nsew")
        scroll.grid_columnconfigure(0, weight=1)

        # ── Pill Hacienda ─────────────────────────────────────────────────────
        self._hacienda_lbl = ctk.CTkLabel(
            scroll, text="",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color=SUCCESS, fg_color="#0d2a1e", corner_radius=8,
            anchor="center",
        )
        self._hacienda_lbl.grid(row=0, column=0, sticky="ew",
                                 padx=12, pady=(12, 0), ipadx=6, ipady=4)

        # ── Panel clasificación contable ──────────────────────────────────────
        clf_border = ctk.CTkFrame(scroll, fg_color=BORDER, corner_radius=12)
        clf_border.grid(row=1, column=0, sticky="ew", padx=12, pady=10)
        clf = ctk.CTkFrame(clf_border, fg_color=CARD, corner_radius=10)
        clf.pack(fill="both", expand=True, padx=1, pady=1)
        clf.grid_columnconfigure(0, weight=1)

        # row 0 -- título sección
        ctk.CTkLabel(clf, text="CLASIFICACIÓN CONTABLE",
                      font=ctk.CTkFont(family="Segoe UI", size=9, weight="bold"),
                      text_color=TEAL).grid(row=0, column=0, sticky="w",
                                             padx=12, pady=(12, 6))

        # row 1-2 -- Categoría (siempre visible)
        ctk.CTkLabel(clf, text="Categoría", font=F_SMALL(),
                      text_color=MUTED).grid(row=1, column=0, sticky="w", padx=12)
        self._cat_var = ctk.StringVar()
        self._cat_cb = ctk.CTkComboBox(
            clf, variable=self._cat_var, values=[],
            state="readonly",
            fg_color=SURFACE, border_color=BORDER,
            button_color=BORDER, button_hover_color=TEAL,
            text_color=TEXT, font=F_LABEL(),
            dropdown_fg_color=CARD, dropdown_text_color=TEXT,
            command=self._on_categoria_change,
        )
        self._cat_cb.grid(row=2, column=0, sticky="ew", padx=12, pady=(2, 8))

        # row 3 -- Tipo (GASTOS: GENERALES/ESPECIFICOS | OGND: OGND/DNR/ORS/CNR)
        self._tipo_frame = ctk.CTkFrame(clf, fg_color="transparent")
        self._tipo_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(self._tipo_frame, text="Tipo", font=F_SMALL(),
                      text_color=MUTED).grid(row=0, column=0, sticky="w", padx=12)
        self._tipo_var = ctk.StringVar()
        self._tipo_cb = ctk.CTkComboBox(
            self._tipo_frame, variable=self._tipo_var, values=[],
            state="readonly",
            fg_color=SURFACE, border_color=BORDER,
            button_color=BORDER, button_hover_color=TEAL,
            text_color=TEXT, font=F_LABEL(),
            dropdown_fg_color=CARD, dropdown_text_color=TEXT,
            command=self._on_subtipo_change,
        )
        self._tipo_cb.grid(row=1, column=0, sticky="ew", padx=12, pady=(2, 8))
        self._tipo_frame.grid(row=3, column=0, sticky="ew")
        self._tipo_frame.grid_remove()  # oculto hasta que se seleccione GASTOS u OGND

        # row 4 -- Cuenta (solo para GASTOS)
        self._cuenta_frame = ctk.CTkFrame(clf, fg_color="transparent")
        self._cuenta_frame.grid_columnconfigure(0, weight=1)

        # Header row: Label + "New Account" button
        header_frame = ctk.CTkFrame(self._cuenta_frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=12)
        header_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header_frame, text="Cuenta", font=F_SMALL(),
                      text_color=MUTED).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(
            header_frame, text="➕", width=30, height=28,
            fg_color=TEAL, hover_color=TEAL_DIM,
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color="#0d1a18",
            command=self._open_new_cuenta_dialog,
        ).grid(row=0, column=1, sticky="e", padx=(8, 0))

        # Search filter
        self._cuenta_search_var = ctk.StringVar()
        self._cuenta_search_var.trace_add("write", self._filter_cuentas)
        ctk.CTkEntry(
            self._cuenta_frame, textvariable=self._cuenta_search_var,
            placeholder_text="Buscar cuenta...",
            fg_color=SURFACE, border_color=BORDER, text_color=TEXT,
            font=F_LABEL(), height=28, corner_radius=6,
        ).grid(row=1, column=0, sticky="ew", padx=12, pady=(4, 2))

        # ComboBox
        self._cuenta_var = ctk.StringVar()
        self._cuenta_cb = ctk.CTkComboBox(
            self._cuenta_frame, variable=self._cuenta_var, values=[],
            state="readonly",
            fg_color=SURFACE, border_color=BORDER,
            button_color=BORDER, button_hover_color=TEAL,
            text_color=TEXT, font=F_LABEL(),
            dropdown_fg_color=CARD, dropdown_text_color=TEXT,
            command=lambda _v: self._update_path_preview(),
        )
        self._cuenta_cb.grid(row=2, column=0, sticky="ew", padx=12, pady=(2, 8))
        self._cuenta_frame.grid(row=4, column=0, sticky="ew")
        self._cuenta_frame.grid_remove()

        # row 5 -- Proveedor (COMPRAS y GASTOS)
        self._prov_frame = ctk.CTkFrame(clf, fg_color="transparent")
        self._prov_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(self._prov_frame, text="Proveedor", font=F_SMALL(),
                      text_color=MUTED).grid(row=0, column=0, sticky="w", padx=12)
        self._prov_var = ctk.StringVar()
        self._prov_var.trace_add("write", lambda *_: self.after(0, self._update_path_preview))
        ctk.CTkEntry(
            self._prov_frame, textvariable=self._prov_var,
            fg_color=SURFACE, border_color=BORDER, text_color=TEXT,
            font=F_LABEL(), height=34, corner_radius=8,
        ).grid(row=1, column=0, sticky="ew", padx=12, pady=(2, 8))
        self._prov_frame.grid(row=5, column=0, sticky="ew")

        # row 6 -- Preview de ruta destino
        self._preview_lbl = ctk.CTkLabel(
            clf, text="",
            font=ctk.CTkFont(family="Segoe UI", size=10),
            text_color=MUTED,
            wraplength=220, justify="left", anchor="w",
        )
        self._preview_lbl.grid(row=6, column=0, sticky="ew", padx=12, pady=(0, 4))

        # row 7 -- Botón Recuperar (para PDFs huérfanos)
        self._btn_recover = ctk.CTkButton(
            clf, text="🔧  Recuperar PDF",
            font=F_BTN(), fg_color=WARNING, hover_color="#e8a61c",
            text_color="#0d1a18", corner_radius=10, height=40,
            state="disabled",
            command=self._recover_selected,
        )
        self._btn_recover.grid_remove()  # Ocultar por defecto, mostrar solo con huérfanos

        # row 7 -- Botón Vincular (para PDFs omitidos)
        self._btn_link = ctk.CTkButton(
            clf, text="🔗  Vincular a XML",
            font=F_BTN(), fg_color="#8b5cf6", hover_color="#7c3aed",
            text_color="#0d1a18", corner_radius=10, height=40,
            state="disabled",
            command=self._link_omitted_to_xml,
        )
        self._btn_link.grid_remove()  # Ocultar por defecto, mostrar solo con omitidos

        # row 7 -- Botón Borrar PDF (para PDFs omitidos)
        self._btn_delete = ctk.CTkButton(
            clf, text="🗑️  Borrar PDF",
            font=F_BTN(), fg_color=DANGER, hover_color="#dc2626",
            text_color="#0d1a18", corner_radius=10, height=40,
            state="disabled",
            command=self._delete_omitido,
        )
        self._btn_delete.grid_remove()  # Ocultar por defecto, mostrar solo con omitidos

        # row 7 -- Botón Crear PDF (para XMLs sin PDF)
        self._btn_create_pdf = ctk.CTkButton(
            clf, text="📄  Crear PDF",
            font=F_BTN(), fg_color="#3b82f6", hover_color="#2563eb",
            text_color="#0d1a18", corner_radius=10, height=40,
            state="disabled",
            command=self._create_pdf_for_selected,
        )
        self._btn_create_pdf.grid_remove()  # Ocultar por defecto

        # row 8 -- Botón Auto-clasificar (para Ingresos y Sin Receptor)
        self._btn_auto_classify = ctk.CTkButton(
            clf, text="⚡  Clasificar todos",
            font=F_BTN(), fg_color="#10b981", hover_color="#059669",
            text_color="#0d1a18", corner_radius=10, height=40,
            state="disabled",
            command=self._auto_classify_current_tab,
        )
        self._btn_auto_classify.grid_remove()  # Ocultar por defecto

        # row 9 -- Botón Clasificar
        self._btn_classify = ctk.CTkButton(
            clf, text="✔  Clasificar",
            font=F_BTN(), fg_color=TEAL, hover_color=TEAL_DIM,
            text_color="#0d1a18", corner_radius=10, height=40,
            state="disabled",
            command=self._classify_selected,
        )
        self._btn_classify.grid(row=9, column=0, sticky="ew", padx=12, pady=(0, 12))

        # ── Clasificación anterior ────────────────────────────────────────────
        self._prev_frame = ctk.CTkFrame(scroll, fg_color=CARD, corner_radius=10)
        self._prev_frame.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 12))
        self._prev_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(self._prev_frame, text="ANTERIOR",
                      font=ctk.CTkFont(family="Segoe UI", size=9, weight="bold"),
                      text_color=MUTED).grid(row=0, column=0, sticky="w",
                                              padx=10, pady=(8, 2))
        self._prev_var = ctk.StringVar(value="--")
        ctk.CTkLabel(self._prev_frame, textvariable=self._prev_var,
                      font=F_SMALL(), text_color="#555e6e",
                      justify="left", wraplength=200, anchor="w").grid(
            row=1, column=0, sticky="ew", padx=10, pady=(0, 8))

        # Label ruta destino (tk.Label nativo para cursor hand2 y color exacto)
        import tkinter as tk
        self._prev_path_lbl = tk.Label(
            self._prev_frame,
            text="",
            font=("Segoe UI", 10),
            fg=TEAL,
            bg=CARD,
            cursor="hand2",
            anchor="w",
            justify="left",
            wraplength=195,
        )
        self._prev_path_lbl.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 8))
        self._prev_path_lbl.bind("<Button-1>", self._open_dest_folder)

    # ── PESTAÑAS DE FACTURAS DEL PERÍODO ────────────────────────────────────────
    def _on_tab_clicked(self, tab: str):
        """Maneja click en pestaña. Filtra registros y actualiza UI."""
        self._active_tab = tab
        self._update_tab_appearance(tab)

        # Filtrar registros según pestaña activa
        if self.session:
            self.records = self._apply_filters()
            self._refresh_tree()
            self._update_progress()

    def _update_tab_appearance(self, active_tab: str):
        """Actualiza colores de botones de pestañas."""
        for tab_id, btn in self._tab_buttons.items():
            if tab_id == active_tab:
                btn.configure(fg_color=TEAL, text_color=BG)
            else:
                btn.configure(fg_color=CARD, text_color=TEXT)

    def _get_client_cedula(self) -> str:
        """Obtiene cédula confiable del cliente desde client_profiles.json.

        Estructura: client_name -> gmail_account -> __email__:{gmail} -> cedula
        Con fallback inteligente a cedula más frecuente en XMLs si no hay match.
        """
        if not self.session:
            return ""

        try:
            from app3.core.client_profiles import load_profiles
            profiles = load_profiles()

            # Paso 1: Buscar perfil del cliente por nombre
            client_name = self.session.nombre
            if client_name in profiles:
                profile = profiles[client_name]
                if isinstance(profile, dict):
                    # Paso 2: Obtener gmail_account del perfil
                    gmail = str(profile.get("gmail_account", "")).strip().lower()
                    if gmail:
                        # Paso 3: Buscar entrada __email__:{gmail}
                        email_key = f"__email__:{gmail}"
                        if email_key in profiles:
                            email_profile = profiles[email_key]
                            if isinstance(email_profile, dict):
                                cedula = str(email_profile.get("cedula", "")).strip()
                                if cedula:
                                    # Limpiar cédula (solo dígitos)
                                    import re
                                    cedula_clean = re.sub(r"\D", "", cedula)
                                    if cedula_clean:
                                        logger.info(f"Cédula obtenida de perfiles: {cedula_clean}")
                                        return cedula_clean
        except Exception as e:
            logger.warning(f"Error obteniendo cédula desde perfiles: {e}")

        # Fallback inteligente: si la cédula no match ningún registro,
        # usar la cédula más frecuente en todos los registros
        if self.all_records:
            from collections import Counter
            import re
            cedula_candidates = Counter()

            for r in self.all_records:
                if r.emisor_cedula:
                    cedula_clean = re.sub(r"\D", "", str(r.emisor_cedula))
                    if cedula_clean:
                        cedula_candidates[cedula_clean] += 1
                if r.receptor_cedula:
                    cedula_clean = re.sub(r"\D", "", str(r.receptor_cedula))
                    if cedula_clean:
                        cedula_candidates[cedula_clean] += 1

            if cedula_candidates:
                most_common_cedula, count = cedula_candidates.most_common(1)[0]
                logger.info(f"Cedula más frecuente en registros: {most_common_cedula} ({count} apariciones)")
                return most_common_cedula

        # Último fallback: cedula de sesión
        cedula_fallback = (self.session.cedula or "").strip()
        logger.info(f"Usando cedula fallback de sesión: {cedula_fallback}")
        return cedula_fallback

    # ── TABLA ─────────────────────────────────────────────────────────────────
    def _refresh_tree(self):
        """Refresca Treeview ordenado: Emisor -> Tipo -> Fecha (con colores por estado)."""
        self.tree.delete(*self.tree.get_children())

        if not self.records:
            return

        # Mapeo de tipo_documento -> abreviatura
        tipo_map = {
            "Factura Electrónica": "FE",
            "Factura electronica": "FE",
            "Nota de Crédito": "NC",
            "Nota de Débito": "ND",
            "Tiquete": "TQ",
        }

        # Reordenar por: Emisor -> Tipo de documento -> Fecha
        def sort_key(r):
            emisor = (r.emisor_nombre or "").lower()
            tipo = (r.tipo_documento or "").lower()
            fecha = r.fecha_emision or ""
            return (emisor, tipo, fecha)

        sorted_records = sorted(self.records, key=sort_key)

        # Preparar items + mapeo de clave -> record
        items_to_insert = []
        self._tree_clave_map = {}  # Mapeo visual: iid -> record (para _on_select)

        for idx, r in enumerate(sorted_records):
            # Estado para etiqueta de color
            estado = (self._db_records.get(r.clave, {}).get("estado") if self.db else None) or r.estado
            tag = estado if estado in ("clasificado", "pendiente", "pendiente_pdf", "sin_xml", "huerfano") else "pendiente"

            # Formatear campos
            tipo_raw = str(r.tipo_documento or "")
            tipo_short = tipo_map.get(tipo_raw, tipo_raw[:4])
            emisor_short = _short_name(r.emisor_nombre)

            # Valores numéricos con formato
            iva_13_fmt = _fmt_amount(r.iva_13)
            impuesto_fmt = _fmt_amount(r.impuesto_total)
            total_fmt = _fmt_amount(r.total_comprobante)

            # Orden de columnas: emisor, fecha, tipo, iva_13, impuesto, total
            row_values = (
                emisor_short,
                r.fecha_emision,
                tipo_short,
                iva_13_fmt,
                impuesto_fmt,
                total_fmt,
            )

            # Generar IID único: usar índice para evitar duplicados (ej: "506040..._{0}")
            iid = f"{r.clave}_{idx}" if r.clave else f"UNKNOWN_{idx}"
            items_to_insert.append((iid, row_values, tag))
            self._tree_clave_map[iid] = r  # Mapeo IID -> record

        # Insertar en batches para que UI responda
        batch_size = 200
        for batch_start in range(0, len(items_to_insert), batch_size):
            batch_end = min(batch_start + batch_size, len(items_to_insert))
            for iid, values, tag in items_to_insert[batch_start:batch_end]:
                self.tree.insert("", "end", iid=iid, values=values, tags=(tag,))

            self.update_idletasks()

            if hasattr(self, '_loading_overlay') and self._loading_overlay and self._loading_overlay.winfo_exists():
                pct = batch_end / len(items_to_insert)
                self._loading_overlay.update_progress(batch_end, len(items_to_insert))

    def _update_progress(self):
        if not self.records:
            self._progress_var.set("")
            return
        total = len(self.records)
        clf = sum(
            1
            for r in self.records
            if ((self._db_records.get(r.clave, {}).get("estado") if self.db else None) or r.estado)
            == "clasificado"
        )
        pct = int(clf / total * 100) if total else 0
        self._progress_var.set(f"{clf}/{total}  ({pct}%)")

    # ── SELECCIÓN ─────────────────────────────────────────────────────────────
    def _on_select(self, _event=None):
        sel = self.tree.selection()

        # Si no hay selección, limpiar estado
        if not sel:
            self.selected = None
            self.selected_records = []
            return

        # Resolver todos los records seleccionados
        records = []
        for iid in sel:
            if hasattr(self, '_tree_clave_map') and iid in self._tree_clave_map:
                records.append(self._tree_clave_map[iid])
            else:
                # Fallback: búsqueda por clave en records
                matches = [r for r in self.records if r.clave == iid]
                if matches:
                    records.append(matches[0])

        self.selected_records = records

        if len(records) == 1:
            # FLUJO EXISTENTE: una sola factura, sin cambios al comportamiento
            self._on_select_single(records[0])
        elif len(records) > 1:
            # MODO LOTE: múltiples facturas del mismo emisor
            self._on_multi_select(records)

    def _on_select_single(self, r: FacturaRecord):
        """Maneja selección de una sola factura (flujo original)."""
        self.selected = r

        # Estado Hacienda -- pill superior
        if r.estado_hacienda:
            esh = r.estado_hacienda.strip()
            color = SUCCESS if "aceptado" in esh.lower() else WARNING
            bg    = "#0d2a1e" if color == SUCCESS else "#2d2010"
            icon  = "✓" if color == SUCCESS else "⚠"
            self._hacienda_lbl.configure(
                text=f"{icon}  Hacienda: {esh}",
                text_color=color, fg_color=bg,
            )
        else:
            self._hacienda_lbl.configure(text="", fg_color="transparent")

        # PDF: Intentar cargar desde ruta original, o desde ruta clasificada si ya fue movido
        pdf_to_load = None

        # 1️⃣ Intenta ruta original (si no fue clasificado aún)
        if r.pdf_path and r.pdf_path.exists():
            pdf_to_load = r.pdf_path

        # 2️⃣ Si no está en original, busca en ruta destino (si ya fue clasificado)
        elif self.db and r.clave:
            db_record = self._db_records.get(r.clave)
            if db_record and db_record.get("ruta_destino"):
                ruta_destino = Path(db_record["ruta_destino"])
                if ruta_destino.exists():
                    pdf_to_load = ruta_destino
                    logger.debug(f"PDF cargado desde ruta clasificada: {ruta_destino}")

        # Cargar o mostrar vacío
        if pdf_to_load:
            self.pdf_viewer.load(pdf_to_load)
        else:
            # Si es un registro sin PDF, mostrar mensaje específico
            if r.estado == "pendiente_pdf":
                self.pdf_viewer.show_message(
                    "📄 XML sin PDF\n\nPresiona «Crear PDF» para generar\n"
                    "una factura a partir de los datos del XML."
                )
            elif r.razon_omisión:
                # Registro omitido - mostrar razón
                razon_text = {
                    "non_invoice": "Detectado como no-factura (borrador, catálogo, comunicado, etc.)",
                    "timeout": "Timeout durante extracción de clave",
                    "extract_failed": "Error al extraer información del PDF",
                }.get(r.razon_omisión, "PDF omitido")
                self.pdf_viewer.release_file_handles(f"⊘ PDF Omitido\n\n{razon_text}")
            else:
                self.pdf_viewer.clear()

        # Prellenar proveedor
        if r.emisor_nombre:
            self._prov_var.set(r.emisor_nombre)

        # Habilitar botones según tipo de registro
        if r.estado == "huerfano":
            # Es un PDF huérfano - permitir recuperación
            self._btn_recover.grid(row=7, column=0, sticky="ew", padx=12, pady=(0, 6))
            self._btn_recover.configure(state="normal")
            self._btn_link.grid_remove()
            self._btn_delete.grid_remove()
            self._btn_auto_classify.grid_remove()
            self._btn_classify.grid(row=9, column=0, sticky="ew", padx=12, pady=(0, 12))
            self._btn_classify.configure(state="disabled", text="⊘ No clasificable")
        elif r.razon_omisión:
            # Registro omitido - permitir vincular a XML o borrar
            self._btn_recover.grid_remove()
            self._btn_link.grid(row=7, column=0, sticky="ew", padx=12, pady=(0, 6))
            self._btn_link.configure(state="normal")
            self._btn_delete.grid(row=8, column=0, sticky="ew", padx=12, pady=(0, 6))
            self._btn_delete.configure(state="normal")
            self._btn_auto_classify.grid_remove()
            self._btn_classify.grid(row=9, column=0, sticky="ew", padx=12, pady=(0, 12))
            self._btn_classify.configure(state="disabled", text="⊘ No clasificable")
        else:
            # Registro normal - permitir clasificación
            self._btn_recover.grid_remove()
            self._btn_link.grid_remove()
            self._btn_delete.grid_remove()
            if r.estado == "pendiente_pdf":
                # XML sin PDF: ofrecer crear PDF
                self._btn_create_pdf.grid(row=7, column=0, sticky="ew", padx=12, pady=(0, 6))
                self._btn_create_pdf.configure(state="normal")
                self._btn_auto_classify.grid_remove()
                self._btn_classify.grid(row=8, column=0, sticky="ew", padx=12, pady=(0, 12))
                self._btn_classify.configure(state="normal", text="✔  Clasificar sin PDF")
            else:
                self._btn_create_pdf.grid_remove()
                # Mostrar botón auto-clasificar solo en Ingresos o Sin Receptor
                if self._active_tab in ("ingreso", "sin_receptor"):
                    self._btn_auto_classify.grid(row=7, column=0, sticky="ew", padx=12, pady=(0, 6))
                    self._btn_auto_classify.configure(state="normal")
                    self._btn_classify.grid(row=8, column=0, sticky="ew", padx=12, pady=(0, 12))
                else:
                    self._btn_auto_classify.grid_remove()
                    self._btn_classify.grid(row=7, column=0, sticky="ew", padx=12, pady=(0, 12))
                self._btn_classify.configure(state="normal", text="✔  Clasificar")

        # Clasificación previa
        if self.db:
            prev = self._db_records.get(r.clave)
            if prev and prev.get("estado") == "clasificado":
                crumbs = " › ".join(
                    p for p in [
                        prev.get("categoria"),
                        prev.get("subtipo"),
                        prev.get("nombre_cuenta"),
                        prev.get("proveedor"),
                    ] if p
                )
                self._prev_var.set(f"{crumbs}\n{prev.get('fecha_clasificacion', '')}")

                # Ruta destino
                ruta_dest = prev.get("ruta_destino", "")
                if ruta_dest:
                    parts = Path(ruta_dest).parts
                    snippet = ("📁 .../" + "/".join(parts[-4:]) + "/") if len(parts) > 4 else f"📁 {ruta_dest}/"
                    self._prev_path_lbl.config(text=snippet)
                else:
                    self._prev_path_lbl.config(text="")
            else:
                self._prev_var.set("--")
                self._prev_path_lbl.config(text="")

    def _on_multi_select(self, records: list[FacturaRecord]):
        """Maneja selección de múltiples facturas (modo lote)."""
        # Verificar que todos pertenecen al mismo emisor
        cedulas = {r.emisor_cedula for r in records}

        if len(cedulas) > 1:
            # Emisores diferentes -> bloquear con aviso
            self.pdf_viewer.show_message(
                f"Advertencia: {len(records)} facturas de {len(cedulas)} emisores distintos.\n"
                "Solo se puede clasificar en lote facturas del MISMO EMISOR."
            )
            self._btn_classify.configure(
                state="disabled",
                text=f"Advertencia: Emisores distintos ({len(cedulas)})"
            )
            self._btn_auto_classify.grid_remove()
            # Ocultar pill Hacienda
            self._hacienda_lbl.configure(text="", fg_color="transparent")
            # Ocultar panel de clasificación anterior
            self._prev_frame.grid_remove()
            return

        # Mismo emisor -> mostrar modo lote
        self.selected = records[0]  # Mantener referencia al primero
        emisor_nombre = records[0].emisor_nombre or "Emisor desconocido"

        # Mostrar mensaje de modo lote en lugar del PDF
        self.pdf_viewer.show_message(
            f"Lote de clasificación\n"
            f"{len(records)} facturas seleccionadas\n\n"
            f"Emisor: {emisor_nombre}\n\n"
            "Complete la clasificacion y haga click\n"
            "en 'Clasificar N facturas'."
        )

        # Pre-llenar proveedor automaticamente con el nombre del emisor
        self._prov_var.set(emisor_nombre)

        # Cambiar texto del botón clasificar
        self._btn_classify.configure(
            state="normal",
            text=f"Clasificar {len(records)} facturas"
        )

        # Ocultar botón auto-clasificar en modo lote
        self._btn_auto_classify.grid_remove()

        # Ocultar pill Hacienda
        self._hacienda_lbl.configure(text="", fg_color="transparent")

        # Ocultar panel de clasificación anterior
        self._prev_frame.grid_remove()

    # ── CATÁLOGO ──────────────────────────────────────────────────────────────
    def _on_categoria_change(self, _value=None):
        cat = self._cat_var.get()
        mgr = self.catalog_mgr

        if cat == "COMPRAS":
            self._tipo_frame.grid_remove()
            self._cuenta_frame.grid_remove()
            self._prov_frame.grid()

        elif cat == "ACTIVO":
            self._tipo_frame.grid_remove()
            self._cuenta_frame.grid_remove()
            self._prov_frame.grid()

        elif cat == "GASTOS":
            tipos = mgr.subtipos("GASTOS") if mgr else []
            self._tipo_cb.configure(values=tipos)
            self._tipo_var.set(tipos[0] if tipos else "")
            self._tipo_frame.grid()
            self._cuenta_frame.grid()
            self._prov_frame.grid()
            self._on_subtipo_change()  # actualiza cuentas
            return  # _on_subtipo_change llama a _update_path_preview

        elif cat == "OGND":
            tipos = mgr.subtipos("OGND") if mgr else ["OGND", "DNR", "ORS", "CNR"]
            self._tipo_cb.configure(values=tipos)
            self._tipo_var.set(tipos[0] if tipos else "")
            self._tipo_frame.grid()
            self._cuenta_frame.grid_remove()
            self._prov_frame.grid_remove()

        elif cat == "INGRESOS" or cat == "SIN_RECEPTOR":
            # Ingresos y Sin Receptor: sin subtipo, cuenta, ni proveedor
            self._tipo_frame.grid_remove()
            self._cuenta_frame.grid_remove()
            self._prov_frame.grid_remove()

        self._update_path_preview()

    def _on_subtipo_change(self, _value=None):
        cat    = self._cat_var.get()
        subtipo = self._tipo_var.get()
        mgr    = self.catalog_mgr

        if cat == "GASTOS" and mgr:
            cuentas = mgr.cuentas("GASTOS", subtipo)
            self._all_cuentas = cuentas  # Store unfiltered list
            self._cuenta_search_var.set("")  # Clear search
            self._cuenta_cb.configure(values=cuentas)
            self._cuenta_var.set(cuentas[0] if cuentas else "")
            self._cuenta_frame.grid()
        else:
            self._cuenta_frame.grid_remove()

        self._update_path_preview()

    def _filter_cuentas(self, *_args):
        """Filter accounts by search term (case-insensitive substring match)."""
        search_term = self._cuenta_search_var.get().strip().lower()
        if search_term:
            filtered = [c for c in self._all_cuentas if search_term in c.lower()]
        else:
            filtered = self._all_cuentas
        self._cuenta_cb.configure(values=filtered)
        # Auto-select first matching item if available
        if filtered and not self._cuenta_var.get():
            self._cuenta_var.set(filtered[0])
        elif filtered and self._cuenta_var.get() not in filtered:
            self._cuenta_var.set(filtered[0])

    def _open_new_cuenta_dialog(self):
        """Open modal dialog to add a new account."""
        if not self.catalog_mgr:
            self._show_warning("Atención", "Catálogo no disponible.")
            return

        cat = self._cat_var.get()
        subtipo = self._tipo_var.get()
        if cat != "GASTOS" or not subtipo:
            self._show_warning("Atención", "Selecciona una categoría y tipo primero.")
            return

        dialog = NewCuentaDialog(
            self,
            categoria=cat,
            subtipo=subtipo,
            existing_cuentas=self._all_cuentas,
            catalog_mgr=self.catalog_mgr,
            on_success=self._on_nueva_cuenta_added,
        )

    def _on_nueva_cuenta_added(self, new_cuenta: str):
        """Called when a new account is successfully added."""
        # Reload the subtipo to get updated list
        self._on_subtipo_change()
        # Select the new account
        if new_cuenta in self._all_cuentas:
            self._cuenta_var.set(new_cuenta)
        self._update_path_preview()
        self._set_status(f"Cuenta '{new_cuenta}' agregada.")

    def _update_path_preview(self):
        """Muestra la ruta de destino estimada debajo del formulario."""
        if not self.session:
            self._preview_lbl.configure(text="")
            return

        cat     = self._cat_var.get()
        subtipo = self._tipo_var.get() if cat in ("GASTOS", "OGND") else ""
        cuenta  = self._cuenta_var.get() if cat == "GASTOS" else ""
        prov    = self._prov_var.get().strip() if cat in ("COMPRAS", "GASTOS", "ACTIVO") else ""
        fecha   = (self.selected.fecha_emision if self.selected else "") or ""

        # Validar mínimos necesarios para construir la ruta
        if not cat:
            self._preview_lbl.configure(text="")
            return
        if cat in ("COMPRAS", "ACTIVO") and not prov:
            self._preview_lbl.configure(text="")
            return
        if cat == "GASTOS" and not (subtipo and cuenta and prov):
            self._preview_lbl.configure(text="")
            return
        if cat == "OGND" and not subtipo:
            self._preview_lbl.configure(text="")
            return
        # INGRESOS y SIN_RECEPTOR no necesitan validación (sin subtipo/cuenta/prov)

        try:
            dest  = build_dest_folder(self.session.folder, fecha, cat, subtipo, cuenta, prov)
            parts = dest.parts
            snippet = (".../" + "/".join(parts[-4:]) + "/") if len(parts) > 4 else str(dest) + "/"
            self._preview_lbl.configure(text=snippet)
        except Exception:
            self._preview_lbl.configure(text="")

    def _on_filter(self):
        if not self.all_records:
            return
        self.records = self._apply_filters()
        self.selected = None
        self.selected_records = []
        self.pdf_viewer.clear()
        self._refresh_tree()
        self._update_progress()
        self._set_status("Filtro aplicado")

    def _open_date_picker(self, target: str):
        self._close_date_picker()
        initial = self.from_var.get() if target == "from" else self.to_var.get()
        entry = self._from_entry if target == "from" else self._to_entry

        def on_pick(value: str):
            if target == "from":
                self.from_var.set(value)
            else:
                self.to_var.set(value)
            self._close_date_picker()

        self.update_idletasks()
        x = entry.winfo_rootx()
        y = entry.winfo_rooty() + entry.winfo_height() + 6
        self._active_calendar = DatePickerDropdown(
            self,
            on_pick=on_pick,
            initial_value=initial,
            x=x,
            y=y,
        )

    def _close_date_picker(self):
        if self._active_calendar is not None:
            self._active_calendar._safe_close()
            self._active_calendar = None

    def _export_report(self):
        if not self.records:
            messagebox.showinfo("Exportar", "No hay registros para exportar.")
            return

        default_name = _default_export_filename(
            self._lbl_cliente.cget("text"),
            self.from_var.get(),
            self.to_var.get(),
        )
        target = filedialog.asksaveasfilename(
            title="Exportar reporte",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            confirmoverwrite=True,
        )
        if not target:
            return

        # Eliminar archivo existente antes de escribir (evita error en Windows si está cerrado)
        target_path = Path(target)
        if target_path.exists():
            try:
                target_path.unlink()
            except PermissionError:
                self._show_error(
                    "Error al exportar",
                    f"El archivo está abierto en otro programa.\nCiérralo e intenta de nuevo:\n{target}",
                )
                return

        client_cedula = self._get_client_cedula()

        # Exportar todos los registros del período (sin filtro de pestaña, sin omitidos)
        period_records = [r for r in self._apply_date_filter(self.all_records) if not r.razon_omisión]

        rows: list[dict] = []
        for r in period_records:
            meta = self._db_records.get(r.clave, {}) if self.db else {}
            estado = meta.get("estado") or r.estado
            rows.append(
                {
                    "clave_numerica": r.clave,
                    "tipo_documento": r.tipo_documento,
                    "fecha_emision": r.fecha_emision,
                    "consecutivo": r.consecutivo,
                    "emisor_nombre": r.emisor_nombre,
                    "emisor_cedula": r.emisor_cedula,
                    "receptor_nombre": r.receptor_nombre,
                    "receptor_cedula": r.receptor_cedula,
                    "moneda": r.moneda,
                    "tipo_cambio": r.tipo_cambio,
                    "subtotal": r.subtotal,
                    "iva_1": r.iva_1,
                    "iva_2": r.iva_2,
                    "iva_4": r.iva_4,
                    "iva_8": r.iva_8,
                    "iva_13": r.iva_13,
                    "iva_otros": r.iva_otros,
                    "impuesto_total": r.impuesto_total,
                    "total_comprobante": r.total_comprobante,
                    "estado_hacienda": r.estado_hacienda,
                    "detalle_estado_hacienda": r.detalle_estado_hacienda,
                    "categoria": str(meta.get("categoria") or ""),
                    "subtipo": str(meta.get("subtipo") or ""),
                    "nombre_cuenta": str(meta.get("nombre_cuenta") or ""),
                    "estado": estado,
                }
            )

        export_columns = [
            "tipo_documento",
            "fecha_emision",
            "consecutivo",
            "emisor_nombre",
            "emisor_cedula",
            "receptor_nombre",
            "receptor_cedula",
            "moneda",
            "tipo_cambio",
            "subtotal",
            "iva_1",
            "iva_2",
            "iva_4",
            "iva_8",
            "iva_13",
            "iva_otros",
            "impuesto_total",
            "total_comprobante",
            "estado_hacienda",
            "detalle_estado_hacienda",
            "categoria",
            "subtipo",
            "nombre_cuenta",
            "estado",
        ]

        # Columnas ocultas: usadas internamente (agrupación Gasto) pero no visibles en ninguna hoja
        _HIDDEN = {"subtipo", "nombre_cuenta", "estado", "categoria", "detalle_estado_hacienda"}
        display_columns = [c for c in export_columns if c not in _HIDDEN]

        numeric_columns = {
            "subtotal", "tipo_cambio",
            "iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros",
            "impuesto_total", "total_comprobante",
        }
        text_columns = {"clave_numerica", "consecutivo", "emisor_cedula", "receptor_cedula"}
        date_column = "fecha_emision"

        pretty_headers = {
            "clave_numerica": "Clave",
            "tipo_documento": "Tipo documento",
            "fecha_emision": "Fecha emisión",
            "consecutivo": "Consecutivo",
            "emisor_nombre": "Emisor",
            "emisor_cedula": "Cédula emisor",
            "receptor_nombre": "Receptor",
            "receptor_cedula": "Cédula receptor",
            "moneda": "Moneda",
            "tipo_cambio": "Tipo cambio",
            "subtotal": "Subtotal",
            "iva_1": "IVA 1%",
            "iva_2": "IVA 2%",
            "iva_4": "IVA 4%",
            "iva_8": "IVA 8%",
            "iva_13": "IVA 13%",
            "impuesto_total": "Impuesto total",
            "total_comprobante": "Total comprobante",
            "estado_hacienda": "Estado Hacienda",
            "detalle_estado_hacienda": "Detalle Estado Hacienda",
            "categoria": "Categoría",
            "subtipo": "Subtipo",
            "nombre_cuenta": "Cuenta",
            "estado": "Estado App 3",
        }

        try:
            if target.lower().endswith(".csv"):
                with open(target, "w", newline="", encoding="utf-8-sig") as fh:
                    writer = csv.DictWriter(fh, fieldnames=export_columns)
                    writer.writeheader()
                    writer.writerows([{col: row.get(col, "") for col in export_columns} for row in rows])
            else:
                import pandas as pd
                from openpyxl.styles import Alignment, Font, PatternFill

                df_all = pd.DataFrame(rows)
                df = df_all[[col for col in export_columns if col in df_all.columns]].copy()

                for col in text_columns:
                    if col in df.columns:
                        df[col] = df[col].fillna("").astype(str).str.strip()

                for col in numeric_columns:
                    if col in df.columns:
                        df[col] = pd.to_numeric(
                            df[col].astype(str).str.replace(" ", "", regex=False).str.replace(",", ".", regex=False),
                            errors="coerce",
                        )

                if date_column in df.columns:
                    df[date_column] = pd.to_datetime(df[date_column], format="%d/%m/%Y", errors="coerce")

                # ── Sheet splitting idéntico a App 2 ─────────────────────────
                emisor_raw = df_all["emisor_cedula"].fillna("").astype(str).str.strip()
                receptor_raw = df_all["receptor_cedula"].fillna("").astype(str).str.strip()
                receptor_is_empty = receptor_raw.str.lower().isin({"", "null", "none", "nan"})

                mask_ventas = emisor_raw.eq(client_cedula)
                mask_egreso = ~mask_ventas & receptor_raw.eq(client_cedula)
                mask_sin_receptor = ~mask_ventas & ~mask_egreso & receptor_is_empty

                # Egresos: GASTOS -> hoja "Gastos"; COMPRAS -> hoja "Compras"; OGND -> hoja "OGND"
                categoria_upper = df_all["categoria"].fillna("").astype(str).str.strip().str.upper()
                estado_export = df_all["estado"].fillna("").astype(str).str.strip().str.lower()

                mask_gasto = mask_egreso & categoria_upper.eq("GASTOS")
                mask_ognd = categoria_upper.eq("OGND")  # OGND: sin restricción de receptor (gastos propios de la empresa)
                mask_compras = mask_egreso & categoria_upper.eq("COMPRAS")
                mask_pendiente = mask_egreso & estado_export.eq("pendiente")

                # Nueva máscara: Rechazados por Hacienda (todos los registros, no solo egresos)
                estado_hacienda_col = df_all["estado_hacienda"].fillna("").astype(str).str.strip()
                mask_rechazados = estado_hacienda_col.eq("Rechazada")

                used_names: set[str] = set()
                sheet_map: dict[str, pd.DataFrame] = {}
                for label, mask in [
                    ("Ingresos", mask_ventas),
                    ("Compras", mask_compras),
                    ("Gastos", mask_gasto),
                    ("OGND", mask_ognd),
                    ("Pendientes", mask_pendiente),
                    ("Sin Receptor", mask_sin_receptor),
                    ("Rechazados", mask_rechazados),
                ]:
                    chunk = df.loc[mask]
                    if not chunk.empty:
                        sheet_map[_safe_excel_sheet_name(label, used_names)] = chunk.copy()

                if not sheet_map:
                    sheet_map[_safe_excel_sheet_name("Reporte", used_names)] = df.copy()

                owner_name = self._lbl_cliente.cget("text") or "REPORTE DE COMPROBANTES"
                date_from_label = self.from_var.get().strip() or "01/01/1900"
                date_to_label = self.to_var.get().strip() or datetime.now().strftime("%d/%m/%Y")

                title_fill = PatternFill(fill_type="solid", fgColor="0B2B66")
                subtitle_fill = PatternFill(fill_type="solid", fgColor="7F7F7F")
                summary_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
                header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
                credit_fill = PatternFill(fill_type="solid", fgColor="DAF2D0")
                title_font = Font(bold=True, color="FFFFFF", size=22)
                subtitle_font = Font(bold=True, color="FFFFFF", size=14)
                summary_font = Font(bold=False, color="111111", size=12)
                header_font = Font(bold=True)

                # ── Función auxiliar: Filtrar columnas IVA con todos ceros ──────
                def _filter_iva_cols(cols, sdf):
                    """Elimina columnas IVA cuyo valor es todo-cero en el DataFrame dado."""
                    IVA_COLS = {"iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros"}
                    ZERO_VALUES = {"", "0", "0.0", "0,00", "0.00", "nan", "none", "null"}
                    result = []
                    for col in cols:
                        if col not in IVA_COLS:
                            result.append(col)
                        else:
                            col_values = sdf[col].astype(str).str.strip().str.lower()
                            if col_values.loc[~col_values.isin(ZERO_VALUES)].any():
                                result.append(col)
                    return result

                with pd.ExcelWriter(target, engine="openpyxl") as writer:
                    # Eliminar hoja vacía por defecto que crea openpyxl
                    if "Sheet" in writer.book.sheetnames:
                        del writer.book["Sheet"]

                    for sheet_name, sheet_df in sheet_map.items():
                        # ── Hoja Gastos: layout agrupado especial ──────────────
                        if sheet_name == "Gastos":
                            ws = writer.book.create_sheet(title=sheet_name)
                            writer.sheets[sheet_name] = ws
                            # Usar solo columnas que existen en sheet_df (CRÍTICO para catálogo de cuentas)
                            # Excluir columnas de receptor (Gastos es egreso, receptor es el cliente)
                            gasto_base = [c for c in display_columns if c not in {"receptor_nombre", "receptor_cedula"} and c in sheet_df.columns]
                            # Aplicar filtro IVA-cero
                            gasto_cols = _filter_iva_cols(gasto_base, sheet_df)
                            _write_gasto_grouped(
                                ws, sheet_df, gasto_cols,
                                numeric_columns, text_columns, date_column,
                                pretty_headers, owner_name, sheet_name,
                                date_from_label, date_to_label,
                                title_fill, subtitle_fill, summary_fill,
                                header_fill, credit_fill,
                                title_font, subtitle_font, summary_font, header_font,
                            )
                            continue

                        # ── Hoja OGND: layout agrupado especial ──────────────
                        if sheet_name == "OGND":
                            ws = writer.book.create_sheet(title=sheet_name)
                            writer.sheets[sheet_name] = ws
                            # Aplicar filtro IVA-cero
                            ognd_cols = _filter_iva_cols([c for c in display_columns if c in sheet_df.columns], sheet_df)
                            _write_gasto_grouped(
                                ws, sheet_df, ognd_cols,
                                numeric_columns, text_columns, date_column,
                                pretty_headers, owner_name, sheet_name,
                                date_from_label, date_to_label,
                                title_fill, subtitle_fill, summary_fill,
                                header_fill, credit_fill,
                                title_font, subtitle_font, summary_font, header_font,
                            )
                            continue

                        # ── Hoja Rechazados: incluye detalle_estado_hacienda ──────────────
                        if sheet_name == "Rechazados":
                            # Usar _HIDDEN sin detalle_estado_hacienda para que la columna sea visible
                            # Excluir también receptor (no aplica a Rechazados)
                            rechazados_hidden = {"subtipo", "nombre_cuenta", "estado", "categoria", "receptor_nombre", "receptor_cedula"}
                            rechazados_cols = [c for c in export_columns
                                               if c not in rechazados_hidden and c in sheet_df.columns]
                            # Filtrar columnas IVA todo-cero (igual que hojas normales)
                            IVA_COLS = {"iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros"}
                            ZERO_VALUES = {"", "0", "0.0", "0,00", "0.00", "nan", "none", "null"}
                            visible_rechazados = []
                            for col in rechazados_cols:
                                if col not in IVA_COLS:
                                    visible_rechazados.append(col)
                                else:
                                    col_values = sheet_df[col].astype(str).str.strip().str.lower()
                                    has_nonzero = col_values.loc[~col_values.isin(ZERO_VALUES)].any()
                                    if has_nonzero:
                                        visible_rechazados.append(col)
                            display_df = sheet_df[visible_rechazados].rename(
                                columns={col: pretty_headers.get(col, col.replace("_", " ").title()) for col in visible_rechazados}
                            )
                            display_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=4)
                            ws = writer.sheets[sheet_name]

                            max_col = ws.max_column if ws.max_column > 0 else 1
                            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
                            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)

                            title_cell = ws.cell(row=1, column=1)
                            title_cell.value = str(owner_name).upper()
                            title_cell.font = title_font
                            title_cell.alignment = Alignment(horizontal="center", vertical="center")
                            title_cell.fill = title_fill

                            subtitle_cell = ws.cell(row=2, column=1)
                            subtitle_cell.value = f"REPORTE DE {sheet_name.upper()} - Período: {date_from_label} al {date_to_label}"
                            subtitle_cell.font = subtitle_font
                            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
                            subtitle_cell.fill = subtitle_fill

                            monto_total = Decimal("0")
                            if "total_comprobante" in sheet_df.columns:
                                valid_amounts = []
                                for value in sheet_df["total_comprobante"].dropna().tolist():
                                    try:
                                        valid_amounts.append(Decimal(str(value)))
                                    except Exception:
                                        continue
                                if valid_amounts:
                                    monto_total = sum(valid_amounts, Decimal("0"))

                            monedas = (
                                sorted({str(m).strip() for m in sheet_df["moneda"].dropna().tolist() if str(m).strip()})
                                if "moneda" in sheet_df.columns else []
                            )
                            moneda_value = (
                                "N/A" if not monedas
                                else monedas[0] if len(monedas) == 1
                                else "MIXTA: " + ", ".join(monedas)
                            )
                            generated = datetime.now().strftime("%d/%m/%Y %H:%M")

                            summary_cell = ws.cell(row=3, column=1)
                            summary_cell.value = (
                                f"Total filas: {len(sheet_df)}   |   Monto Total: {_format_amount_es(monto_total)}   |   "
                                f"Moneda: {moneda_value}   |   Generado: {generated}"
                            )
                            summary_cell.font = summary_font
                            summary_cell.alignment = Alignment(horizontal="center", vertical="center")
                            summary_cell.fill = summary_fill

                            # Centrar y negrear encabezados
                            header_row = 5
                            for col_idx in range(1, ws.max_column + 1):
                                header_cell = ws.cell(row=header_row, column=col_idx)
                                header_cell.font = header_font
                                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                                header_cell.fill = header_fill

                            # Loop de formato de celdas de datos
                            for col_idx, col_name in enumerate(visible_rechazados, start=1):
                                for row_idx in range(header_row + 1, len(sheet_df) + header_row + 1):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    if col_name in text_columns:
                                        cell.number_format = "@"
                                        cell.value = "" if cell.value is None else str(cell.value)
                                    elif col_name == date_column and cell.value is not None:
                                        cell.number_format = "dd/mm/yyyy"
                                    elif col_name in numeric_columns and cell.value is not None:
                                        cell.number_format = "#,##0.00"
                                        if isinstance(cell.value, Decimal):
                                            cell.value = float(cell.value)

                            # Resaltado de NC en verde
                            tipo_idx = (
                                visible_rechazados.index("tipo_documento") + 1
                                if "tipo_documento" in visible_rechazados else None
                            )
                            if tipo_idx is not None:
                                for row_idx in range(header_row + 1, len(sheet_df) + header_row + 1):
                                    if ws.cell(row=row_idx, column=tipo_idx).value == "Nota de Crédito":
                                        for col in range(1, ws.max_column + 1):
                                            ws.cell(row=row_idx, column=col).fill = credit_fill

                            # Auto-width de columnas
                            for col_idx in range(1, ws.max_column + 1):
                                max_len = 0
                                for row_idx in range(header_row, ws.max_row + 1):
                                    value = ws.cell(row=row_idx, column=col_idx).value
                                    if value is None:
                                        continue
                                    max_len = max(max_len, len(str(value)))
                                ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = min(max(max_len + 3, 12), 65)

                            # Freeze panes
                            ws.freeze_panes = ws["A6"]

                            continue

                        # Hojas normales: solo columnas visibles (filtrar IVA todo-cero)
                        # Excluir receptor excepto para "Sin Receptor"
                        exclude_receptor = sheet_name != "Sin Receptor"
                        visible_cols_base = [
                            c for c in display_columns
                            if c in sheet_df.columns and not (exclude_receptor and c in {"receptor_nombre", "receptor_cedula"})
                        ]

                        # Aplicar filtro IVA-cero usando la función auxiliar
                        visible_cols_filtered = _filter_iva_cols(visible_cols_base, sheet_df)

                        display_df = sheet_df[visible_cols_filtered].rename(
                            columns={col: pretty_headers.get(col, col.replace("_", " ").title()) for col in visible_cols_filtered}
                        )
                        display_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=4)
                        ws = writer.sheets[sheet_name]

                        max_col = ws.max_column if ws.max_column > 0 else 1
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
                        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)

                        title_cell = ws.cell(row=1, column=1)
                        title_cell.value = str(owner_name).upper()
                        title_cell.font = title_font
                        title_cell.alignment = Alignment(horizontal="center", vertical="center")
                        title_cell.fill = title_fill

                        subtitle_cell = ws.cell(row=2, column=1)
                        subtitle_cell.value = f"REPORTE DE {sheet_name.upper()} - Período: {date_from_label} al {date_to_label}"
                        subtitle_cell.font = subtitle_font
                        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
                        subtitle_cell.fill = subtitle_fill

                        monto_total = Decimal("0")
                        if "total_comprobante" in sheet_df.columns:
                            valid_amounts = []
                            for value in sheet_df["total_comprobante"].dropna().tolist():
                                try:
                                    valid_amounts.append(Decimal(str(value)))
                                except Exception:
                                    continue
                            if valid_amounts:
                                monto_total = sum(valid_amounts, Decimal("0"))

                        monedas = (
                            sorted({str(m).strip() for m in sheet_df["moneda"].dropna().tolist() if str(m).strip()})
                            if "moneda" in sheet_df.columns
                            else []
                        )
                        moneda_value = (
                            "N/A" if not monedas
                            else monedas[0] if len(monedas) == 1
                            else "MIXTA: " + ", ".join(monedas)
                        )
                        generated = datetime.now().strftime("%d/%m/%Y %H:%M")

                        summary_cell = ws.cell(row=3, column=1)
                        summary_cell.value = (
                            f"Total filas: {len(sheet_df)}   |   Monto Total: {_format_amount_es(monto_total)}   |   "
                            f"Moneda: {moneda_value}   |   Generado: {generated}"
                        )
                        summary_cell.font = summary_font
                        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
                        summary_cell.fill = summary_fill

                        header_row = 5
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=header_row, column=col_idx)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        tipo_idx = (
                            visible_cols_filtered.index("tipo_documento") + 1
                            if "tipo_documento" in visible_cols_filtered else None
                        )

                        for col_idx, col_name in enumerate(visible_cols_filtered, start=1):
                            for row_idx in range(header_row + 1, len(sheet_df) + header_row + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if col_name in text_columns:
                                    cell.number_format = "@"
                                    cell.value = "" if cell.value is None else str(cell.value)
                                elif col_name == date_column and cell.value is not None:
                                    cell.number_format = "dd/mm/yyyy"
                                elif col_name in numeric_columns and cell.value is not None:
                                    cell.number_format = "#,##0.00"
                                    if isinstance(cell.value, Decimal):
                                        cell.value = float(cell.value)

                        if tipo_idx is not None:
                            for row_idx in range(header_row + 1, len(sheet_df) + header_row + 1):
                                if ws.cell(row=row_idx, column=tipo_idx).value == "Nota de Crédito":
                                    for col in range(1, ws.max_column + 1):
                                        ws.cell(row=row_idx, column=col).fill = credit_fill

                        for col_idx in range(1, ws.max_column + 1):
                            max_len = 0
                            for row_idx in range(header_row, ws.max_row + 1):
                                value = ws.cell(row=row_idx, column=col_idx).value
                                if value is None:
                                    continue
                                max_len = max(max_len, len(str(value)))
                            ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = min(max(max_len + 3, 12), 65)

                        ws.freeze_panes = ws["A6"]

            messagebox.showinfo("Exportar", f"Reporte guardado en:\n{target}")
            self._set_status("Reporte exportado")
        except Exception as exc:
            self._show_error("Error al exportar", str(exc))

    # ── RECUPERACIÓN DE PDFs HUÉRFANOS ────────────────────────────────────────
    def _recover_selected(self):
        """Recupera el PDF huérfano seleccionado."""
        if not self.selected or not self.db:
            return

        # Obtener metadata del PDF huérfano
        orphaned_info = getattr(self.selected, "_orphaned_info", None)
        if not orphaned_info:
            self._show_warning("Error", "Este PDF no tiene información de recuperación")
            return

        ruta_actual = orphaned_info.get("ruta_actual")
        ruta_esperada = orphaned_info.get("ruta_esperada")
        clave = orphaned_info.get("clave")

        if not ruta_esperada:
            self._show_warning("Error", "No se pudo determinar la ruta de destino")
            return

        # Confirmar
        motivo = orphaned_info.get("motivo", "desconocido")
        if not self._ask(
            "Recuperar PDF",
            f"¿Recuperar este PDF?\n\n"
            f"Clave: {clave}\n"
            f"Motivo: {motivo}\n"
            f"De: {Path(ruta_actual).name}\n"
            f"A: {Path(ruta_esperada).name}"
        ):
            return

        # Ejecutar recuperación
        def worker():
            try:
                from app3.core.classifier import recover_orphaned_pdf
                if recover_orphaned_pdf(orphaned_info, self.db):
                    self.after(0, lambda: self._show_info("✓ Recuperado", "PDF movido exitosamente"))
                    self.after(0, self._load_session, self.session)
                else:
                    self.after(0, lambda: self._show_error("Error", "No se pudo recuperar el PDF"))
            except Exception as e:
                self.after(0, lambda error=e: self._show_error("Error", str(error)))

        threading.Thread(target=worker, daemon=True).start()

    def _link_omitted_to_xml(self):
        """Vincula un PDF omitido a un XML sin PDF disponible."""
        if not self.selected or not self.db:
            return

        # Obtener XMLs sin PDF que puedan ser vinculados a este PDF
        xmls_without_pdf = [
            r for r in self.all_records
            if r.xml_path and not r.pdf_path and r.estado == "pendiente_pdf"
        ]

        if not xmls_without_pdf:
            self._show_warning(
                "Sin XMLs disponibles",
                "No hay XMLs sin PDF disponibles para vincular.\n\n"
                "Asegúrate de que haya archivos XML en la carpeta que aún no tengan PDF."
            )
            return

        # Crear diálogo de selección con Toplevel para mejor UX
        selection_window = ctk.CTkToplevel(self)
        selection_window.title("Vincular PDF a XML")
        selection_window.geometry("600x400")
        selection_window.configure(fg_color=BG)

        # Header
        header = ctk.CTkFrame(selection_window, fg_color=SURFACE, corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(
            header,
            text="🔗 Seleccionar XML para vincular",
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
            text_color=TEXT,
        ).pack(side="left", padx=16, pady=12)

        # Body con Treeview
        body = ctk.CTkFrame(selection_window, fg_color=BG)
        body.pack(fill="both", expand=True, padx=12, pady=12)

        # Treeview
        from tkinter import ttk
        tree_frame = ctk.CTkFrame(body, fg_color=CARD, corner_radius=8)
        tree_frame.pack(fill="both", expand=True, pady=(0, 12))
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.grid(row=0, column=1, sticky="ns")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background=CARD, foreground=TEXT, fieldbackground=CARD)
        style.configure("Treeview.Heading", background=SURFACE, foreground=TEXT)

        tree = ttk.Treeview(
            tree_frame,
            columns=("emisor", "fecha", "tipo", "clave"),
            height=12,
            yscrollcommand=scrollbar.set,
            selectmode="browse",
        )
        scrollbar.configure(command=tree.yview)

        tree.column("#0", width=0, stretch=False)
        tree.column("emisor", anchor="w", width=200)
        tree.column("fecha", anchor="w", width=80)
        tree.column("tipo", anchor="w", width=60)
        tree.column("clave", anchor="w", width=150)

        tree.heading("#0", text="")
        tree.heading("emisor", text="Emisor")
        tree.heading("fecha", text="Fecha")
        tree.heading("tipo", text="Tipo")
        tree.heading("clave", text="Clave")

        # Llenar árbol con XMLs disponibles
        for idx, r in enumerate(xmls_without_pdf):
            tree.insert(
                "", "end",
                iid=str(idx),
                values=(
                    r.emisor_nombre[:30],
                    r.fecha_emision,
                    (r.tipo_documento or "")[:10],
                    r.clave,
                )
            )

        tree.grid(row=0, column=0, sticky="nsew")

        # Botones
        button_frame = ctk.CTkFrame(body, fg_color="transparent")
        button_frame.pack(fill="x")

        def on_cancel():
            selection_window.destroy()

        def on_select():
            selection = tree.selection()
            if not selection:
                self._show_warning("Selección", "Selecciona un XML para vincular")
                return

            idx = int(selection[0])
            selected_xml = xmls_without_pdf[idx]

            selection_window.destroy()

            # Confirmar vinculación
            if not self._ask(
                "Confirmar vinculación",
                f"¿Vincular el PDF a este XML?\n\n"
                f"Emisor: {selected_xml.emisor_nombre}\n"
                f"Fecha: {selected_xml.fecha_emision}\n"
                f"Tipo: {selected_xml.tipo_documento}\n"
                f"Clave: {selected_xml.clave}"
            ):
                return

            # Ejecutar vinculación
            def worker():
                try:
                    # Actualizar el PDF omitido para que tenga la misma clave que el XML
                    self.selected.clave = selected_xml.clave
                    self.selected.estado = "pendiente"
                    self.selected.razon_omisión = None

                    # Actualizar DB
                    if self.db:
                        self.db.upsert(
                            clave_numerica=selected_xml.clave,
                            estado="pendiente",
                            ruta_origen=str(selected_xml.xml_path or ""),
                            ruta_destino="",
                        )

                    self.after(0, lambda: self._show_info(
                        "✓ Vinculación completada",
                        f"PDF vinculado al XML de {selected_xml.emisor_nombre}"
                    ))
                    self.after(0, self._load_session, self.session)
                except Exception as e:
                    self.after(0, lambda error=e: self._show_error("Error en vinculación", str(error)))

            threading.Thread(target=worker, daemon=True).start()

        ctk.CTkButton(
            button_frame, text="Cancelar", fg_color=SURFACE, hover_color=BORDER,
            text_color=TEXT, command=on_cancel, width=100, corner_radius=8,
        ).pack(side="right", padx=(8, 0))

        ctk.CTkButton(
            button_frame, text="Vincular", fg_color="#8b5cf6", hover_color="#7c3aed",
            text_color="#0d1a18", command=on_select, width=100, corner_radius=8,
        ).pack(side="right")

    def _create_pdf_for_selected(self):
        """Crea un PDF desde los datos del XML para facturas sin PDF."""
        r = self.selected
        if not r or r.estado != "pendiente_pdf" or not r.xml_path:
            return

        # Ruta destino: .../CLIENT/PDF/{clave}.pdf
        client_folder = r.xml_path.parent.parent
        pdf_root = client_folder / "PDF"
        output_path = pdf_root / f"{r.clave}.pdf"

        self._btn_create_pdf.configure(state="disabled", text="Generando...")

        def _do_generate():
            from app3.core.pdf_generator import generate_factura_pdf, extract_items_from_xml
            try:
                # Intentar extraer items del XML
                items = extract_items_from_xml(r.xml_path) if r.xml_path and r.xml_path.exists() else None
                generate_factura_pdf(r, output_path, items=items)
                self.after(0, lambda: _on_done())
            except Exception as exc:
                err = str(exc)
                self.after(0, lambda e=err: _on_error(e))

        def _on_done():
            r.pdf_path = output_path
            r.estado = "pendiente"
            self._refresh_tree()
            self.pdf_viewer.load(output_path)
            self._on_select_single(r)  # Re-trigger para actualizar botones

        def _on_error(err):
            self._btn_create_pdf.configure(state="normal", text="📄  Crear PDF")
            messagebox.showerror("Error al generar PDF", f"No se pudo generar el PDF:\n{err}", parent=self)

        threading.Thread(target=_do_generate, daemon=True).start()

    def _delete_omitido(self):
        """Borra uno o más PDFs omitidos del disco."""
        # Determinar qué registros borrar
        records_to_delete = []
        if self.selected_records:
            # Multi-selección: verificar que todos sean omitidos
            records_to_delete = [r for r in self.selected_records if r.razon_omisión]
        elif self.selected and self.selected.razon_omisión:
            # Selección simple
            records_to_delete = [self.selected]

        if not records_to_delete:
            self._show_warning("Advertencia", "Selecciona PDFs omitidos para borrar")
            return

        # Preparar mensaje de confirmación
        if len(records_to_delete) == 1:
            msg = (
                f"¿Borrar este PDF omitido?\n\n"
                f"Archivo: {records_to_delete[0].pdf_path.name if records_to_delete[0].pdf_path else 'desconocido'}\n"
                f"Razón: {records_to_delete[0].razon_omisión}\n\n"
                "Esta acción no se puede deshacer."
            )
        else:
            msg = (
                f"¿Borrar {len(records_to_delete)} PDFs omitidos?\n\n"
                f"Se eliminarán {len(records_to_delete)} archivos de forma permanente.\n"
                "Esta acción no se puede deshacer."
            )

        if not self._ask("Borrar PDFs", msg):
            return

        # Ejecutar borrado en worker
        def worker():
            deleted = 0
            errors = []

            for record in records_to_delete:
                try:
                    if record.pdf_path:
                        # Cerrar el documento PDF en el viewer para liberar el archivo
                        if self.selected and self.selected.clave == record.clave:
                            self.after(0, self.pdf_viewer._close_doc)
                            time.sleep(0.1)

                        Path(record.pdf_path).unlink(missing_ok=True)
                        # Eliminar del registro en memoria
                        self.all_records = [r for r in self.all_records if r.clave != record.clave]
                        deleted += 1
                except Exception as e:
                    errors.append(f"{record.pdf_path.name if record.pdf_path else 'desconocido'}: {str(e)}")

            # Actualizar UI
            if deleted > 0:
                self.selected = None
                self.selected_records = []
                self.after(0, self._refresh_tree)
                self.after(0, self.pdf_viewer._close_doc)

                if deleted == 1:
                    self.after(0, lambda: self._show_info("✓ PDF borrado", f"Archivo eliminado exitosamente"))
                else:
                    self.after(0, lambda: self._show_info("✓ PDFs borrados", f"{deleted} archivos eliminados exitosamente"))

            if errors:
                msg_error = "Errores en el borrado:\n\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    msg_error += f"\n... y {len(errors) - 5} errores más"
                self.after(0, lambda: self._show_error("Error", msg_error))

        threading.Thread(target=worker, daemon=True).start()

    def _show_info(self, title: str, message: str):
        """Muestra un mensaje de información."""
        messagebox.showinfo(title, message)

    # ── CLASIFICACIÓN ─────────────────────────────────────────────────────────
    def _classify_selected(self):
        if not self.session or not self.selected or not self.db:
            return

        # Bloquear clasificación de facturas rechazadas por Hacienda
        all_selected = self.selected_records if self.selected_records else ([self.selected] if self.selected else [])
        rechazados = [r for r in all_selected if r.estado_hacienda == "Rechazada"]
        if rechazados:
            nombres = "\n".join(r.emisor_nombre[:50] for r in rechazados[:5])
            suffix = f"\n... y {len(rechazados) - 5} más" if len(rechazados) > 5 else ""
            self._show_warning(
                "Factura rechazada",
                f"Las siguientes facturas fueron rechazadas por Hacienda y no pueden clasificarse:\n{nombres}{suffix}"
            )
            return

        cat    = self._cat_var.get().strip().upper()
        subtipo = self._tipo_var.get().strip().upper() if cat in ("GASTOS", "OGND") else ""
        cuenta  = self._cuenta_var.get().strip().upper() if cat == "GASTOS" else ""
        prov    = self._prov_var.get().strip().upper() if cat in ("COMPRAS", "GASTOS", "ACTIVO") else ""

        if not cat:
            self._show_warning("Atención", "Selecciona una categoría.")
            return
        if cat == "GASTOS" and not subtipo:
            self._show_warning("Atención", "Selecciona el tipo de gasto.")
            return
        if cat == "GASTOS" and not cuenta:
            self._show_warning("Atención", "Selecciona la cuenta contable.")
            return
        if cat in ("COMPRAS", "GASTOS", "ACTIVO") and not prov:
            self._show_warning("Atención", "Ingresa el proveedor.")
            return
        if cat == "OGND" and not subtipo:
            self._show_warning("Atención", "Selecciona el tipo OGND.")
            return
        # INGRESOS y SIN_RECEPTOR no necesitan validación adicional

        if len(self.selected_records) <= 1:
            # FLUJO EXISTENTE: clasificar una sola factura
            if self._db_records.get(self.selected.clave, {}).get("estado") == "clasificado":
                if not self._ask("Reclasificar",
                                  "Esta factura ya fue clasificada.\n¿Deseas reclasificarla?"):
                    return

            self._btn_classify.configure(state="disabled", text="Clasificando...")

            # Liberar lock del PDF mostrado antes de mover/eliminar en Windows.
            self.pdf_viewer.release_file_handles("Procesando clasificación...")

            record  = self.selected
            session = self.session
            db      = self.db

            def worker():
                try:
                    classify_record(record, session.folder, db, cat, subtipo, cuenta, prov)
                    self.after(0, self._on_classify_ok)
                except Exception as exc:
                    self.after(0, lambda e=exc: self._on_classify_error(str(e)))

            threading.Thread(target=worker, daemon=True).start()
        else:
            # MODO LOTE: clasificar múltiples facturas
            records_to_classify = list(self.selected_records)
            n = len(records_to_classify)

            # Pedir confirmación si alguno ya estaba clasificado
            ya_clasificados = [
                r for r in records_to_classify
                if self._db_records.get(r.clave, {}).get("estado") == "clasificado"
            ]
            if ya_clasificados:
                if not self._ask("Reclasificar en lote",
                                  f"{len(ya_clasificados)} de {n} facturas ya estan clasificadas.\n"
                                  f"¿Deseas reclasificarlas?"):
                    return

            self._btn_classify.configure(state="disabled", text=f"Clasificando 0/{n}...")
            self.pdf_viewer.release_file_handles("Clasificando en lote...")

            record_list = records_to_classify
            session = self.session
            db = self.db

            def worker():
                errores = []
                for i, record in enumerate(record_list):
                    self.after(0, lambda i=i: self._btn_classify.configure(
                        text=f"Clasificando {i+1}/{n}..."
                    ))
                    try:
                        classify_record(record, session.folder, db, cat, subtipo, cuenta, prov)
                    except Exception as exc:
                        errores.append((record, str(exc)))

                self.after(0, lambda: self._on_batch_classify_done(n, errores))

            threading.Thread(target=worker, daemon=True).start()

    def _on_classify_ok(self):
        if self.db and self.selected:
            updated = self.db.get_record(self.selected.clave)
            if updated:
                self._db_records[self.selected.clave] = updated
        saved_clave = self.selected.clave if self.selected else None
        self._btn_classify.configure(state="normal", text="✔  Clasificar")
        self._refresh_tree()
        self._update_progress()
        # Auto-avance: seleccionar el siguiente registro si existe
        if saved_clave:
            # Buscar IID correcto (formato: clave_idx) usando _tree_clave_map
            saved_iid = next(
                (iid for iid, r in self._tree_clave_map.items() if r.clave == saved_clave),
                None
            )
            if saved_iid and self.tree.exists(saved_iid):
                # Intentar obtener el siguiente registro
                next_iid = self.tree.next(saved_iid)
                # Si hay siguiente, seleccionarlo; si no, quedarse en el actual
                target_iid = next_iid if next_iid else saved_iid
                self.tree.selection_set(target_iid)
                self.tree.focus(target_iid)
                self.tree.see(target_iid)
        self._on_select()

    def _on_batch_classify_done(self, total: int, errores: list[tuple]):
        """Maneja finalización de clasificación en lote."""
        exitosos = total - len(errores)

        # Actualizar BD local en memoria con los nuevos estados
        for r in self.selected_records:
            db_rec = self.db.get_record(r.clave) if self.db else None
            if db_rec:
                self._db_records[r.clave] = db_rec

        # Refrescar árbol para actualizar estado visual de las filas
        self._refresh_tree()
        self._update_progress()

        # Restaurar botón y mostrar resultado
        if errores:
            self._btn_classify.configure(
                state="normal",
                text=f"Clasificacion: {exitosos}/{total} OK"
            )
            # Mostrar primeros errores (máximo 5)
            msgs = "\n".join(
                f"• {r.emisor_nombre}: {e}"
                for r, e in errores[:5]
            )
            if len(errores) > 5:
                msgs += f"\n... y {len(errores) - 5} errores mas"
            self._show_error("Errores en clasificacion en lote", msgs)
        else:
            self._btn_classify.configure(
                state="normal",
                text=f"Clasificacion completada: {total} OK"
            )
            self._show_info("Exito", f"{total} facturas clasificadas correctamente.")

        # Limpiar selección multi
        self.selected_records = []
        self.tree.selection_remove(self.tree.selection())

    def _auto_classify_current_tab(self):
        """Auto-clasifica todos los registros en la pestaña actual (Ingresos o Sin Receptor)."""
        if not self.session or not self.db or self._active_tab not in ("ingreso", "sin_receptor"):
            return

        # Determinar categoría según tab activo
        if self._active_tab == "ingreso":
            categoria = "INGRESOS"
        elif self._active_tab == "sin_receptor":
            categoria = "SIN_RECEPTOR"
        else:
            return

        # Obtener registros de la pestaña actual que no estén clasificados
        registros_a_clasificar = [
            r for r in self.records
            if not r.razon_omisión
            and self._db_records.get(r.clave, {}).get("estado") != "clasificado"
        ]

        if not registros_a_clasificar:
            self._show_info("Sin registros", f"No hay registros pendientes en {self._active_tab}")
            return

        # Pedir confirmación
        if not self._ask(
            f"Auto-clasificar {categoria}",
            f"¿Clasificar {len(registros_a_clasificar)} registros como {categoria}?\n\n"
            f"Se moverán automáticamente a:\n"
            f"Contabilidades/[mes]/[cliente]/{categoria}/"
        ):
            return

        # Ejecutar clasificación en worker thread
        self._btn_auto_classify.configure(state="disabled", text=f"Clasificando 0/{len(registros_a_clasificar)}...")
        self.pdf_viewer.release_file_handles("Clasificando en lote...")

        records_list = registros_a_clasificar
        session = self.session
        db = self.db
        n = len(records_list)

        def worker():
            errores = []
            for i, record in enumerate(records_list):
                self.after(0, lambda i=i: self._btn_auto_classify.configure(
                    text=f"Clasificando {i+1}/{n}..."
                ))
                try:
                    # Para INGRESOS y SIN_RECEPTOR, no se necesitan subtipo, cuenta, proveedor
                    classify_record(record, session.folder, db, categoria, "", "", "")
                except Exception as exc:
                    errores.append((record, str(exc)))

            self.after(0, lambda: self._on_batch_classify_done(n, errores))

        threading.Thread(target=worker, daemon=True).start()

    @staticmethod
    def _parse_ui_date(value: str):
        text = (value or "").strip()
        if not text:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def _apply_filters(self) -> list[FacturaRecord]:
        """Aplica filtro de pestaña activa + filtro de fecha sobre all_records."""
        tab_filtered = filter_records_by_tab(
            self.all_records,
            self._active_tab,
            self._get_client_cedula(),
            self._db_records,
        )
        return self._apply_date_filter(tab_filtered)

    def _apply_date_filter(self, records: list[FacturaRecord]) -> list[FacturaRecord]:
        from_dt = self._parse_ui_date(self.from_var.get())
        to_dt = self._parse_ui_date(self.to_var.get())
        if not from_dt and not to_dt:
            return list(records)

        filtered: list[FacturaRecord] = []
        for record in records:
            fecha_txt = (record.fecha_emision or "").strip()
            try:
                fecha = datetime.strptime(fecha_txt, "%d/%m/%Y").date()
            except ValueError:
                # Mantener visibles los PDF sin XML para revisión manual.
                if record.estado == "sin_xml":
                    filtered.append(record)
                continue
            if from_dt and fecha < from_dt:
                continue
            if to_dt and fecha > to_dt:
                continue
            filtered.append(record)
        return filtered

    def _on_classify_error(self, msg: str):
        self._btn_classify.configure(state="normal", text="✔  Clasificar")
        self._show_error("Error al clasificar", msg)

    # ── HELPERS ───────────────────────────────────────────────────────────────
    def _set_status(self, text: str):
        self._status_var.set(text)

    def _show_error(self, title: str, msg: str):
        win = ctk.CTkToplevel(self)
        win.title(title)
        win.geometry("420x200")
        win.configure(fg_color=CARD)
        win.grab_set()
        ctk.CTkLabel(win, text=f"✗  {title}", font=F_BTN(),
                      text_color=DANGER).pack(pady=(24, 8))
        ctk.CTkLabel(win, text=msg, font=F_SMALL(), text_color=TEXT,
                      wraplength=380, justify="center").pack(pady=(0, 16))
        ctk.CTkButton(win, text="Cerrar", fg_color=SURFACE,
                       hover_color=BORDER, text_color=TEXT,
                       command=win.destroy).pack()

    def _show_warning(self, title: str, msg: str):
        win = ctk.CTkToplevel(self)
        win.title(title)
        win.geometry("440x220")
        win.configure(fg_color=CARD)
        win.grab_set()
        ctk.CTkLabel(win, text=f"⚠  {title}", font=F_BTN(),
                      text_color=WARNING).pack(pady=(24, 8))
        ctk.CTkLabel(win, text=msg, font=F_SMALL(), text_color=TEXT,
                      wraplength=400, justify="center").pack(pady=(0, 16))
        ctk.CTkButton(win, text="Entendido", fg_color=SURFACE,
                       hover_color=BORDER, text_color=TEXT,
                       command=win.destroy).pack()

    def _ask(self, title: str, msg: str) -> bool:
        result = [False]
        win = ctk.CTkToplevel(self)
        win.title(title)
        win.geometry("380x180")
        win.configure(fg_color=CARD)
        win.grab_set()
        ctk.CTkLabel(win, text=msg, font=F_LABEL(), text_color=TEXT,
                      wraplength=340, justify="center").pack(pady=(28, 16))
        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack()

        def yes():
            result[0] = True
            win.destroy()

        ctk.CTkButton(btns, text="Sí, reclasificar", fg_color=TEAL,
                       hover_color=TEAL_DIM, text_color="#0d1a18",
                       command=yes).pack(side="left", padx=8)
        ctk.CTkButton(btns, text="Cancelar", fg_color=SURFACE,
                       hover_color=BORDER, text_color=TEXT,
                       command=win.destroy).pack(side="left", padx=8)
        win.wait_window()
        return result[0]

    def _open_dest_folder(self, event=None):
        """Abre en el Explorador de Windows la carpeta donde está el PDF clasificado."""
        if not self.selected or not self.db:
            return
        db_rec = self._db_records.get(self.selected.clave)
        if not db_rec or not db_rec.get("ruta_destino"):
            return
        try:
            ruta = Path(db_rec["ruta_destino"])
            import subprocess
            if ruta.exists():
                # Seleccionar el archivo en la carpeta
                subprocess.Popen(["explorer", f"/select,{ruta}"])
            elif ruta.parent.exists():
                # El archivo no existe pero la carpeta sí — abrir la carpeta
                subprocess.Popen(["explorer", str(ruta.parent)])
        except Exception as e:
            logger.warning(f"No se pudo abrir carpeta destino: {e}")
