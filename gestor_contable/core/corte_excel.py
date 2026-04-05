"""Generador del reporte Excel de Corte Mensual.

Produce un workbook con hasta 4 hojas:
  INGRESOS  — facturas donde el cliente es emisor
  COMPRAS   — egresos clasificados como compras/insumos
  GASTOS    — egresos clasificados como gastos operativos
  AMBIGUO   — facturas sin clasificar (solo si quedan pendientes)

Sigue el mismo estilo visual del reporte de contabilidad completo
(mismos colores, estructura de encabezado, freeze panes, auto-ancho).
"""
from __future__ import annotations

import calendar
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from gestor_contable.core.corte_engine import (
    CorteItem,
    CATEGORIA_INGRESOS,
    CATEGORIA_COMPRAS,
    CATEGORIA_GASTOS,
    CATEGORIA_AMBIGUO,
)
from gestor_contable.core.iva_utils import apply_exchange_rate, parse_decimal_value
from gestor_contable.core.report_paths import month_abbr_es

LOGGER = logging.getLogger(__name__)

# ── Paleta Excel (idéntica al reporte de contabilidad) ────────────────────────
_TITLE_COLOR    = "0B2B66"   # azul oscuro — nombre del cliente
_SUBTITLE_COLOR = "7F7F7F"   # gris — título de hoja
_SUMMARY_COLOR  = "EDEDED"   # gris claro — resumen estadístico
_HEADER_COLOR   = "D9E1F2"   # azul claro — encabezados de columna
_CREDIT_COLOR   = "DAF2D0"   # verde — notas de crédito
_TOTAL_COLOR    = "BDD7EE"   # azul medio — fila de totales
_AMBIGUO_COLOR  = "FFF2CC"   # amarillo claro — fila AMBIGUO

# ── Columnas exportadas ───────────────────────────────────────────────────────
# Mismas columnas para todas las hojas; la columna de contraparte cambia según
# la perspectiva (emisor para egresos, receptor para ingresos).

_COLS_EGRESO = [
    "fecha_emision",
    "tipo_documento",
    "consecutivo",
    "emisor_nombre",
    "emisor_cedula",
    "moneda",
    "tipo_cambio",
    "subtotal",
    "iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros",
    "impuesto_total",
    "total_comprobante",
    "estado_hacienda",
]

_COLS_INGRESO = [
    "fecha_emision",
    "tipo_documento",
    "consecutivo",
    "receptor_nombre",
    "receptor_cedula",
    "moneda",
    "tipo_cambio",
    "subtotal",
    "iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros",
    "impuesto_total",
    "total_comprobante",
    "estado_hacienda",
]

_PRETTY_HEADERS = {
    "fecha_emision":    "Fecha",
    "tipo_documento":   "Tipo Doc.",
    "consecutivo":      "Consecutivo",
    "emisor_nombre":    "Proveedor",
    "emisor_cedula":    "Cédula Prov.",
    "receptor_nombre":  "Cliente / Receptor",
    "receptor_cedula":  "Cédula Receptor",
    "moneda":           "Moneda",
    "tipo_cambio":      "Tipo Cambio",
    "subtotal":         "Subtotal",
    "iva_1":            "IVA 1%",
    "iva_2":            "IVA 2%",
    "iva_4":            "IVA 4%",
    "iva_8":            "IVA 8%",
    "iva_13":           "IVA 13%",
    "iva_otros":        "IVA Otros",
    "impuesto_total":   "Total IVA",
    "total_comprobante":"Total",
    "estado_hacienda":  "Estado Hacienda",
}

_IVA_COLS = {"iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros"}
# Columnas de monto que se suman en totales y se convierten a CRC si aplica.
_NUMERIC_COLS = {"subtotal", "iva_1", "iva_2", "iva_4", "iva_8", "iva_13",
                 "iva_otros", "impuesto_total", "total_comprobante"}
# Columnas numéricas que se formatean como número pero NO se suman ni se convierten.
_NUMERIC_DISPLAY_COLS = {"tipo_cambio"}
_TEXT_COLS = {"consecutivo", "emisor_cedula", "receptor_cedula", "moneda"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_amount(raw: Any) -> Decimal:
    """Convierte texto de monto (coma decimal) a Decimal."""
    s = str(raw or "").strip()
    if not s:
        return Decimal("0")
    s = s.replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _to_float(raw: Any) -> float | None:
    """Convierte monto a float para celda Excel, o None si vacío."""
    d = _parse_amount(raw)
    return float(d) if d != Decimal("0") or str(raw or "").strip() else None


def _record_value(item: CorteItem, col: str) -> Any:
    """Extrae el valor de un campo del FacturaRecord del CorteItem."""
    return getattr(item.record, col, None)


def _filter_iva_cols(cols: list[str], items: list[CorteItem]) -> list[str]:
    """Elimina columnas IVA que son todas cero en la lista de items."""
    result = []
    for col in cols:
        if col in _IVA_COLS:
            total = sum(_parse_amount(_record_value(i, col)) for i in items)
            if total == Decimal("0"):
                continue
        result.append(col)
    return result


def _safe_sheet_name(raw: str, used: set[str]) -> str:
    invalid = {"\\", "/", "*", "?", ":", "[", "]"}
    cleaned = "".join("_" if ch in invalid else ch for ch in str(raw).strip())[:31]
    name = cleaned or "HOJA"
    suffix = 1
    while name in used:
        sfx = f" ({suffix})"
        name = cleaned[:31 - len(sfx)] + sfx
        suffix += 1
    used.add(name)
    return name


def _month_label(dt: datetime) -> str:
    meses = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
             7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
    return meses.get(dt.month, "MES")


# ── Escritura de una hoja ──────────────────────────────────────────────────────

def _write_sheet(
    ws,
    items: list[CorteItem],
    cols: list[str],
    sheet_label: str,
    client_name: str,
    period_label: str,
) -> None:
    """
    Escribe una hoja completa del reporte de corte.

    Estructura:
        Fila 1 : nombre del cliente  (title_fill)
        Fila 2 : "CORTE DE [LABEL] — Período: [period]"  (subtitle_fill)
        Fila 3 : resumen  (summary_fill)
        Fila 4 : vacía
        Fila 5 : encabezados  (header_fill)
        Fila 6+ : datos
        Última  : TOTALES  (total_fill)
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    title_fill    = PatternFill("solid", fgColor=_TITLE_COLOR)
    subtitle_fill = PatternFill("solid", fgColor=_SUBTITLE_COLOR)
    summary_fill  = PatternFill("solid", fgColor=_SUMMARY_COLOR)
    header_fill   = PatternFill("solid", fgColor=_HEADER_COLOR)
    credit_fill   = PatternFill("solid", fgColor=_CREDIT_COLOR)
    total_fill    = PatternFill("solid", fgColor=_TOTAL_COLOR)
    ambiguo_fill  = PatternFill("solid", fgColor=_AMBIGUO_COLOR)

    title_font    = Font(bold=True, color="FFFFFF", size=22)
    subtitle_font = Font(bold=True, color="FFFFFF", size=14)
    summary_font  = Font(bold=False, color="111111", size=12)
    header_font   = Font(bold=True)
    total_font    = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")
    n = len(cols)

    # ── Filas 1-3: encabezado ─────────────────────────────────────────────────
    for row in (1, 2, 3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22

    ws.cell(1, 1).value     = client_name.upper()
    ws.cell(1, 1).font      = title_font
    ws.cell(1, 1).fill      = title_fill
    ws.cell(1, 1).alignment = center

    ws.cell(2, 1).value     = f"CORTE DE {sheet_label.upper()} - Período: {period_label}"
    ws.cell(2, 1).font      = subtitle_font
    ws.cell(2, 1).fill      = subtitle_fill
    ws.cell(2, 1).alignment = center

    # Calcular monto total (referencia CRC para el resumen)
    monto_total = Decimal("0")
    for _i in items:
        _rec = _i.record
        _raw_total = _parse_amount(getattr(_rec, "total_comprobante", "") or "")
        _moneda    = str(getattr(_rec, "moneda", "") or "").strip().upper()
        _tc        = parse_decimal_value(getattr(_rec, "tipo_cambio", "") or "") or Decimal("0")
        if _moneda and _moneda != "CRC" and _tc > Decimal("0"):
            monto_total += apply_exchange_rate(_raw_total, _moneda, _tc)
        else:
            monto_total += _raw_total

    monedas = sorted({str(getattr(i.record, "moneda", "") or "CRC").strip() for i in items})
    moneda_str = monedas[0] if len(monedas) == 1 else "MIXTA"

    _monto_str = "{:,.2f}".format(float(monto_total)).replace(",", " ").replace(".", ",")
    ws.cell(3, 1).value = (
        f"Total filas: {len(items)}"
        f"   |   Monto Total CRC: {_monto_str}"
        f"   |   Moneda: {moneda_str}"
        f"   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws.cell(3, 1).font      = summary_font
    ws.cell(3, 1).fill      = summary_fill
    ws.cell(3, 1).alignment = center

    # ── Fila 5: encabezados de columna ────────────────────────────────────────
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(5, ci)
        cell.value     = _PRETTY_HEADERS.get(col, col.replace("_", " ").title())
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    # ── Filas de datos ────────────────────────────────────────────────────────
    totals: dict[str, Decimal] = {c: Decimal("0") for c in cols if c in _NUMERIC_COLS}
    current_row = 6

    for item in items:
        rec = item.record
        is_nc = "crédito" in str(getattr(rec, "tipo_documento", "")).lower()
        is_ambiguo = item.categoria == CATEGORIA_AMBIGUO

        _rec_moneda = str(getattr(rec, "moneda", "") or "").strip().upper()
        _rec_tc     = parse_decimal_value(getattr(rec, "tipo_cambio", "") or "") or Decimal("0")

        for ci, col in enumerate(cols, 1):
            raw = getattr(rec, col, None)
            cell = ws.cell(current_row, ci)

            if col in _NUMERIC_COLS:
                raw_dec = _parse_amount(raw)
                crc_dec = apply_exchange_rate(raw_dec, _rec_moneda, _rec_tc)
                val = float(crc_dec) if (crc_dec != Decimal("0") or str(raw or "").strip()) else None
                cell.value         = val
                cell.number_format = "#,##0.00"
                if val is not None:
                    totals[col] += crc_dec
            elif col in _NUMERIC_DISPLAY_COLS:
                # Numérico de solo visualización (tipo_cambio): mostrar pero no sumar ni convertir
                val = _to_float(raw)
                cell.value         = val if val is not None else ""
                cell.number_format = "#,##0.00" if val is not None else "@"
            elif col in _TEXT_COLS:
                cell.value         = str(raw or "")
                cell.number_format = "@"
            elif col == "fecha_emision":
                cell.value         = str(raw or "")
                cell.number_format = "@"
            else:
                cell.value = str(raw or "") if raw is not None else ""

        # Colores de fila
        row_fill = (
            ambiguo_fill if is_ambiguo else
            credit_fill  if is_nc      else
            None
        )
        if row_fill:
            for ci in range(1, n + 1):
                ws.cell(current_row, ci).fill = row_fill

        current_row += 1

    # ── Fila de TOTALES ───────────────────────────────────────────────────────
    for ci in range(1, n + 1):
        ws.cell(current_row, ci).fill = total_fill

    ws.cell(current_row, 1).value = "TOTAL"
    ws.cell(current_row, 1).font  = total_font
    ws.cell(current_row, 1).alignment = center

    for ci, col in enumerate(cols, 1):
        if col in totals:
            cell = ws.cell(current_row, ci)
            cell.value         = float(totals[col])
            cell.number_format = "#,##0.00"
            cell.font          = total_font

    # ── Auto-ancho de columnas ────────────────────────────────────────────────
    for ci in range(1, n + 1):
        max_len = 0
        for ri in range(5, current_row + 1):
            v = ws.cell(ri, ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        col_letter = ws.cell(5, ci).column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 65)

    ws.freeze_panes = ws["A6"]


# ── Función principal ─────────────────────────────────────────────────────────

def generar_corte_excel(
    resultados: list[CorteItem],
    client_name: str,
    output_path: Path | str,
    mes: int | None = None,
    anio: int | None = None,
) -> Path:
    """
    Genera el archivo Excel del corte mensual.

    Args:
        resultados:   Lista de CorteItem clasificados por el motor.
        client_name:  Nombre del cliente (aparece en encabezado).
        output_path:  Ruta destino del .xlsx.
        mes:          Mes del período (1-12). None = detecta del primero que tenga fecha.
        anio:         Año del período. None = año actual.

    Returns:
        Path del archivo generado.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError(f"Dependencia faltante para exportar Excel: {e}")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Detectar período ──────────────────────────────────────────────────────
    if mes is None or anio is None:
        for item in resultados:
            fecha_raw = getattr(item.record, "fecha_emision", "") or ""
            try:
                dt = datetime.strptime(fecha_raw.strip(), "%d/%m/%Y")
                mes  = mes  or dt.month
                anio = anio or dt.year
                break
            except ValueError:
                pass
    mes  = mes  or datetime.now().month
    anio = anio or datetime.now().year

    try:
        ultimo_dia = calendar.monthrange(anio, mes)[1]
        period_label = f"01/{mes:02d}/{anio} al {ultimo_dia:02d}/{mes:02d}/{anio}"
    except ValueError:
        period_label = f"{mes:02d}/{anio}"

    # ── Separar por categoría ─────────────────────────────────────────────────
    ingresos  = [i for i in resultados if i.categoria == CATEGORIA_INGRESOS]
    compras   = [i for i in resultados if i.categoria == CATEGORIA_COMPRAS]
    gastos    = [i for i in resultados if i.categoria == CATEGORIA_GASTOS]
    ambiguos  = [i for i in resultados if i.categoria == CATEGORIA_AMBIGUO]

    # ── Construir workbook ────────────────────────────────────────────────────
    wb = Workbook()
    # Eliminar hoja vacía por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    used_names: set[str] = set()

    sheets_to_write = [
        (CATEGORIA_INGRESOS, ingresos,  _COLS_INGRESO),
        (CATEGORIA_COMPRAS,  compras,   _COLS_EGRESO),
        (CATEGORIA_GASTOS,   gastos,    _COLS_EGRESO),
    ]
    if ambiguos:
        sheets_to_write.append((CATEGORIA_AMBIGUO, ambiguos, _COLS_EGRESO))

    for label, items, base_cols in sheets_to_write:
        if not items:
            continue  # no crear hoja vacía

        # Filtrar columnas IVA con todos ceros
        cols = _filter_iva_cols(base_cols, items)

        sheet_name = _safe_sheet_name(label, used_names)
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, items, cols, label, client_name, period_label)
        LOGGER.info("Hoja '%s': %d filas", sheet_name, len(items))

    if not wb.sheetnames:
        # Sin datos: crear hoja placeholder
        ws = wb.create_sheet("SIN DATOS")
        ws.cell(1, 1).value = "No hay facturas clasificadas para este período."

    wb.save(output_path)
    LOGGER.info("Corte Excel guardado en %s", output_path)
    return output_path


_SOCIEDADES = [
    ("SOCIEDAD DE RESPONSABILIDAD LIMITADA", "S.R.L."),
    ("SOCIEDAD ANONIMA",                     "S.A."),
    ("SOCIEDAD EN NOMBRE COLECTIVO",         "S.N.C."),
    ("EMPRESA INDIVIDUAL DE RESPONSABILIDAD LIMITADA", "E.I.R.L."),
]


def _abreviar_sociedad(name: str) -> str:
    """Reemplaza el tipo de sociedad por su sigla según Hacienda."""
    upper = name.upper()
    for forma, sigla in _SOCIEDADES:
        if forma in upper:
            return upper.replace(forma, sigla).strip().rstrip(",").strip()
    return upper


def default_filename(client_name: str, mes: int, anio: int) -> str:
    """Nombre de archivo sugerido para el corte."""
    mes_str  = month_abbr_es(mes)
    name_safe = _abreviar_sociedad(client_name or "CLIENTE").replace("/","").replace("\\","").strip()[:50]
    return f"CORTE {mes_str} {anio} - {name_safe}.xlsx"