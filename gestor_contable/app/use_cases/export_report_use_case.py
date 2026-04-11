"""Caso de uso: exportar reporte de período a Excel o CSV.

Toda la lógica de extracción de datos, transformación y escritura de archivo
vive aquí. No importa nada de customtkinter ni de gui/.
La vista llama a export_period_report() y recibe éxito (None) o excepción.
"""
from __future__ import annotations

import csv
import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from gestor_contable.core.classification_utils import (
    classify_transaction,
    get_hacienda_review_status,
)
from gestor_contable.core.iva_utils import apply_exchange_rate, compute_tax_base_rows, parse_decimal_value
from gestor_contable.core.models import FacturaRecord
from gestor_contable.core.report_paths import month_name_es

logger = logging.getLogger(__name__)


# ── Helpers internos ─────────────────────────────────────────────────────────

def _parse_date_for_filename(text: str):
    raw = (text or "").strip()
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _month_name_es(dt: datetime) -> str:
    return month_name_es(dt.month)


def _format_amount_es(number: Decimal) -> str:
    sign = "-" if number < 0 else ""
    n = abs(number)
    text = f"{n:,.2f}"
    text = text.replace(",", "_").replace(".", ",").replace("_", " ")
    return f"{sign}{text}"


def _safe_excel_sheet_name(raw_name: str, used_names: set[str]) -> str:
    """Sanitiza y hace único un nombre de hoja de Excel (máx. 31 chars)."""
    invalid_chars = {"\\", "/", "*", "?", ":", "[", "]"}
    cleaned = "".join("_" if ch in invalid_chars else ch for ch in str(raw_name or "").strip())
    cleaned = cleaned.strip("'")
    base = (cleaned or "SIN CLASIFICAR")[:31]

    candidate = base
    suffix = 1
    while candidate in used_names:
        suffix_txt = f" ({suffix})"
        allowed = 31 - len(suffix_txt)
        candidate = f"{base[:allowed]}{suffix_txt}"
        suffix += 1

    used_names.add(candidate)
    return candidate


_GASTO_PREFIX = {
    "GASTOS GENERALES": "GG",
    "GASTOS ESPECÍFICOS": "GE",
    "GASTOS ESPECIFICOS": "GE",
}
_EXCEL_AMOUNT_FORMAT = "#,##0.00"

_TOTAL_AMOUNT_COLUMNS = {
    "subtotal",
    "iva_1",
    "iva_2",
    "iva_4",
    "iva_8",
    "iva_13",
    "iva_otros",
    "impuesto_total",
    "total_comprobante",
}


def _sum_visible_amounts(sheet_df, visible_cols, total_columns):
    """Suma columnas monetarias visibles para la fila final de TOTAL."""
    totals: dict[str, Decimal] = {}
    for col in visible_cols:
        if col not in total_columns or col not in sheet_df.columns:
            continue
        total = Decimal("0")
        for value in sheet_df[col].dropna().tolist():
            parsed = parse_decimal_value(value)
            if parsed is not None:
                total += parsed
        totals[col] = total
    return totals


def _write_total_row(ws, row_idx, visible_cols, totals, total_fill, total_font, label="TOTAL"):
    """Escribe una fila final de totales con el estilo del reporte de cortes."""
    from openpyxl.styles import Alignment

    if not visible_cols:
        return row_idx

    for col_idx in range(1, len(visible_cols) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = total_fill
        cell.font = total_font

    label_cell = ws.cell(row=row_idx, column=1)
    label_cell.value = label
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(visible_cols, start=1):
        if col_name not in totals:
            continue
        total_cell = ws.cell(row=row_idx, column=col_idx)
        total_cell.value = float(totals[col_name])
        total_cell.number_format = _EXCEL_AMOUNT_FORMAT
        total_cell.font = total_font

    return row_idx


def _compact_tax_base_label(label: str) -> str:
    if label == "Base exenta":
        return "EXENTO"
    if label.startswith("Base imponible "):
        return f"BASE IMP {label.removeprefix('Base imponible ').upper()}"
    return str(label).upper()


def _tax_block_anchor_columns(visible_cols):
    amount_indices = [
        idx + 1 for idx, col_name in enumerate(visible_cols)
        if col_name in _TOTAL_AMOUNT_COLUMNS
    ]
    if len(amount_indices) >= 2:
        return amount_indices[-2], amount_indices[-1]
    if len(amount_indices) == 1:
        return max(1, amount_indices[0] - 1), amount_indices[0]
    return 1, 2


def _write_tax_base_block(ws, total_row_idx, visible_cols, totals, row_font):
    """Escribe bases imponibles como recuadro compacto a la derecha."""
    from openpyxl.styles import Alignment, Border, PatternFill, Side

    if not visible_cols:
        return total_row_idx

    base_rows = compute_tax_base_rows(totals, visible_cols)
    if not base_rows:
        return total_row_idx

    label_col_idx, amount_col_idx = _tax_block_anchor_columns(visible_cols)
    current_row = total_row_idx + 2
    last_row = current_row + len(base_rows) - 1

    outer = Side(style="medium", color="000000")
    inner = Side(style="thin", color="000000")
    solid_white = PatternFill(fill_type="solid", fgColor="FFFFFF")

    for label, amount in base_rows:
        label_cell = ws.cell(row=current_row, column=label_col_idx)
        label_cell.value = _compact_tax_base_label(label)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.font = row_font
        label_cell.fill = solid_white

        amount_cell = ws.cell(row=current_row, column=amount_col_idx)
        amount_cell.value = float(amount)
        amount_cell.number_format = _EXCEL_AMOUNT_FORMAT
        amount_cell.alignment = Alignment(horizontal="right", vertical="center")
        amount_cell.font = row_font
        amount_cell.fill = solid_white

        top_side = outer if current_row == total_row_idx + 2 else inner
        bottom_side = outer if current_row == last_row else inner
        label_cell.border = Border(left=outer, right=inner, top=top_side, bottom=bottom_side)
        amount_cell.border = Border(left=inner, right=outer, top=top_side, bottom=bottom_side)

        current_row += 1

    return current_row - 1


def _auto_fit_columns(ws, start_row: int) -> None:
    """Auto-ajusta el ancho de todas las columnas visibles."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(start_row, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=start_row, column=col_idx).column_letter].width = min(max(max_len + 3, 12), 65)


def _write_rechazados_sheet(
    ws, sheet_df, display_cols,
    numeric_columns, text_columns, date_column,
    pretty_headers, owner_name, date_from_label, date_to_label,
    title_fill, subtitle_fill, summary_fill, header_fill, credit_fill, total_fill,
    title_font, subtitle_font, summary_font, header_font, total_font,
):
    """Hoja Rechazados con bloques separados por ingresos y egresos."""
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    n_cols = max(len(display_cols), 1)

    for row in (1, 2, 3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)

    ws.cell(row=1, column=1).value = str(owner_name).upper()
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).fill = title_fill

    ws.cell(row=2, column=1).value = (
        f"REPORTE DE RECHAZADOS - Período: {date_from_label} al {date_to_label}"
    )
    ws.cell(row=2, column=1).font = subtitle_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).fill = subtitle_fill

    monto_total = Decimal("0")
    if "total_comprobante" in sheet_df.columns:
        for value in sheet_df["total_comprobante"].dropna().tolist():
            parsed = parse_decimal_value(value)
            if parsed is not None:
                monto_total += parsed

    monedas = (
        sorted({str(m).strip() for m in sheet_df["moneda"].dropna().tolist() if str(m).strip()})
        if "moneda" in sheet_df.columns else []
    )
    moneda_value = (
        "N/A" if not monedas
        else monedas[0] if len(monedas) == 1
        else "MIXTA: " + ", ".join(monedas)
    )
    generated = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.cell(row=3, column=1).value = (
        f"Total filas: {len(sheet_df)}   |   Monto Total: {_format_amount_es(monto_total)}   |   "
        f"Moneda: {moneda_value}   |   Generado: {generated}"
    )
    ws.cell(row=3, column=1).font = summary_font
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=1).fill = summary_fill

    section_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    section_font = Font(bold=True, color="111111", size=12)

    def _safe(value):
        try:
            return None if pd.isna(value) else value
        except (TypeError, ValueError):
            return value

    tx_series = (
        sheet_df["clasificacion_tx"].fillna("").astype(str).str.strip().str.lower()
        if "clasificacion_tx" in sheet_df.columns
        else pd.Series([""] * len(sheet_df), index=sheet_df.index)
    )

    def _sort_block(sdf):
        cols = [c for c in ("emisor_nombre", "fecha_emision") if c in sdf.columns]
        return sdf.sort_values(cols) if cols else sdf

    block_defs = [
        ("Rechazados Ingresos", _sort_block(sheet_df.loc[tx_series.eq("ingreso")].copy())),
        ("Rechazados Egresos",  _sort_block(sheet_df.loc[~tx_series.eq("ingreso")].copy())),
    ]

    tipo_idx = (display_cols.index("tipo_documento") + 1) if "tipo_documento" in display_cols else None
    current_row = 5

    for block_label, block_df in block_defs:
        if block_df.empty:
            continue

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        section_cell = ws.cell(row=current_row, column=1)
        section_cell.value = block_label.upper()
        section_cell.font = section_font
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        section_cell.fill = section_fill
        current_row += 1

        for col_idx, col_name in enumerate(display_cols, start=1):
            header_cell = ws.cell(row=current_row, column=col_idx)
            header_cell.value = pretty_headers.get(col_name, col_name.replace("_", " ").title())
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.fill = header_fill

        current_row += 1

        for _, row_data in block_df.iterrows():
            for col_idx, col_name in enumerate(display_cols, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                value = _safe(row_data.get(col_name)) if col_name in row_data.index else None
                cell.value = value
                if col_name in text_columns:
                    cell.number_format = "@"
                    cell.value = "" if cell.value is None else str(cell.value)
                elif col_name == date_column and cell.value is not None:
                    cell.number_format = "dd/mm/yyyy"
                elif col_name in numeric_columns and cell.value is not None:
                    cell.number_format = _EXCEL_AMOUNT_FORMAT
                    if isinstance(cell.value, Decimal):
                        cell.value = float(cell.value)

            if tipo_idx is not None and ws.cell(row=current_row, column=tipo_idx).value == "Nota de Crédito":
                for col in range(1, n_cols + 1):
                    ws.cell(row=current_row, column=col).fill = credit_fill

            current_row += 1

        block_totals = _sum_visible_amounts(block_df, display_cols, _TOTAL_AMOUNT_COLUMNS)
        _write_total_row(
            ws,
            current_row,
            display_cols,
            block_totals,
            total_fill,
            total_font,
            label=f"SUBTOTAL {block_label.upper()}",
        )
        current_row += 2

    overall_totals = _sum_visible_amounts(sheet_df, display_cols, _TOTAL_AMOUNT_COLUMNS)
    _write_total_row(ws, current_row, display_cols, overall_totals, total_fill, total_font, label="TOTAL")
    current_row = _write_tax_base_block(ws, current_row, display_cols, overall_totals, total_font)

    _auto_fit_columns(ws, 6)
    ws.freeze_panes = ws["A7"]


def _write_gasto_grouped(
    ws, sheet_df, display_cols,
    numeric_columns, text_columns, date_column,
    pretty_headers, owner_name, sheet_name, date_from_label, date_to_label,
    title_fill, subtitle_fill, summary_fill, header_fill, credit_fill, total_fill,
    title_font, subtitle_font, summary_font, header_font, total_font,
):
    """Hoja Gasto -- agrupación por (subtipo, nombre_cuenta).

    Layout por grupo:
        [filas de datos -- sin color]
        [fila subtotal: sumas numéricas + label "GG/GE / NOMBRE" en última col]  <- fill azul
        [fila vacía]

    display_cols: columnas a mostrar.  subtipo/nombre_cuenta se leen del DataFrame
    para agrupar aunque no aparezcan en display_cols.
    """
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    n_cols = len(display_cols)

    # Filas 1-3: título / subtítulo / resumen
    for row in (1, 2, 3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)

    ws.cell(row=1, column=1).value = str(owner_name).upper()
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).fill = title_fill

    ws.cell(row=2, column=1).value = (
        f"REPORTE DE {sheet_name.upper()} - Período: {date_from_label} al {date_to_label}"
    )
    ws.cell(row=2, column=1).font = subtitle_font
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).fill = subtitle_fill

    monto_total = Decimal("0")
    if "total_comprobante" in sheet_df.columns:
        for v in sheet_df["total_comprobante"].dropna():
            try:
                monto_total += Decimal(str(v))
            except Exception:
                pass

    monedas = (
        sorted({str(m).strip() for m in sheet_df["moneda"].dropna() if str(m).strip()})
        if "moneda" in sheet_df.columns else []
    )
    moneda_value = (
        "N/A" if not monedas
        else monedas[0] if len(monedas) == 1
        else "MIXTA: " + ", ".join(monedas)
    )

    ws.cell(row=3, column=1).value = (
        f"Total filas: {len(sheet_df)}   |   Monto Total: {_format_amount_es(monto_total)}   |   "
        f"Moneda: {moneda_value}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws.cell(row=3, column=1).font = summary_font
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=1).fill = summary_fill

    # Fila 5: encabezados de columna
    for col_idx, col_name in enumerate(display_cols, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = pretty_headers.get(col_name, col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tipo_col_idx = (display_cols.index("tipo_documento") + 1) if "tipo_documento" in display_cols else None
    numeric_display = [c for c in display_cols if c in numeric_columns and c in display_cols]

    subtotal_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    subtotal_font = Font(bold=True)

    group_cols = [c for c in ("subtipo", "nombre_cuenta") if c in sheet_df.columns]
    sort_cols  = group_cols + [c for c in ("emisor_nombre", "fecha_emision") if c in sheet_df.columns]
    sorted_df  = sheet_df.sort_values(sort_cols) if sort_cols else sheet_df

    def _safe(v):
        try:
            return None if pd.isna(v) else v
        except (TypeError, ValueError):
            return v

    current_row = 6

    if group_cols:
        for group_keys, group_df in sorted_df.groupby(group_cols, sort=False):
            if isinstance(group_keys, tuple) and len(group_keys) == 2:
                subtipo_val = str(group_keys[0]).strip().upper()
                cuenta_val  = str(group_keys[1]).strip()
            else:
                subtipo_val = ""
                cuenta_val  = str(group_keys).strip()

            group_sums: dict[str, Decimal] = {c: Decimal("0") for c in numeric_display}

            for _, row_data in group_df.iterrows():
                for col_idx, col_name in enumerate(display_cols, start=1):
                    val = _safe(row_data.get(col_name)) if col_name in row_data.index else None
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = val
                    if col_name in text_columns:
                        cell.number_format = "@"
                        cell.value = "" if cell.value is None else str(cell.value)
                    elif col_name == date_column and cell.value is not None:
                        cell.number_format = "dd/mm/yyyy"
                    elif col_name in numeric_columns and cell.value is not None:
                        cell.number_format = _EXCEL_AMOUNT_FORMAT
                        if isinstance(cell.value, Decimal):
                            cell.value = float(cell.value)

                if tipo_col_idx and ws.cell(row=current_row, column=tipo_col_idx).value == "Nota de Crédito":
                    for c in range(1, n_cols + 1):
                        ws.cell(row=current_row, column=c).fill = credit_fill

                for col_name in numeric_display:
                    tv = _safe(row_data.get(col_name)) if col_name in row_data.index else None
                    try:
                        if tv is not None:
                            group_sums[col_name] += Decimal(str(tv))
                    except Exception:
                        pass

                current_row += 1

            for col_idx in range(1, n_cols + 1):
                ws.cell(row=current_row, column=col_idx).fill = subtotal_fill

            for col_name in numeric_display:
                ci = display_cols.index(col_name) + 1
                tc = ws.cell(row=current_row, column=ci)
                tc.value         = float(group_sums[col_name])
                tc.number_format = _EXCEL_AMOUNT_FORMAT
                tc.font          = subtotal_font

            prefix      = _GASTO_PREFIX.get(subtipo_val, "")
            cuenta_label = cuenta_val.upper() if cuenta_val else subtipo_val
            label        = f"{prefix} / {cuenta_label}" if prefix else cuenta_label
            lbl = ws.cell(row=current_row, column=n_cols)
            lbl.value = label
            lbl.font = subtotal_font
            label_alignment = "left" if sheet_name == "Gastos" and label.startswith("GG /") else "right"
            lbl.alignment = Alignment(horizontal=label_alignment, vertical="center")

            current_row += 2  # subtotal + fila vacía

    else:
        for _, row_data in sorted_df.iterrows():
            for col_idx, col_name in enumerate(display_cols, start=1):
                val  = _safe(row_data[col_name])
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = val
                if col_name in text_columns:
                    cell.number_format = "@"
                    cell.value = "" if cell.value is None else str(cell.value)
                elif col_name == date_column and cell.value is not None:
                    cell.number_format = "dd/mm/yyyy"
                elif col_name in numeric_columns and cell.value is not None:
                    cell.number_format = _EXCEL_AMOUNT_FORMAT
            current_row += 1

    overall_totals = _sum_visible_amounts(sheet_df, display_cols, _TOTAL_AMOUNT_COLUMNS)
    _write_total_row(ws, current_row, display_cols, overall_totals, total_fill, total_font, label="TOTAL")
    current_row = _write_tax_base_block(ws, current_row, display_cols, overall_totals, total_font)

    _auto_fit_columns(ws, 5)

    ws.freeze_panes = ws["A6"]


# ── API pública ──────────────────────────────────────────────────────────────

def default_export_filename(
    client_name: str,
    from_date: str,
    to_date: str,
    *,
    mes: int | None = None,
    anio: int | None = None,
) -> str:
    """Nombre de archivo sugerido para el reporte de exportación."""
    if mes is None or anio is None:
        base_dt = _parse_date_for_filename(from_date) or _parse_date_for_filename(to_date) or datetime.now()
        anio = base_dt.year
        mes = base_dt.month

    year = f"{int(anio):04d}"
    month_txt = month_name_es(int(mes))
    client_clean = (str(client_name or "REPORTE")
                    .replace("/", " ")
                    .replace("\\", " ")
                    .strip())
    if len(client_clean) > 42:
        client_clean = client_clean[:42].strip()
    return f"PF-{year} - {client_clean} - REPORTE - {month_txt}.xlsx"


def export_period_report(
    records: list[FacturaRecord],
    db_records: dict,
    client_cedula: str,
    target_path: Path,
    owner_name: str,
    date_from_label: str,
    date_to_label: str,
) -> dict[str, object]:
    """Exporta los registros del período a Excel (.xlsx) o CSV.

    Args:
        records: lista de FacturaRecord ya filtrados por período y sin omitidos.
        db_records: {clave: {estado, categoria, subtipo, nombre_cuenta}} de la BD.
        client_cedula: cédula del receptor (cliente activo) para classify_transaction.
        target_path: ruta destino (extensión determina formato: .xlsx o .csv).
        owner_name: nombre del cliente para encabezados del Excel.
        date_from_label: fecha de inicio en formato "dd/mm/yyyy" (solo para encabezados).
        date_to_label: fecha de fin en formato "dd/mm/yyyy" (solo para encabezados).

    Returns:
        Métricas de cobertura del reporte exportado.

    Raises:
        ValueError: si no hay registros o la extensión no es soportada.
        IOError/OSError: si no se puede escribir el archivo.
    """
    rows: list[dict] = []
    for r in records:
        meta = db_records.get(r.clave, {}) if db_records else {}
        estado = meta.get("estado") or r.estado
        rows.append(
            {
                "clave_numerica": r.clave,
                "tipo_documento": r.tipo_documento,
                "fecha_emision": r.fecha_emision,
                "consecutivo": r.consecutivo,
                "emisor_nombre": r.emisor_nombre,
                "emisor_cedula": r.emisor_cedula,
                "receptor_nombre": r.receptor_nombre,
                "receptor_cedula": r.receptor_cedula,
                "moneda": r.moneda,
                "tipo_cambio": r.tipo_cambio,
                "subtotal": r.subtotal,
                "iva_1": r.iva_1,
                "iva_2": r.iva_2,
                "iva_4": r.iva_4,
                "iva_8": r.iva_8,
                "iva_13": r.iva_13,
                "iva_otros": r.iva_otros,
                "impuesto_total": r.impuesto_total,
                "total_comprobante": r.total_comprobante,
                "estado_hacienda": r.estado_hacienda,
                "detalle_estado_hacienda": r.detalle_estado_hacienda,
                "hacienda_review_status": get_hacienda_review_status(r) or "",
                "categoria": str(meta.get("categoria") or ""),
                "subtipo": str(meta.get("subtipo") or ""),
                "nombre_cuenta": str(meta.get("nombre_cuenta") or ""),
                "estado": estado,
                "clasificacion_tx": classify_transaction(r, client_cedula),
            }
        )

    export_columns = [
        "clave_numerica",
        "tipo_documento",
        "fecha_emision",
        "consecutivo",
        "emisor_nombre",
        "emisor_cedula",
        "receptor_nombre",
        "receptor_cedula",
        "moneda",
        "tipo_cambio",
        "subtotal",
        "iva_1",
        "iva_2",
        "iva_4",
        "iva_8",
        "iva_13",
        "iva_otros",
        "impuesto_total",
        "total_comprobante",
        "estado_hacienda",
        "detalle_estado_hacienda",
        "categoria",
        "subtipo",
        "nombre_cuenta",
        "estado",
    ]

    _HIDDEN = {"clave_numerica", "subtipo", "nombre_cuenta", "estado", "categoria", "detalle_estado_hacienda", "clasificacion_tx"}
    display_columns = [c for c in export_columns if c not in _HIDDEN]

    numeric_columns = {
        "subtotal", "tipo_cambio",
        "iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros",
        "impuesto_total", "total_comprobante",
    }
    text_columns = {"clave_numerica", "consecutivo", "emisor_cedula", "receptor_cedula"}
    date_column = "fecha_emision"

    pretty_headers = {
        "clave_numerica": "Clave",
        "tipo_documento": "Tipo documento",
        "fecha_emision": "Fecha emisión",
        "consecutivo": "Consecutivo",
        "emisor_nombre": "Emisor",
        "emisor_cedula": "Cédula emisor",
        "receptor_nombre": "Receptor",
        "receptor_cedula": "Cédula receptor",
        "moneda": "Moneda",
        "tipo_cambio": "Tipo cambio",
        "subtotal": "Subtotal",
        "iva_1": "IVA 1%",
        "iva_2": "IVA 2%",
        "iva_4": "IVA 4%",
        "iva_8": "IVA 8%",
        "iva_13": "IVA 13%",
        "impuesto_total": "Impuesto total",
        "total_comprobante": "Total comprobante",
        "estado_hacienda": "Estado Hacienda",
        "detalle_estado_hacienda": "Detalle Estado Hacienda",
        "categoria": "Categoría",
        "subtipo": "Subtipo",
        "nombre_cuenta": "Cuenta",
        "estado": "Estado App 3",
        "clasificacion_tx": "Clasificación Tx",
    }

    target = str(target_path)
    coverage_info: dict[str, object] = {
        "unassigned_count": 0,
        "unassigned_keys": [],
    }

    if target.lower().endswith(".csv"):
        with open(target, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=export_columns)
            writer.writeheader()
            writer.writerows([{col: row.get(col, "") for col in export_columns} for row in rows])
        return coverage_info

    # Excel
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill

    df_all = pd.DataFrame(rows)
    df = df_all[[col for col in export_columns if col in df_all.columns]].copy()

    for col in text_columns:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(" ", "", regex=False).str.replace(",", ".", regex=False),
                errors="coerce",
            )

    amounts_to_convert = {
        "subtotal", "iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros",
        "impuesto_total", "total_comprobante"
    }

    if "moneda" in df_all.columns and "tipo_cambio" in df_all.columns:
        for idx in df.index:
            moneda_str = str(df_all.loc[idx, "moneda"] or "").strip().upper()
            if moneda_str and moneda_str != "CRC":
                tc = parse_decimal_value(df_all.loc[idx, "tipo_cambio"])
                if tc and tc > Decimal("0"):
                    for col in amounts_to_convert:
                        if col in df.columns and pd.notna(df.loc[idx, col]):
                            amount_val = df.loc[idx, col]
                            if amount_val and isinstance(amount_val, (int, float)):
                                amount = Decimal(str(amount_val))
                                converted = apply_exchange_rate(amount, moneda_str, tc)
                                df.loc[idx, col] = float(converted)

    df["clasificacion_tx"] = df_all["clasificacion_tx"].fillna("").astype(str).str.strip().str.lower()

    if date_column in df.columns:
        df[date_column] = pd.to_datetime(df[date_column], format="%d/%m/%Y", errors="coerce")

    clasificacion_tx = df["clasificacion_tx"]

    categoria_upper = df_all["categoria"].fillna("").astype(str).str.strip().str.upper()
    hacienda_review_col = df_all["hacienda_review_status"].fillna("").astype(str).str.strip().str.lower()
    mask_rechazados = hacienda_review_col.eq("rechazada")
    mask_sin_respuesta = hacienda_review_col.eq("sin_respuesta")
    mask_exportable = ~mask_rechazados & ~mask_sin_respuesta

    assigned_mask = pd.Series(False, index=df.index)

    mask_ventas = clasificacion_tx.eq("ingreso") & mask_exportable & ~assigned_mask
    assigned_mask = assigned_mask | mask_ventas

    mask_compras = categoria_upper.eq("COMPRAS") & mask_exportable & ~assigned_mask
    assigned_mask = assigned_mask | mask_compras

    mask_gasto = categoria_upper.eq("GASTOS") & mask_exportable & ~assigned_mask
    assigned_mask = assigned_mask | mask_gasto

    mask_activos = categoria_upper.eq("ACTIVO") & mask_exportable & ~assigned_mask
    assigned_mask = assigned_mask | mask_activos

    mask_ognd = categoria_upper.eq("OGND") & mask_exportable & ~assigned_mask
    assigned_mask = assigned_mask | mask_ognd

    mask_sin_receptor = (
        categoria_upper.eq("SIN_RECEPTOR") | clasificacion_tx.eq("sin_receptor")
    ) & mask_exportable & ~assigned_mask
    assigned_mask = assigned_mask | mask_sin_receptor

    mask_pendiente = clasificacion_tx.eq("egreso") & mask_exportable & ~assigned_mask
    assigned_mask = assigned_mask | mask_pendiente

    mask_fuera_reporte = mask_exportable & ~assigned_mask
    if bool(mask_fuera_reporte.any()):
        unassigned_keys = (
            df_all.loc[mask_fuera_reporte, "clave_numerica"]
            .fillna("")
            .astype(str)
            .tolist()
        )
        coverage_info = {
            "unassigned_count": len(unassigned_keys),
            "unassigned_keys": unassigned_keys,
        }
        logger.warning(
            "Exportación con %d registro(s) fuera de hoja: %s",
            len(unassigned_keys),
            ", ".join(unassigned_keys[:10]),
        )

    used_names: set[str] = set()
    sheet_map: dict[str, pd.DataFrame] = {}
    for label, mask in [
        ("Ingresos", mask_ventas),
        ("Compras", mask_compras),
        ("Gastos", mask_gasto),
        ("OGND", mask_ognd),
        ("Activos", mask_activos),
        ("Pendientes", mask_pendiente),
        ("Sin Receptor", mask_sin_receptor),
        ("Rechazados", mask_rechazados),
        ("Sin Respuesta", mask_sin_respuesta),
        ("Fuera Reporte", mask_fuera_reporte),
    ]:
        chunk = df.loc[mask]
        if not chunk.empty:
            sheet_map[_safe_excel_sheet_name(label, used_names)] = chunk.copy()

    if not sheet_map:
        sheet_map[_safe_excel_sheet_name("Reporte", used_names)] = df.copy()

    title_fill = PatternFill(fill_type="solid", fgColor="0B2B66")
    subtitle_fill = PatternFill(fill_type="solid", fgColor="7F7F7F")
    summary_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    credit_fill = PatternFill(fill_type="solid", fgColor="DAF2D0")
    total_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    title_font = Font(bold=True, color="FFFFFF", size=22)
    subtitle_font = Font(bold=True, color="FFFFFF", size=14)
    summary_font = Font(bold=False, color="111111", size=12)
    header_font = Font(bold=True)
    total_font = Font(bold=True)

    def _filter_iva_cols(cols, sdf):
        """Elimina columnas IVA cuyo valor es todo-cero en el DataFrame dado."""
        IVA_COLS = {"iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros"}
        ZERO_VALUES = {"", "0", "0.0", "0,00", "0.00", "nan", "none", "null"}
        result = []
        for col in cols:
            if col not in IVA_COLS:
                result.append(col)
            else:
                col_values = sdf[col].astype(str).str.strip().str.lower()
                if col_values.loc[~col_values.isin(ZERO_VALUES)].any():
                    result.append(col)
        return result

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        if "Sheet" in writer.book.sheetnames:
            del writer.book["Sheet"]

        for sheet_name, sheet_df in sheet_map.items():
            # Hoja Gastos: layout agrupado especial
            if sheet_name == "Gastos":
                ws = writer.book.create_sheet(title=sheet_name)
                writer.sheets[sheet_name] = ws
                gasto_base = [c for c in display_columns if c not in {"receptor_nombre", "receptor_cedula"} and c in sheet_df.columns]
                gasto_cols = _filter_iva_cols(gasto_base, sheet_df)
                _write_gasto_grouped(
                    ws, sheet_df, gasto_cols,
                    numeric_columns, text_columns, date_column,
                    pretty_headers, owner_name, sheet_name,
                    date_from_label, date_to_label,
                    title_fill, subtitle_fill, summary_fill,
                    header_fill, credit_fill, total_fill,
                    title_font, subtitle_font, summary_font, header_font, total_font,
                )
                continue

            # Hoja OGND: layout agrupado especial
            if sheet_name == "OGND":
                ws = writer.book.create_sheet(title=sheet_name)
                writer.sheets[sheet_name] = ws
                ognd_cols = _filter_iva_cols([c for c in display_columns if c in sheet_df.columns], sheet_df)
                _write_gasto_grouped(
                    ws, sheet_df, ognd_cols,
                    numeric_columns, text_columns, date_column,
                    pretty_headers, owner_name, sheet_name,
                    date_from_label, date_to_label,
                    title_fill, subtitle_fill, summary_fill,
                    header_fill, credit_fill, total_fill,
                    title_font, subtitle_font, summary_font, header_font, total_font,
                )
                continue

            # Hoja Rechazados: incluye detalle_estado_hacienda
            if sheet_name == "Rechazados":
                rechazados_hidden = {"subtipo", "nombre_cuenta", "estado", "categoria", "receptor_nombre", "receptor_cedula"}
                rechazados_cols = [c for c in export_columns
                                   if c not in rechazados_hidden and c in sheet_df.columns]
                IVA_COLS = {"iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros"}
                ZERO_VALUES = {"", "0", "0.0", "0,00", "0.00", "nan", "none", "null"}
                visible_rechazados = []
                for col in rechazados_cols:
                    if col not in IVA_COLS:
                        visible_rechazados.append(col)
                    else:
                        col_values = sheet_df[col].astype(str).str.strip().str.lower()
                        has_nonzero = col_values.loc[~col_values.isin(ZERO_VALUES)].any()
                        if has_nonzero:
                            visible_rechazados.append(col)
                _write_rechazados_sheet(
                    writer.book.create_sheet(title=sheet_name),
                    sheet_df,
                    visible_rechazados,
                    numeric_columns,
                    text_columns,
                    date_column,
                    pretty_headers,
                    owner_name,
                    date_from_label,
                    date_to_label,
                    title_fill,
                    subtitle_fill,
                    summary_fill,
                    header_fill,
                    credit_fill,
                    total_fill,
                    title_font,
                    subtitle_font,
                    summary_font,
                    header_font,
                    total_font,
                )
                writer.sheets[sheet_name] = writer.book[sheet_name]
                continue

            # Hojas normales
            if sheet_name == "Fuera Reporte":
                visible_cols_base = [
                    c for c in [
                        "clave_numerica",
                        "tipo_documento",
                        "fecha_emision",
                        "emisor_nombre",
                        "emisor_cedula",
                        "receptor_nombre",
                        "receptor_cedula",
                        "categoria",
                        "estado",
                        "clasificacion_tx",
                        "estado_hacienda",
                        "detalle_estado_hacienda",
                        "total_comprobante",
                    ]
                    if c in sheet_df.columns
                ]
            else:
                exclude_receptor = sheet_name != "Sin Receptor"
                visible_cols_base = [
                    c for c in display_columns
                    if c in sheet_df.columns and not (exclude_receptor and c in {"receptor_nombre", "receptor_cedula"})
                ]

            visible_cols_filtered = _filter_iva_cols(visible_cols_base, sheet_df)

            _sort_simple = [c for c in ("emisor_nombre", "fecha_emision") if c in sheet_df.columns]
            if _sort_simple:
                sheet_df = sheet_df.sort_values(_sort_simple)
            display_df = sheet_df[visible_cols_filtered].rename(
                columns={col: pretty_headers.get(col, col.replace("_", " ").title()) for col in visible_cols_filtered}
            )
            display_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=4)
            ws = writer.sheets[sheet_name]

            max_col = ws.max_column if ws.max_column > 0 else 1
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)

            title_cell = ws.cell(row=1, column=1)
            title_cell.value = str(owner_name).upper()
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = title_fill

            subtitle_cell = ws.cell(row=2, column=1)
            subtitle_cell.value = f"REPORTE DE {sheet_name.upper()} - Período: {date_from_label} al {date_to_label}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
            subtitle_cell.fill = subtitle_fill

            monto_total = Decimal("0")
            if "total_comprobante" in sheet_df.columns:
                valid_amounts = []
                for value in sheet_df["total_comprobante"].dropna().tolist():
                    try:
                        valid_amounts.append(Decimal(str(value)))
                    except Exception:
                        continue
                if valid_amounts:
                    monto_total = sum(valid_amounts, Decimal("0"))

            monedas = (
                sorted({str(m).strip() for m in sheet_df["moneda"].dropna().tolist() if str(m).strip()})
                if "moneda" in sheet_df.columns
                else []
            )
            moneda_value = (
                "N/A" if not monedas
                else monedas[0] if len(monedas) == 1
                else "MIXTA: " + ", ".join(monedas)
            )
            generated = datetime.now().strftime("%d/%m/%Y %H:%M")

            summary_cell = ws.cell(row=3, column=1)
            summary_cell.value = (
                f"Total filas: {len(sheet_df)}   |   Monto Total: {_format_amount_es(monto_total)}   |   "
                f"Moneda: {moneda_value}   |   Generado: {generated}"
            )
            summary_cell.font = summary_font
            summary_cell.alignment = Alignment(horizontal="center", vertical="center")
            summary_cell.fill = summary_fill

            header_row = 5
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            tipo_idx = (
                visible_cols_filtered.index("tipo_documento") + 1
                if "tipo_documento" in visible_cols_filtered else None
            )

            for col_idx, col_name in enumerate(visible_cols_filtered, start=1):
                for row_idx in range(header_row + 1, len(sheet_df) + header_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_name in text_columns:
                        cell.number_format = "@"
                        cell.value = "" if cell.value is None else str(cell.value)
                    elif col_name == date_column and cell.value is not None:
                        cell.number_format = "dd/mm/yyyy"
                    elif col_name in numeric_columns and cell.value is not None:
                        cell.number_format = _EXCEL_AMOUNT_FORMAT
                        if isinstance(cell.value, Decimal):
                            cell.value = float(cell.value)

            if tipo_idx is not None:
                for row_idx in range(header_row + 1, len(sheet_df) + header_row + 1):
                    if ws.cell(row=row_idx, column=tipo_idx).value == "Nota de Crédito":
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col).fill = credit_fill

            totals = _sum_visible_amounts(sheet_df, visible_cols_filtered, _TOTAL_AMOUNT_COLUMNS)
            total_row = header_row + len(sheet_df) + 1
            _write_total_row(ws, total_row, visible_cols_filtered, totals, total_fill, total_font, label="TOTAL")
            _write_tax_base_block(ws, total_row, visible_cols_filtered, totals, total_font)

            _auto_fit_columns(ws, header_row)

            ws.freeze_panes = ws["A6"]

    return coverage_info
