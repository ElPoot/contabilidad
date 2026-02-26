"""Capa de interfaz gráfica (CustomTkinter) para la aplicación."""
from __future__ import annotations

import json
import threading
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from queue import Queue
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import customtkinter as ctk
except ModuleNotFoundError:  # Permite usar backend aunque falte GUI.
    ctk = None

from .config import LOGGER, resolve_default_data_dir
from .xml_manager import CRXMLManager

class MainApp(ctk.CTk if ctk else object):
    """Interfaz principal optimizada para miles de registros."""
    PAGE_SIZE = 500
    def __init__(self) -> None:
        if ctk is None:
            raise RuntimeError("customtkinter no está instalado. Ejecuta: pip install -r requirements.txt")
        super().__init__()
        self.title("Procesador XML Facturación Electrónica CR")
        self.geometry("1450x850")
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")
        self.manager = CRXMLManager()
        self.df_full = pd.DataFrame()
        self.df_filtered = pd.DataFrame()
        self.column_vars: dict[str, ctk.BooleanVar] = {}
        self.column_aliases: dict[str, str] = {}
        self.section_collapsed: dict[str, bool] = {}
        self._column_default_selected: set[str] = set()
        self.column_mode_var = ctk.StringVar(value="simple")
        self.current_page = 0
        self._loading_queue: Queue[tuple[str, Any]] = Queue()
        self._loading_thread: threading.Thread | None = None
        self.template_path = Path.home() / ".facturacion_column_templates.json"
        self.last_audit_report: dict[str, Any] = {}
        self._tooltip_window: ctk.CTkToplevel | None = None
        self._tooltip_after_id: str | None = None
        self._filter_after_id: str | None = None
        self._build_layout()
    def _build_layout(self) -> None:
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.left_panel = ctk.CTkFrame(self, width=340)
        self.left_panel.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(10, 5), pady=10)
        self.left_panel.grid_propagate(False)
        ctk.CTkLabel(
            self.left_panel,
            text="Selector de columnas",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).pack(anchor="w", padx=10, pady=(10, 6))

        self.mode_switch = ctk.CTkSegmentedButton(
            self.left_panel,
            values=["simple", "avanzado"],
            variable=self.column_mode_var,
            command=lambda _value: self.refresh_column_selector(),
        )
        self.mode_switch.pack(fill="x", padx=10, pady=(0, 6))

        self.selector_actions = ctk.CTkFrame(self.left_panel)
        self.selector_actions.pack(fill="x", padx=10, pady=(0, 6))
        self.selector_actions.grid_columnconfigure((0, 1), weight=1)

        ctk.CTkButton(self.selector_actions, text="Básicas", height=28, command=self.select_basic_columns).grid(row=0, column=0, padx=2, pady=2, sticky="ew")
        ctk.CTkButton(self.selector_actions, text="Todo", height=28, command=self.select_all_visible_columns).grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        ctk.CTkButton(self.selector_actions, text="Ninguna", height=28, command=self.clear_visible_columns).grid(row=1, column=0, padx=2, pady=2, sticky="ew")
        ctk.CTkButton(self.selector_actions, text="Guardar", height=28, command=self.save_column_template).grid(row=1, column=1, padx=2, pady=2, sticky="ew")
        ctk.CTkButton(self.selector_actions, text="Cargar", height=28, command=self.load_column_template).grid(row=2, column=0, columnspan=2, padx=2, pady=2, sticky="ew")

        self.columns_scroll = ctk.CTkScrollableFrame(self.left_panel)
        self.columns_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.top_panel = ctk.CTkFrame(self)
        self.top_panel.grid(row=0, column=1, sticky="ew", padx=(5, 10), pady=(10, 5))
        self.top_panel.grid_columnconfigure(12, weight=1)
        self.load_button = ctk.CTkButton(
            self.top_panel,
            text="Cargar carpeta XML",
            command=self.load_folder,
        )
        self.load_button.grid(row=0, column=0, padx=8, pady=8)
        self.export_button = ctk.CTkButton(
            self.top_panel,
            text="Exportar Excel",
            command=self.export_excel,
        )
        self.export_button.grid(row=0, column=1, padx=8, pady=8)
        self.audit_button = ctk.CTkButton(
            self.top_panel,
            text="Ver Reporte",
            command=self.show_audit_report_window,
            state="disabled",
        )
        self.audit_button.grid(row=0, column=2, padx=8, pady=8)
        ctk.CTkLabel(self.top_panel, text="Cédula cliente:").grid(row=0, column=3, padx=(12, 4), pady=8)
        self.search_tax_id_var = ctk.StringVar()
        tax_id_entry = ctk.CTkEntry(self.top_panel, textvariable=self.search_tax_id_var, width=190)
        tax_id_entry.grid(row=0, column=4, padx=4, pady=8)
        tax_id_entry.bind("<KeyRelease>", lambda _event: self.schedule_apply_filters())
        ctk.CTkLabel(self.top_panel, text="Desde (DD/MM/AAAA):").grid(row=0, column=5, padx=(12, 4), pady=8)
        self.search_date_from_var = ctk.StringVar()
        date_from_entry = ctk.CTkEntry(self.top_panel, textvariable=self.search_date_from_var, width=120)
        date_from_entry.grid(row=0, column=6, padx=4, pady=8)
        date_from_entry.bind("<KeyRelease>", lambda _event: self.schedule_apply_filters())
        ctk.CTkLabel(self.top_panel, text="Hasta (DD/MM/AAAA):").grid(row=1, column=5, padx=(12, 4), pady=4)
        self.search_date_to_var = ctk.StringVar()
        date_to_entry = ctk.CTkEntry(self.top_panel, textvariable=self.search_date_to_var, width=120)
        date_to_entry.grid(row=1, column=6, padx=4, pady=4)
        date_to_entry.bind("<KeyRelease>", lambda _event: self.schedule_apply_filters())
        self.prev_button = ctk.CTkButton(self.top_panel, text="◀", width=36, command=self.prev_page)
        self.prev_button.grid(row=0, column=7, padx=(12, 2), pady=8)
        self.next_button = ctk.CTkButton(self.top_panel, text="▶", width=36, command=self.next_page)
        self.next_button.grid(row=0, column=8, padx=2, pady=8)
        self.page_label = ctk.CTkLabel(self.top_panel, text="Página 0/0")
        self.page_label.grid(row=0, column=9, padx=(6, 8), pady=8)
        self.status_label = ctk.CTkLabel(self.top_panel, text="Listo", anchor="w")
        self.status_label.grid(row=0, column=12, padx=4, pady=8, sticky="ew")
        self.center_panel = ctk.CTkFrame(self)
        self.center_panel.grid(row=1, column=1, sticky="nsew", padx=(5, 10), pady=(5, 10))
        self.center_panel.grid_columnconfigure(0, weight=1)
        self.center_panel.grid_rowconfigure(0, weight=1)
        self.tree = ttk.Treeview(self.center_panel, show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(self.center_panel, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(self.center_panel, orient="horizontal", command=self.tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
    def load_folder(self) -> None:
        default_data_dir = resolve_default_data_dir()
        folder = filedialog.askdirectory(
            title="Selecciona carpeta con XML",
            initialdir=str(default_data_dir),
        )
        if not folder:
            return
        if self._loading_thread and self._loading_thread.is_alive():
            messagebox.showinfo("Carga en progreso", "Espera a que termine la carga actual.")
            return
        self.status_label.configure(text="Cargando archivos...")
        self.load_button.configure(state="disabled")
        def worker() -> None:
            try:
                df, audit_report = self.manager.load_xml_folder(folder)
                self._loading_queue.put(("ok", (df, audit_report)))
            except Exception as exc:  # noqa: BLE001
                self._loading_queue.put(("error", str(exc)))
        self._loading_thread = threading.Thread(target=worker, daemon=True)
        self._loading_thread.start()
        self.after(150, self._poll_loading_queue)
    def _poll_loading_queue(self) -> None:
        if self._loading_queue.empty():
            if self._loading_thread and self._loading_thread.is_alive():
                self.after(150, self._poll_loading_queue)
            return
        status, payload = self._loading_queue.get()
        self.load_button.configure(state="normal")
        if status == "error":
            self.status_label.configure(text="Error en carga")
            messagebox.showerror("Error", f"No se pudo cargar XML: {payload}")
            return
        df, audit_report = payload
        self.last_audit_report = audit_report
        self.audit_button.configure(state="normal")
        if df.empty:
            self.status_label.configure(text="Sin datos")
            messagebox.showinfo("Sin datos", "No se encontraron XML en la carpeta/subcarpetas.")
            return
        self.df_full = df.fillna("")
        emisor_series = self.df_full["emisor_cedula"] if "emisor_cedula" in self.df_full.columns else pd.Series("", index=self.df_full.index)
        receptor_series = self.df_full["receptor_cedula"] if "receptor_cedula" in self.df_full.columns else pd.Series("", index=self.df_full.index)
        fecha_series = self.df_full["fecha_emision"] if "fecha_emision" in self.df_full.columns else pd.Series("", index=self.df_full.index)
        self.df_full["_emisor_cedula_norm"] = emisor_series.astype(str).map(self.manager.normalize_identification)
        self.df_full["_receptor_cedula_norm"] = receptor_series.astype(str).map(self.manager.normalize_identification)
        self.df_full["_fecha_emision_dt"] = pd.to_datetime(fecha_series, format="%d/%m/%Y", errors="coerce")
        self.df_filtered = self.df_full
        self.current_page = 0
        self.refresh_column_selector()
        self.refresh_table()
        total = int(audit_report.get("total_files_found", 0))
        ok = int(audit_report.get("successfully_processed", 0))
        failed_count = len(audit_report.get("failed_files", []))
        duplicate_count = len(audit_report.get("duplicate_files", []))
        omitted = failed_count + duplicate_count
        seconds = float(audit_report.get("processing_time_seconds", 0.0))
        status_text = f"✓ {ok}/{total} procesados ({omitted} omitidos/fallidos) · {seconds:.1f}s"
        if omitted > 0:
            self.status_label.configure(text=status_text, text_color=("#B45309", "#FBBF24"))
        else:
            self.status_label.configure(text=status_text, text_color=("#047857", "#34D399"))
    def refresh_column_selector(self) -> None:
        """Renderiza selector de columnas con alias, secciones y modo simple/avanzado."""
        for widget in self.columns_scroll.winfo_children():
            widget.destroy()

        priority = self.get_priority_columns()
        all_cols = [col for col in self.df_full.columns if not str(col).startswith("_")]
        ordered_cols = [c for c in priority if c in all_cols] + [c for c in all_cols if c not in priority]
        default_selected = self.get_default_selected_columns()
        self._column_default_selected = default_selected
        iva_columns = {"iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros"}

        for col in ordered_cols:
            if col in self.column_vars:
                continue
            is_selected = col in default_selected
            if col in iva_columns and col in self.df_full.columns:
                col_values = self.df_full[col].astype(str).str.strip().str.lower()
                has_values = col_values[~col_values.isin({"", "0", "0,00", "0.00", "nan", "none", "null"})].any()
                is_selected = bool(has_values)
            self.column_vars[col] = ctk.BooleanVar(value=is_selected)

        mode = self.column_mode_var.get()
        if mode == "simple":
            visible_cols = [c for c in ordered_cols if c in priority]
        else:
            visible_cols = ordered_cols

        sectioned: dict[str, list[str]] = {
            "Operativas": [],
            "Impuestos y montos": [],
            "Estado Hacienda": [],
            "XML técnico": [],
        }
        for col in visible_cols:
            sectioned[self.get_column_section(col)].append(col)

        for section_name, cols in sectioned.items():
            if not cols:
                continue
            is_collapsed = self.section_collapsed.get(section_name, section_name == "XML técnico" and mode == "avanzado")
            self.section_collapsed[section_name] = is_collapsed

            section_frame = ctk.CTkFrame(self.columns_scroll)
            section_frame.pack(fill="x", expand=False, padx=2, pady=3)

            header_text = f"{section_name} ({len(cols)}) {'▸' if is_collapsed else '▾'}"
            header_btn = ctk.CTkButton(
                section_frame,
                text=header_text,
                height=28,
                anchor="w",
                command=lambda name=section_name: self.toggle_section(name),
            )
            header_btn.pack(fill="x", padx=4, pady=(4, 2))

            if is_collapsed:
                continue

            items_frame = ctk.CTkFrame(section_frame, fg_color="transparent")
            items_frame.pack(fill="x", padx=4, pady=(0, 4))

            for col in cols:
                alias = self.get_column_alias(col)
                checkbox = ctk.CTkCheckBox(
                    items_frame,
                    text=self.truncate_label(alias, 40),
                    variable=self.column_vars[col],
                    command=self.refresh_table,
                )
                checkbox.pack(anchor="w", padx=4, pady=2)
                self.attach_tooltip(checkbox, f"{alias}\n\nCampo técnico: {col}")

    def get_priority_columns(self) -> list[str]:
        return [
            "archivo",
            "clave_numerica",
            "tipo_documento",
            "fecha_emision",
            "consecutivo",
            "emisor_nombre",
            "emisor_cedula",
            "receptor_nombre",
            "receptor_cedula",
            "moneda",
            "tipo_cambio",
            "subtotal",
            "iva_1",
            "iva_2",
            "iva_4",
            "iva_8",
            "iva_13",
            "iva_otros",
            "impuesto_total",
            "total_comprobante",
            "estado_hacienda",
            "detalle_estado_hacienda",
        ]

    def get_default_selected_columns(self) -> set[str]:
        return {
            "tipo_documento",
            "fecha_emision",
            "consecutivo",
            "emisor_nombre",
            "emisor_cedula",
            "moneda",
            "tipo_cambio",
            "subtotal",
            "impuesto_total",
            "total_comprobante",
            "estado_hacienda",
        }

    def get_column_alias(self, column_name: str) -> str:
        aliases = {
            "archivo": "Archivo",
            "clave_numerica": "Clave",
            "tipo_documento": "Tipo documento",
            "fecha_emision": "Fecha emisión",
            "consecutivo": "Consecutivo",
            "emisor_nombre": "Emisor",
            "emisor_cedula": "Cédula emisor",
            "receptor_nombre": "Receptor",
            "receptor_cedula": "Cédula receptor",
            "moneda": "Moneda",
            "tipo_cambio": "Tipo cambio",
            "subtotal": "Subtotal",
            "iva_1": "IVA 1%",
            "iva_2": "IVA 2%",
            "iva_4": "IVA 4%",
            "iva_8": "IVA 8%",
            "iva_13": "IVA 13%",
            "iva_otros": "Otros impuestos",
            "impuesto_total": "Impuesto total",
            "total_comprobante": "Total comprobante",
            "estado_hacienda": "Estado Hacienda",
            "detalle_estado_hacienda": "Detalle estado Hacienda",
        }
        if column_name in aliases:
            return aliases[column_name]

        if "_" in column_name and column_name.count("_") > 2:
            parts = column_name.split("_")
            if len(parts) >= 3:
                return f"{parts[0]} · {parts[1]} · {' '.join(parts[2:])}"

        return column_name.replace("_", " ").strip().title()

    def get_column_section(self, column_name: str) -> str:
        if column_name in {"estado_hacienda", "detalle_estado_hacienda", "estado_hacienda_xml", "detalle_estado_hacienda_xml"}:
            return "Estado Hacienda"
        if column_name in {"subtotal", "tipo_cambio", "iva_1", "iva_2", "iva_4", "iva_8", "iva_13", "iva_otros", "impuesto_total", "total_comprobante", "moneda"}:
            return "Impuestos y montos"
        if column_name in {"archivo", "clave_numerica", "tipo_documento", "fecha_emision", "consecutivo", "emisor_nombre", "emisor_cedula", "receptor_nombre", "receptor_cedula", "cliente_nombre", "cliente_cedula"}:
            return "Operativas"
        return "XML técnico"

    @staticmethod
    def truncate_label(value: str, max_len: int = 40) -> str:
        return value if len(value) <= max_len else f"{value[:max_len - 1]}…"

    def attach_tooltip(self, widget: Any, text: str) -> None:
        """Adjunta tooltip estable (único y con retardo) para evitar parpadeos."""

        def hide_tooltip() -> None:
            if self._tooltip_after_id is not None:
                try:
                    self.after_cancel(self._tooltip_after_id)
                except ValueError:
                    pass
                self._tooltip_after_id = None
            if self._tooltip_window is not None:
                self._tooltip_window.destroy()
                self._tooltip_window = None

        def show_tooltip() -> None:
            hide_tooltip()
            tooltip = ctk.CTkToplevel(self)
            tooltip.wm_overrideredirect(True)
            tooltip.attributes("-topmost", True)
            label = ctk.CTkLabel(
                tooltip,
                text=text,
                justify="left",
                padx=8,
                pady=6,
                fg_color=("#1F2937", "#111827"),
                corner_radius=6,
            )
            label.pack()
            x = widget.winfo_pointerx() + 12
            y = widget.winfo_pointery() + 12
            screen_w = self.winfo_screenwidth()
            screen_h = self.winfo_screenheight()
            tooltip.update_idletasks()
            tw = tooltip.winfo_width()
            th = tooltip.winfo_height()
            x = min(x, screen_w - tw - 8)
            y = min(y, screen_h - th - 8)
            tooltip.geometry(f"+{max(0, x)}+{max(0, y)}")
            self._tooltip_window = tooltip

        def on_enter(_event: Any) -> None:
            hide_tooltip()
            self._tooltip_after_id = self.after(450, show_tooltip)

        def on_leave(_event: Any) -> None:
            hide_tooltip()

        widget.bind("<Enter>", on_enter, add="+")
        widget.bind("<Leave>", on_leave, add="+")
        widget.bind("<ButtonPress>", on_leave, add="+")
        widget.bind("<Destroy>", on_leave, add="+")

    def toggle_section(self, section_name: str) -> None:
        self.section_collapsed[section_name] = not self.section_collapsed.get(section_name, False)
        self.refresh_column_selector()

    def select_basic_columns(self) -> None:
        defaults = self.get_default_selected_columns()
        for col, var in self.column_vars.items():
            var.set(col in defaults)
        self.refresh_table()

    def select_all_visible_columns(self) -> None:
        mode = self.column_mode_var.get()
        visible_reference = self.get_priority_columns() if mode == "simple" else list(self.df_full.columns)
        visible_set = set(visible_reference)
        for col, var in self.column_vars.items():
            if col in visible_set:
                var.set(True)
        self.refresh_table()

    def clear_visible_columns(self) -> None:
        mode = self.column_mode_var.get()
        visible_reference = self.get_priority_columns() if mode == "simple" else list(self.df_full.columns)
        visible_set = set(visible_reference)
        for col, var in self.column_vars.items():
            if col in visible_set:
                var.set(False)
        self.refresh_table()

    def save_column_template(self) -> None:
        template_name = simpledialog.askstring("Guardar plantilla", "Nombre de la plantilla:", parent=self)
        if not template_name:
            return
        data: dict[str, Any] = {}
        if self.template_path.exists():
            try:
                data = json.loads(self.template_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                LOGGER.warning("Plantillas dañadas. Se recreará el archivo de plantillas.")
                data = {}
        data[template_name] = [col for col, var in self.column_vars.items() if var.get()]
        self.template_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        self.status_label.configure(text=f"Plantilla guardada: {template_name}")

    def load_column_template(self) -> None:
        if not self.template_path.exists():
            messagebox.showinfo("Plantillas", "No hay plantillas guardadas todavía.")
            return
        try:
            data = json.loads(self.template_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            messagebox.showerror("Plantillas", "El archivo de plantillas está dañado.")
            return

        if not data:
            messagebox.showinfo("Plantillas", "No hay plantillas guardadas todavía.")
            return

        names = sorted(data.keys())
        selected = simpledialog.askstring(
            "Cargar plantilla",
            "Plantillas disponibles\n- " + "\n- ".join(names) + "\n\nEscribe el nombre exacto:",
            parent=self,
        )
        if not selected or selected not in data:
            return

        selected_cols = set(data[selected])
        for col, var in self.column_vars.items():
            var.set(col in selected_cols)
        self.refresh_table()
        self.status_label.configure(text=f"Plantilla cargada: {selected}")

    @staticmethod
    def format_amount_es(value: float) -> str:
        """Formato 10 000,00 para texto de resumen."""
        try:
            text = f"{float(value):,.2f}"
        except (TypeError, ValueError):
            return "0,00"
        return text.replace(",", " ").replace(".", ",")

    @staticmethod
    def short_company_name(name: str, max_len: int = 42) -> str:
        base = str(name or "").strip().upper()
        replacements = {
            "SOCIEDAD ANONIMA": "S. A.",
            "SOCIEDAD ANÓNIMA": "S. A.",
            "SOCIEDAD DE RESPONSABILIDAD LIMITADA": "S. R. L.",
            "SOCIEDAD RESPONSABILIDAD LIMITADA": "S. R. L.",
            "COMPANIA LIMITADA": "LTDA.",
            "COMPAÑIA LIMITADA": "LTDA.",
            "LIMITADA": "LTDA.",
        }
        for long_form, short_form in replacements.items():
            base = base.replace(long_form, short_form)
        base = " ".join(base.split())
        return base[:max_len].rstrip()

    @staticmethod
    def month_name_es(dt: datetime) -> str:
        months = {
            1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
            7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
        }
        return months.get(dt.month, "mes")

    def build_default_export_filename(self, customer_name: str) -> str:
        date_from = self.parse_filter_date(self.search_date_from_var.get().strip())
        date_to = self.parse_filter_date(self.search_date_to_var.get().strip())
        base_date = date_from or date_to or datetime.now()
        year = base_date.strftime("%Y")
        month_text = self.month_name_es(base_date)
        short_name = self.short_company_name(customer_name)
        return f"PF-{year} - {short_name} - REPORTE - {month_text}.xlsx"

    def filter_by_date_range(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aplica solo filtro por rango de fechas DD/MM/AAAA."""
        if df.empty:
            return df

        date_from = self.parse_filter_date(self.search_date_from_var.get().strip())
        date_to = self.parse_filter_date(self.search_date_to_var.get().strip())
        if not date_from and not date_to:
            return df

        filtered = df
        date_series = pd.to_datetime(filtered["fecha_emision"], format="%d/%m/%Y", errors="coerce")
        if date_from:
            filtered = filtered[date_series >= pd.Timestamp(date_from)]
            date_series = pd.to_datetime(filtered["fecha_emision"], format="%d/%m/%Y", errors="coerce")
        if date_to:
            filtered = filtered[date_series <= pd.Timestamp(date_to)]
        return filtered

    def schedule_apply_filters(self) -> None:
        """Debounce de filtros para evitar recalcular por cada tecla inmediatamente."""
        if self._filter_after_id is not None:
            try:
                self.after_cancel(self._filter_after_id)
            except ValueError:
                pass
        self._filter_after_id = self.after(300, self.apply_filters)

    def apply_filters(self) -> None:
        if self.df_full.empty:
            return
        self._filter_after_id = None

        tax_id_text = self.search_tax_id_var.get().strip().lower()
        filtered = self.df_full
        if tax_id_text:
            filtered = filtered[
                filtered["_emisor_cedula_norm"].astype(str).str.contains(tax_id_text, na=False, regex=False)
                | filtered["_receptor_cedula_norm"].astype(str).str.contains(tax_id_text, na=False, regex=False)
            ]

        date_from = self.parse_filter_date(self.search_date_from_var.get().strip())
        date_to = self.parse_filter_date(self.search_date_to_var.get().strip())
        if date_from or date_to:
            date_series = filtered["_fecha_emision_dt"]
            if date_from:
                filtered = filtered[date_series >= pd.Timestamp(date_from)]
                date_series = filtered["_fecha_emision_dt"]
            if date_to:
                filtered = filtered[date_series <= pd.Timestamp(date_to)]

        self.df_filtered = filtered
        self.current_page = 0
        self.refresh_table()
    @staticmethod
    def parse_filter_date(value: str) -> datetime | None:
        if not value:
            return None
        try:
            return datetime.strptime(value, "%d/%m/%Y")
        except ValueError:
            return None

    def copy_to_clipboard(self, text: str) -> None:
        """Copia texto al portapapeles y muestra confirmación."""
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.update_idletasks()
            messagebox.showinfo("Portapapeles", "Ruta copiada al portapapeles.")
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("Error", f"No se pudo copiar al portapapeles: {exc}")

    def export_audit_log_json(self, report: dict[str, Any]) -> None:
        """Exporta reporte como JSON con timestamp en nombre."""
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        documents_dir = Path.home() / "Documents"
        initial_dir = documents_dir if documents_dir.exists() else Path.home()
        target = filedialog.asksaveasfilename(
            title="Exportar log de auditoría (JSON)",
            defaultextension=".json",
            initialdir=str(initial_dir),
            initialfile=f"audit_log_{stamp}.json",
            confirmoverwrite=True,
            filetypes=[("JSON", "*.json")],
        )
        if not target:
            return
        try:
            Path(target).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
            messagebox.showinfo("Exportación completada", f"Log JSON guardado en:\n{target}")
        except OSError as exc:
            messagebox.showerror("Error", f"No se pudo exportar JSON: {exc}")

    def export_audit_log_txt(self, report: dict[str, Any]) -> None:
        """Exporta reporte como TXT legible con secciones claras."""
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        documents_dir = Path.home() / "Documents"
        initial_dir = documents_dir if documents_dir.exists() else Path.home()
        target = filedialog.asksaveasfilename(
            title="Exportar log de auditoría (TXT)",
            defaultextension=".txt",
            initialdir=str(initial_dir),
            initialfile=f"audit_log_{stamp}.txt",
            confirmoverwrite=True,
            filetypes=[("Texto", "*.txt")],
        )
        if not target:
            return

        total = int(report.get("total_files_found", 0))
        processed = int(report.get("successfully_processed", 0))
        failed_files = report.get("failed_files", []) or []
        duplicate_files = report.get("duplicate_files", []) or []
        elapsed = float(report.get("processing_time_seconds", 0.0))
        success_pct = (processed / total * 100.0) if total else 0.0

        lines = [
            "REPORTE DE AUDITORÍA - FACTURACIÓN",
            "=" * 46,
            "",
            "RESUMEN",
            "-" * 46,
            f"Total encontrados: {total}",
            f"Procesados OK: {processed}",
            f"Fallidos: {len(failed_files)}",
            f"Duplicados: {len(duplicate_files)}",
            f"Tiempo total: {elapsed:.3f} s",
            f"Porcentaje de éxito: {success_pct:.2f}%",
            "",
            "DISTRIBUCIÓN POR TIPO DE DOCUMENTO",
            "-" * 46,
        ]

        files_by_type = report.get("files_by_type", {}) or {}
        ordered_types = sorted(files_by_type.items(), key=lambda item: item[1], reverse=True)
        if ordered_types:
            lines.extend([f"- {doc_type}: {count}" for doc_type, count in ordered_types])
        else:
            lines.append("- Sin datos")

        lines.extend(["", "ARCHIVOS FALLIDOS", "-" * 46])
        if failed_files:
            for idx, item in enumerate(failed_files, start=1):
                lines.append(f"{idx}. {item.get('archivo', 'Sin nombre')}")
                lines.append(f"   Ruta: {item.get('ruta', '')}")
                lines.append(f"   Error: {item.get('error', 'Error desconocido')}")
        else:
            lines.append("- Sin archivos fallidos")

        lines.extend(["", "ARCHIVOS DUPLICADOS", "-" * 46])
        if duplicate_files:
            for idx, item in enumerate(duplicate_files, start=1):
                lines.append(
                    f"{idx}. {item.get('archivo', 'Sin nombre')} -> Duplicado de: {item.get('original', 'Desconocido')}"
                )
                lines.append(f"   Ruta duplicado: {item.get('ruta', '')}")
                lines.append(f"   Ruta original: {item.get('original_ruta', '')}")
                lines.append(f"   Hash: {item.get('hash', '')}")
        else:
            lines.append("- Sin duplicados")

        try:
            Path(target).write_text("\n".join(lines), encoding="utf-8")
            messagebox.showinfo("Exportación completada", f"Log TXT guardado en:\n{target}")
        except OSError as exc:
            messagebox.showerror("Error", f"No se pudo exportar TXT: {exc}")

    def show_audit_report_window(self) -> None:
        """Muestra una vista amigable del reporte de auditoría del último procesamiento."""
        report = self.last_audit_report
        if not report:
            messagebox.showinfo("Reporte de auditoría", "No hay reporte disponible todavía.")
            return

        summary = report.get("files_by_status", {}) or {}
        total = int(report.get("total_files_found", 0))
        processed = int(summary.get("ok", report.get("successfully_processed", 0)))
        failed_files = report.get("failed_files", []) or []
        duplicate_files = report.get("duplicate_files", []) or []
        failed = int(summary.get("failed", len(failed_files)))
        duplicates = int(summary.get("duplicates", len(duplicate_files)))
        elapsed = float(report.get("processing_time_seconds", 0.0))
        success_pct = (processed / total * 100.0) if total else 0.0
        success_color = "#10B981" if success_pct >= 95 else "#F59E0B"

        top = ctk.CTkToplevel(self)
        top.title("Reporte de auditoría")
        top.geometry("1100x760")
        top.minsize(980, 680)
        top.transient(self)
        top.grab_set()

        container = ctk.CTkFrame(top)
        container.pack(fill="both", expand=True, padx=12, pady=12)

        header = ctk.CTkLabel(
            container,
            text="Resumen de Auditoría de XML",
            font=ctk.CTkFont(size=20, weight="bold"),
            anchor="w",
        )
        header.pack(fill="x", padx=10, pady=(10, 8))

        cards_frame = ctk.CTkFrame(container)
        cards_frame.pack(fill="x", padx=10, pady=(0, 10))
        for col in range(3):
            cards_frame.grid_columnconfigure(col, weight=1)

        metrics = [
            ("📁 Total encontrados", str(total), "#3B82F6"),
            ("✓ Procesados", str(processed), "#10B981"),
            ("✗ Fallidos", str(failed), "#EF4444"),
            ("≈ Duplicados", str(duplicates), "#F59E0B"),
            ("⏱ Tiempo", f"{elapsed:.2f}s", "#8B5CF6"),
            ("% Éxito", f"{success_pct:.1f}%", success_color),
        ]

        for idx, (title, value, color) in enumerate(metrics):
            row = idx // 3
            col = idx % 3
            card = ctk.CTkFrame(cards_frame)
            card.grid(row=row, column=col, padx=6, pady=6, sticky="nsew")
            ctk.CTkLabel(card, text=title, text_color=color, font=ctk.CTkFont(size=14, weight="bold")).pack(
                anchor="w", padx=12, pady=(10, 2)
            )
            ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=22, weight="bold")).pack(
                anchor="w", padx=12, pady=(0, 10)
            )

        body = ctk.CTkScrollableFrame(container)
        body.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        ctk.CTkLabel(body, text="Distribución por tipo", font=ctk.CTkFont(size=16, weight="bold"), anchor="w").pack(
            fill="x", pady=(4, 4)
        )
        files_by_type = report.get("files_by_type", {}) or {}
        ordered_types = sorted(files_by_type.items(), key=lambda item: item[1], reverse=True)
        if ordered_types:
            for doc_type, count in ordered_types:
                ctk.CTkLabel(body, text=f"• {doc_type}: {count}", anchor="w", justify="left").pack(
                    fill="x", padx=8, pady=1
                )
        else:
            ctk.CTkLabel(body, text="• Sin datos de tipos", anchor="w").pack(fill="x", padx=8, pady=1)

        if failed_files:
            ctk.CTkLabel(body, text="Archivos fallidos", font=ctk.CTkFont(size=16, weight="bold"), anchor="w").pack(
                fill="x", pady=(12, 4)
            )
            failed_frame = ctk.CTkScrollableFrame(body, height=180)
            failed_frame.pack(fill="x", padx=4, pady=(0, 4))
            for idx, item in enumerate(failed_files, start=1):
                row = ctk.CTkFrame(failed_frame)
                row.pack(fill="x", padx=4, pady=4)
                file_name = str(item.get("archivo", "Sin nombre"))
                file_path = str(item.get("ruta", ""))
                error_text = str(item.get("error", "Error desconocido")).replace("\n", " ").strip()
                if len(error_text) > 100:
                    error_text = f"{error_text[:100]}..."
                ctk.CTkLabel(
                    row,
                    text=f"{idx}. {file_name}\nError: {error_text}",
                    anchor="w",
                    justify="left",
                ).pack(side="left", fill="x", expand=True, padx=10, pady=8)
                ctk.CTkButton(
                    row,
                    text="📋 Copiar ruta",
                    width=120,
                    command=lambda p=file_path: self.copy_to_clipboard(p),
                ).pack(side="right", padx=10, pady=8)

        if duplicate_files:
            ctk.CTkLabel(body, text="Duplicados detectados", font=ctk.CTkFont(size=16, weight="bold"), anchor="w").pack(
                fill="x", pady=(12, 4)
            )
            dup_frame = ctk.CTkFrame(body)
            dup_frame.pack(fill="x", padx=4, pady=(0, 6))
            for item in duplicate_files:
                duplicate_name = str(item.get("archivo", "Sin nombre"))
                original_name = str(item.get("original", "Desconocido"))
                ctk.CTkLabel(
                    dup_frame,
                    text=f"• {duplicate_name} → Duplicado de: {original_name}",
                    anchor="w",
                    justify="left",
                ).pack(fill="x", padx=8, pady=2)

        actions = ctk.CTkFrame(container)
        actions.pack(fill="x", padx=10, pady=(0, 10))
        ctk.CTkButton(
            actions,
            text="📥 Exportar Log Completo (JSON)",
            command=lambda: self.export_audit_log_json(report),
        ).pack(side="left", padx=(8, 6), pady=8)
        ctk.CTkButton(
            actions,
            text="📄 Exportar Log Legible (TXT)",
            command=lambda: self.export_audit_log_txt(report),
        ).pack(side="left", padx=6, pady=8)
        ctk.CTkButton(actions, text="Cerrar", command=top.destroy).pack(side="right", padx=(6, 8), pady=8)

    def export_excel(self) -> None:
        if self.df_full.empty:
            messagebox.showinfo("Sin datos", "Primero carga datos para exportar.")
            return

        tax_id_raw = self.search_tax_id_var.get().strip()
        tax_id = self.manager.normalize_identification(tax_id_raw)
        documents_dir = Path.home() / "Documents"
        initial_dir = documents_dir if documents_dir.exists() else Path.home()

        source_df = self.df_filtered if not self.df_filtered.empty else self.df_full
        source_df_for_tax_split = self.df_full if not self.df_full.empty else source_df

        if tax_id:
            # Para clasificación Ventas/Receptor/ORS/Sin receptor se usa el universo completo
            # acotado únicamente por rango de fechas, no por el filtro de texto de cédula.
            export_df = self.filter_by_date_range(source_df_for_tax_split)
        else:
            date_filtered_df = self.filter_by_date_range(source_df)
            has_active_filters = bool(self.search_date_from_var.get().strip() or self.search_date_to_var.get().strip())
            export_df = date_filtered_df if has_active_filters else source_df

        if export_df.empty:
            messagebox.showinfo("Sin datos", "No hay datos para exportar con los filtros actuales.")
            return

        selected_columns = [col for col, var in self.column_vars.items() if var.get()]
        if not selected_columns:
            selected_columns = [col for col in export_df.columns if not str(col).startswith("_")]

        report_owner_name = ""
        if tax_id_raw:
            report_owner_name = self.manager.resolve_party_name(tax_id_raw, "")
        if not report_owner_name and "emisor_nombre" in export_df.columns and not export_df.empty:
            report_owner_name = str(export_df["emisor_nombre"].iloc[0] or "")
        if not report_owner_name:
            report_owner_name = "REPORTE DE COMPROBANTES"

        default_filename = self.build_default_export_filename(report_owner_name)
        save_options = {
            "title": "Guardar reporte Excel",
            "defaultextension": ".xlsx",
            "initialfile": default_filename,
            "initialdir": str(initial_dir),
            "confirmoverwrite": True,
            "filetypes": [("Excel", "*.xlsx")],
        }
        save_path = filedialog.asksaveasfilename(**save_options)

        base_export = export_df.copy()
        to_export = export_df[selected_columns].copy()

        numeric_columns = [
            "subtotal",
            "tipo_cambio",
            "iva_1",
            "iva_2",
            "iva_4",
            "iva_8",
            "iva_13",
            "iva_otros",
            "impuesto_total",
            "total_comprobante",
        ]
        text_columns = {
            "clave_numerica",
            "consecutivo",
            "emisor_cedula",
            "receptor_cedula",
            "cliente_cedula",
        }

        for col in text_columns:
            if col in to_export.columns:
                to_export[col] = to_export[col].fillna("").astype(str).str.strip()

        for col in numeric_columns:
            if col in to_export.columns:
                to_export[col] = to_export[col].apply(self.manager.parse_decimal_value)

        date_column = "fecha_emision"
        if date_column in to_export.columns:
            to_export[date_column] = pd.to_datetime(
                to_export[date_column],
                format="%d/%m/%Y",
                errors="coerce",
            )

        if tax_id and {"emisor_cedula", "receptor_cedula"}.issubset(base_export.columns):
            emisor_raw = base_export["emisor_cedula"].fillna("").astype(str).str.strip()
            receptor_raw = base_export["receptor_cedula"].fillna("").astype(str).str.strip()

            emisor = emisor_raw.map(self.manager.normalize_identification)
            receptor = receptor_raw.map(self.manager.normalize_identification)
            receptor_is_empty = receptor_raw.str.lower().isin({"", "null", "none", "nan"}) | receptor.eq("")

            # Jerarquía estricta:
            # 1) Ventas
            mask_ventas = emisor.eq(tax_id)
            # 2) Receptor
            mask_receptor = ~mask_ventas & receptor.eq(tax_id)
            # 3) Sin receptor
            mask_sin_receptor = ~mask_ventas & ~mask_receptor & receptor_is_empty
            # 4) ORS (resto)
            mask_ors = ~mask_ventas & ~mask_receptor & ~mask_sin_receptor

            sheet_map = {
                "ventas": to_export.loc[mask_ventas],
                "receptor": to_export.loc[mask_receptor],
                "sin receptor": to_export.loc[mask_sin_receptor],
                "ORS": to_export.loc[mask_ors],
            }
        else:
            sheet_map = {"Reporte": to_export}

        pretty_headers = {
            "clave_numerica": "Clave",
            "tipo_documento": "Tipo documento",
            "fecha_emision": "Fecha emisión",
            "consecutivo": "Consecutivo",
            "emisor_nombre": "Emisor",
            "emisor_cedula": "Cédula emisor",
            "receptor_nombre": "Receptor",
            "receptor_cedula": "Cédula receptor",
            "moneda": "Moneda",
            "tipo_cambio": "Tipo cambio",
            "subtotal": "Subtotal",
            "iva_1": "IVA 1%",
            "iva_2": "IVA 2%",
            "iva_4": "IVA 4%",
            "iva_8": "IVA 8%",
            "iva_13": "IVA 13%",
            "iva_otros": "Otros impuestos",
            "impuesto_total": "Impuesto total",
            "total_comprobante": "Total comprobante",
            "estado_hacienda": "Estado Hacienda",
            "detalle_estado_hacienda": "Detalle estado Hacienda",
        }

        date_from_label = self.search_date_from_var.get().strip() or "01/01/1900"
        date_to_label = self.search_date_to_var.get().strip() or datetime.now().strftime("%d/%m/%Y")

        title_fill = PatternFill(fill_type="solid", fgColor="0B2B66")
        subtitle_fill = PatternFill(fill_type="solid", fgColor="7F7F7F")
        summary_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
        header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        credit_fill = PatternFill(fill_type="solid", fgColor="DAF2D0")
        title_font = Font(bold=True, color="FFFFFF", size=22)
        subtitle_font = Font(bold=True, color="FFFFFF", size=14)
        summary_font = Font(bold=False, color="111111", size=12)
        header_font = Font(bold=True)

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheet_map.items():
                display_df = sheet_df.rename(columns={col: pretty_headers.get(col, col.replace("_", " ").title()) for col in sheet_df.columns})
                display_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=4)
                ws = writer.sheets[sheet_name]

                max_col = ws.max_column if ws.max_column > 0 else 1
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = report_owner_name.upper()
                title_cell.font = title_font
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = title_fill

                subtitle_cell = ws.cell(row=2, column=1)
                subtitle_cell.value = f"REPORTE DE {sheet_name.upper()} - Período: {date_from_label} al {date_to_label}"
                subtitle_cell.font = subtitle_font
                subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
                subtitle_cell.fill = subtitle_fill

                monto_total = Decimal("0")
                if "total_comprobante" in sheet_df.columns:
                    valid_amounts = [value for value in sheet_df["total_comprobante"].tolist() if isinstance(value, Decimal)]
                    if valid_amounts:
                        monto_total = sum(valid_amounts, Decimal("0"))
                if "moneda" in sheet_df.columns and not sheet_df.empty:
                    monedas = sorted({str(m).strip() for m in sheet_df["moneda"].dropna().tolist() if str(m).strip()})
                    if not monedas:
                        moneda_value = "N/A"
                    elif len(monedas) == 1:
                        moneda_value = monedas[0]
                    else:
                        moneda_value = "MIXTA: " + ", ".join(monedas)
                else:
                    moneda_value = "N/A"
                generated = datetime.now().strftime("%d/%m/%Y %H:%M")
                summary_cell = ws.cell(row=3, column=1)
                summary_cell.value = (
                    f"Total filas: {len(sheet_df)}   |   Monto Total: {self.format_amount_es(monto_total)}   |   "
                    f"Moneda: {moneda_value}   |   Generado: {generated}"
                )
                summary_cell.font = summary_font
                summary_cell.alignment = Alignment(horizontal="center", vertical="center")
                summary_cell.fill = summary_fill

                header_row = 5
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=header_row, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                tipo_idx = list(sheet_df.columns).index("tipo_documento") + 1 if "tipo_documento" in sheet_df.columns else None

                for idx, col in enumerate(sheet_df.columns, start=1):
                    col_name = str(col)
                    for row in range(header_row + 1, len(sheet_df) + header_row + 1):
                        cell = ws.cell(row=row, column=idx)
                        if col_name in text_columns:
                            cell.number_format = "@"
                            cell.value = "" if cell.value is None else str(cell.value)
                        elif col_name == date_column and cell.value is not None:
                            cell.number_format = "dd/mm/yyyy"
                        elif col_name in numeric_columns and cell.value is not None:
                            cell.number_format = "#,##0.00"
                            if isinstance(cell.value, Decimal):
                                cell.value = float(cell.value)

                if tipo_idx is not None:
                    for row in range(header_row + 1, len(sheet_df) + header_row + 1):
                        if ws.cell(row=row, column=tipo_idx).value == "Nota de Crédito":
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = credit_fill

                for col_idx in range(1, ws.max_column + 1):
                    max_len = 0
                    for row in range(header_row, ws.max_row + 1):
                        value = ws.cell(row=row, column=col_idx).value
                        if value is None:
                            continue
                        max_len = max(max_len, len(str(value)))
                    ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = min(max(max_len + 3, 12), 65)

                ws.freeze_panes = ws["A6"]

        messagebox.showinfo("Exportación completada", f"Reporte exportado en:\n{save_path}")

    def prev_page(self) -> None:
        if self.current_page > 0:
            self.current_page -= 1
            self.refresh_table()
    def next_page(self) -> None:
        if self.df_filtered.empty:
            return
        total_pages = max((len(self.df_filtered) - 1) // self.PAGE_SIZE + 1, 1)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self.refresh_table()
    def refresh_table(self) -> None:
        if self.df_filtered.empty and self.df_full.empty:
            return
        df = self.df_filtered if not self.df_filtered.empty else self.df_full
        selected_columns = [col for col, var in self.column_vars.items() if var.get()]
        if not selected_columns:
            selected_columns = ["archivo"] if "archivo" in df.columns else list(df.columns[:1])
        start = self.current_page * self.PAGE_SIZE
        end = start + self.PAGE_SIZE
        page_df = df.iloc[start:end]
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = selected_columns
        for col in selected_columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=170, minwidth=70, stretch=True)
        for row in page_df[selected_columns].astype(str).to_numpy().tolist():
            self.tree.insert("", "end", values=row)
        total_rows = len(df)
        total_pages = max((total_rows - 1) // self.PAGE_SIZE + 1, 1)
        if self.current_page >= total_pages:
            self.current_page = max(total_pages - 1, 0)
        self.page_label.configure(text=f"Página {self.current_page + 1}/{total_pages} · {total_rows:,} filas")

def main() -> None:
    app = MainApp()
    app.mainloop()
if __name__ == "__main__":
    main()
